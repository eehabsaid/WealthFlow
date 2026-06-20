"""
Excel export — exact replica of original Balance.xlsx style, populated from live DB.
Every font, border, number format, column width, row height, merge, and freeze pane
is matched to the original file as inspected cell-by-cell.
"""

import io
from itertools import groupby
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# ── Exact formats from original ───────────────────────────────────────────────
FMT_EGP = "[$ج.م.\u200f-C01]\\ #,##0.00_-"
FMT_EGP_RED = "[$ج.م.\u200f-C01]\\ #,##0.00;[Red][$ج.م.\u200f-C01]\\ #,##0.00"
FMT_USD = '"$"#,##0.00;[Red]"$"#,##0.00'
FMT_EUR = "[$EUR]\\ #,##0.00;[Red][$EUR]\\ #,##0.00"
FMT_SAR = "[$SAR]\\ #,##0.00;[Red][$SAR]\\ #,##0.00"
FMT_GOLD = '0\\ "Grams"'
FMT_EGP_CERT = "[$EGP]\\ #,##0.00"
FMT_EGP_CERT_R = "[$EGP]\\ #,##0.00;[Red][$EGP]\\ #,##0.00"
FMT_PCT = "0.00%"
FMT_DATE = "[$-F800]dddd/\\ mmmm\\ dd/\\ yyyy"
FMT_INT = "0"

GREY = "FF7F7F7F"
RED = "FFFF0000"
WHITE = "FFFFFFFF"
EXP_BG = "FF203864"
EXP_MONTH_BG = "FFD9E1F2"
EXP_YEAR_BG = "FFBDD7EE"


def auto_adjust_columns(ws):
    """
    Adjusts column widths based on content, using a larger buffer
    specifically for date columns.
    """
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        is_date_col = False

        for cell in col:
            # Check if cell contains a date/datetime object
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                is_date_col = True

            # Check if format is a date format
            # Many date formats contain 'y', 'm', or 'd'
            fmt = str(cell.number_format).lower()
            if any(x in fmt for x in ["yyyy", "mmmm", "dddd"]):
                is_date_col = True

            # Get length of the rendered string
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_length:
                max_length = len(val)

        # Apply specific buffer
        # Dates are naturally longer when formatted (e.g., "Monday, June 14, 2026")
        if is_date_col:
            ws.column_dimensions[column].width = max_length + 20
        else:
            # Standard buffer for other data
            ws.column_dimensions[column].width = max_length + 10


def _f(bold=False, size=11, color=None, name="Arial"):
    kw = dict(bold=bold, size=size, name=name)
    if color:
        kw["color"] = color
    return Font(**kw)


def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _thin_lr():
    s = Side(style="thin")
    return Border(left=s, right=s)


def _thin_tb():
    s = Side(style="thin")
    return Border(top=s, bottom=s)


def _thin_b():
    return Border(bottom=Side(style="thin"))


def _thin_t():
    return Border(top=Side(style="thin"))


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _align(h="general", v="bottom", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _center():
    return Alignment(horizontal="center", vertical="bottom")


# ── Salary sheet ──────────────────────────────────────────────────────────────

MONTH_ORDER = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
    "Quarter-Bonuses",
]


def _msort(m):
    m = str(m).strip()
    for i, n in enumerate(MONTH_ORDER):
        if n.lower().startswith(m.lower()[:3]):
            return i
    return 99


# Column widths per sheet (exact from original)
SALARY_COL_WIDTHS = {
    "NTG": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
    "Giza Systems": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
    "Giza Systems (2)": {
        "A": 13.7,
        "B": 16.1,
        "C": 16.0,
        "D": 19.3,
        "E": 15.6,
        "F": 14.3,
    },
    "ElSeweedy Technology": {
        "A": 13.7,
        "B": 16.1,
        "C": 16.0,
        "D": 19.3,
        "E": 15.6,
        "F": 14.3,
    },
    "Dedalus": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
    "Globemed": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
    "Giza Systems (3)": {
        "A": 13.7,
        "B": 16.1,
        "C": 16.0,
        "D": 19.3,
        "E": 15.6,
        "F": 14.3,
    },
}

# Row heights for structural rows
SALARY_ROW_HT = {
    1: 14.25,  # col header
    2: 14.25,  # title row (merged with row 3)
    3: 20.25,  # title row part 2
    4: 22.8,  # first year heading (always)
}
YEAR_ROW_HT = 22.8  # all other year heading rows

# Freeze pane per company (original)
SALARY_FREEZE = {
    "NTG": "A4",  # freeze header+title rows only, data scrolls
    "Giza Systems": "A4",
    "Giza Systems (2)": "A4",
    "ElSeweedy Technology": "A4",
    "Dedalus": "A4",
    "Globemed": "A4",
    "Giza Systems (3)": "A4",
}


def _apply_data_row(ws, row, has_bonus=False):
    cols = 6 if has_bonus else 5
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _f(name="Arial")
        cell.border = _thin()
        if c in (3, 4, 5):
            cell.number_format = FMT_EGP
        if has_bonus and c == 6:
            cell.number_format = FMT_EGP


def _apply_total_row(ws, row, has_bonus=False):
    cols = 6 if has_bonus else 5
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _f(bold=True, name="Arial")
        cell.border = _thin()
        cell.alignment = _center()
        if c in (3, 4):
            cell.number_format = FMT_EGP
        if c == 5:
            cell.number_format = FMT_EGP
        if has_bonus and c == 6:
            cell.number_format = FMT_EGP


def build_salary_sheet(ws, company, entries):
    """Exact replica of original salary sheet styling using theme RGB values."""
    name = company.name
    # has_bonus = (name == 'Giza Systems (3)')
    has_bonus = True
    cols = 6
    last_col = "F"

    # ── Theme RGB fills (from original Balance.xlsx theme1.xml) ──────────────
    FILL_WHITE = _fill("FFFFFFFF")  # theme 1 lt1  — Salary Details row
    FILL_DARK_BLUE = _fill("FF1F497D")  # theme 2 dk2  — headers, year hdg, total rows
    FILL_RED_DATA = _fill("FFC0504D")  # theme 5 acc2 — data rows
    FILL_BLACK = _fill("FF000000")  # theme 0 dk1  — SUMMARY/grand total row

    # ── Font colors ───────────────────────────────────────────────────────────
    GREY = "FF7F7F7F"  # header col labels
    RED_TTL = "FFFF0000"  # "Salary Details" title
    WHITE = "FFFFFFFF"  # bold on dark backgrounds
    CREAM = "FFEEECE1"  # year heading font (theme 3 lt2)

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 14.25
    ws.row_dimensions[2].height = 14.25
    ws.row_dimensions[3].height = 20.25

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = SALARY_COL_WIDTHS.get(
        name, {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3}
    )
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Freeze pane ───────────────────────────────────────────────────────────
    fp = SALARY_FREEZE.get(name)
    if fp:
        ws.freeze_panes = fp

    # ── Row 1: column headers ─────────────────────────────────────────────────
    # GS3: bold+italic, white text, dark blue fill
    # Others: not bold, italic, grey text, dark blue fill
    hdrs = [
        "Year",
        "Month",
        "Expected",
        "Paid (Salary + Bonus)" if has_bonus else "Paid",
        "Remaining",
    ]
    if has_bonus:
        hdrs.append("Bonus")

    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        if has_bonus:
            cell.font = Font(bold=True, italic=True, size=11, name="Arial")
        else:
            cell.font = Font(bold=False, italic=True, size=11, name="Arial", color=GREY)
        cell.fill = FILL_DARK_BLUE
        cell.alignment = _align("center")
        cell.border = _thin()

    # ── Rows 2-3: "Salary Details" merged, white bg, red bold text ───────────
    ws.merge_cells(f"A2:{last_col}3")
    c2 = ws.cell(row=2, column=1, value=" Salary Details")
    c2.font = Font(bold=True, size=18, name="Times New Roman", color=RED_TTL)
    c2.fill = FILL_BLACK
    c2.alignment = _align("center")
    if has_bonus:
        c2.border = Border(top=Side(style="thin"))
    else:
        c2.border = Border(bottom=Side(style="thin"))

    # ── Data ──────────────────────────────────────────────────────────────────
    row = 4
    total_rows = []
    sorted_entries = sorted(entries, key=lambda e: (e.year, _msort(str(e.month))))

    for year, ygrp in groupby(sorted_entries, key=lambda e: e.year):
        year_entries = list(ygrp)

        # Year heading: merged, bold sz18, Times NR, cream font, dark blue fill
        yr_row = row
        ws.row_dimensions[yr_row].height = YEAR_ROW_HT
        yr_merge = f"A{yr_row}:{last_col}{yr_row}"
        try:
            ws.merge_cells(yr_merge)
        except Exception:
            pass
        yc = ws.cell(row=yr_row, column=1, value=int(year))
        yc.font = Font(bold=True, size=18, name="Times New Roman")
        yc.fill = FILL_DARK_BLUE
        yc.alignment = _align("center")
        # NTG and multi-year sheets: top+bottom border; GS3: no border
        if has_bonus:
            yc.border = Border()
        else:
            yc.border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
        row += 1

        data_start = row
        for entry in year_entries:
            # Data rows: red (accent2) fill, theme color 0 (dark/auto) font
            for c in range(1, cols + 1):
                ws.cell(row=row, column=c).fill = FILL_RED_DATA
                ws.cell(row=row, column=c).font = Font(size=11, name="Arial")
                ws.cell(row=row, column=c).border = _thin()

            ws.cell(row=row, column=1, value=int(entry.year))
            ws.cell(row=row, column=2, value=str(entry.month))
            ws.cell(row=row, column=3, value=float(entry.expected))
            ws.cell(row=row, column=3).number_format = FMT_EGP
            ws.cell(row=row, column=4, value=float(entry.paid))
            ws.cell(row=row, column=4).number_format = FMT_EGP

            # Remaining formula per company
            if name in ("NTG", "Giza Systems", "Giza Systems (2)"):
                rem = f"=D{row}-C{row}"
            else:
                rem = f"=IF(C{row}>D{row},C{row}-D{row},0)"
            ws.cell(row=row, column=5, value=rem)
            ws.cell(row=row, column=5).number_format = FMT_EGP

            if has_bonus:
                bonus_val = float(getattr(entry, "bonus", 0) or 0)
                ws.cell(row=row, column=6, value=bonus_val)
                ws.cell(row=row, column=6).number_format = FMT_EGP
            row += 1

        data_end = row - 1

        # ── Annual Total row: dark blue fill, white bold font, center ─────────
        paid_count = sum(1 for e in year_entries if float(e.paid) > 0)

        for c in range(1, cols + 1):
            tc = ws.cell(row=row, column=c)
            tc.fill = FILL_DARK_BLUE
            tc.font = Font(bold=True, size=11, name="Arial")
            tc.alignment = _align("center")
            tc.border = _thin()

        ws.cell(row=row, column=1, value="Total")
        # B col: NTG first year plain count; GS3/GS/GS2+ use COUNTIF on D
        if name == "NTG" and len(total_rows) == 0:
            ws.cell(row=row, column=2, value=paid_count)
        else:
            ws.cell(
                row=row,
                column=2,
                value=f'=COUNTIF(D{data_start}:D{data_end}, "<> 0.00")',
            )
        ws.cell(row=row, column=3, value=f"=SUM(C{data_start}:C{data_end})")
        ws.cell(row=row, column=3).number_format = FMT_EGP
        ws.cell(row=row, column=4, value=f"=SUM(D{data_start}:D{data_end})")
        ws.cell(row=row, column=4).number_format = FMT_EGP

        # E col: NTG uses D-C; GS uses SUM(E:E); GS3 uses SUM(E:E)
        if name == "NTG":
            ws.cell(row=row, column=5, value=f"=D{row}-C{row}")
        else:
            ws.cell(row=row, column=5, value=f"=SUM(E{data_start}:E{data_end})")
        ws.cell(row=row, column=5).number_format = FMT_EGP

        if has_bonus:
            ws.cell(row=row, column=6, value=f"=SUM(F{data_start}:F{data_end})")
            ws.cell(row=row, column=6).number_format = FMT_EGP

        total_rows.append(row)
        row += 1

    # ── Grand summary / outer Total row ──────────────────────────────────────
    sr = row

    # Single-year companies (ElSeweedy, Dedalus, Globemed) have a blank merged
    # row then an outer Total row that mirrors the inner one

    # Multi-year sheets: SUMMARY / Total grand row — BLACK fill, white font
    label_map = {
        "NTG": "Total",
        "Giza Systems": "Total",
        "Giza Systems (2)": "Total",
        "Giza Systems (3)": "Total",
        "ElSeweedy Technology": "Total",
        "Dedalus": "Total",
        "Globemed": "Total",
    }
    label = label_map.get(name, "Total")

    for c in range(1, cols + 1):
        sc = ws.cell(row=sr, column=c)
        sc.fill = FILL_BLACK
        sc.font = Font(bold=True, size=11, name="Arial", color="FFFFFFFF")
        sc.border = _thin()

    ws.cell(row=sr, column=1, value=label)
    b_ref = "+".join(f"B{r}" for r in total_rows)
    ws.cell(row=sr, column=2, value=f"={b_ref}")
    ws.cell(row=sr, column=2).number_format = "0" if name != "NTG" else "General"
    c_ref = "+".join(f"C{r}" for r in total_rows)
    d_ref = "+".join(f"D{r}" for r in total_rows)
    ws.cell(row=sr, column=3, value=f"={c_ref}")
    ws.cell(row=sr, column=3).number_format = FMT_EGP
    ws.cell(row=sr, column=4, value=f"={d_ref}")
    ws.cell(row=sr, column=4).number_format = FMT_EGP
    ws.cell(row=sr, column=5, value=f"=D{sr}-C{sr}")
    ws.cell(row=sr, column=5).number_format = FMT_EGP

    if has_bonus:
        f_ref = "+".join(f"F{r}" for r in total_rows)
        ws.cell(row=sr, column=6, value=f"={f_ref}")
        ws.cell(row=sr, column=6).number_format = FMT_EGP

    ws.row_dimensions[sr].height = 21.0 if name == "NTG" else None

    return sr


CURRENCIES = [
    ("USD", "دولار أمريكى"),
    ("EUR", "يورو"),
    ("GBP", "جنيَــه إسترليـنى"),
    ("CAD", "دولار كنـدى"),
    ("DKK", "كرون دانمركى"),
    ("NOK", "كرون نرويجى"),
    ("SEK", "كرون سَــويدى"),
    ("CHF", "فرنك سويسرى"),
    ("JPY", "100 ين يابانى"),
    ("SAR", "ريـــال سعــودى"),
    ("KWD", "دينــار كويتى"),
    ("AED", "درهم اماراتى"),
    ("AUD", "دولار اســـترالى"),
    ("BHD", "دينــار البحــرين"),
    ("OMR", "ريـــال عمـــانى"),
    ("QAR", "ريـــال قطــــرى"),
    ("JOD", "دينار اردنى"),
    ("CNY", "يوان صينى"),
]

# --- Add these to excel_export.py ---
ZEBRA_DARK = _fill("FFD9D9D9")
ZEBRA_LIGHT = _fill("FFF2F2F2")
FILL_BLACK = _fill("FF000000")


def _apply_zebra_striping(ws, row, col_count):
    # This automatically determines color:
    # Row 2 (even) -> Dark, Row 3 (odd) -> Light, Row 4 (even) -> Dark...
    is_dark = row % 2 == 0
    fill = ZEBRA_DARK if is_dark else ZEBRA_LIGHT
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = Font(bold=True, name="Arial", color="FF000000")
        cell.border = _thin()


def build_exchange_rates_sheet(ws, rates_list, balance_entries):
    FILL_BLACK = _fill("FF000000")
    # Header
    for c, h in enumerate(["العملة", "شراء", "بيع"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, size=12, name="Arial", color=WHITE)
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    rate_map = {r.currency_code: r for r in rates_list}
    eur_row = 0  # To store row index of EUR for the formula

    for i, (code, arabic) in enumerate(CURRENCIES, 2):
        _apply_zebra_striping(ws, i, 3)
        r = rate_map.get(code)
        ws.cell(row=i, column=1, value=arabic)
        if r:
            val_buy = (
                round(float(r.buy_rate) * 100, 4)
                if code == "JPY"
                else float(r.buy_rate)
            )
            val_sell = (
                round(float(r.sell_rate) * 100, 4)
                if code == "JPY"
                else float(r.sell_rate)
            )
            ws.cell(row=i, column=2, value=val_buy)
            ws.cell(row=i, column=3, value=val_sell)

        if code == "EUR":
            eur_row = i

    # --- Side block (Cols F-H, Rows 5-6) ---
    from core.models import Currency

    try:
        usd_cur = Currency.objects.get(code="USD")
        eur_cur = Currency.objects.get(code="EUR")
        home_usd = sum(
            float(be.amount)
            for be in balance_entries
            if be.bank_id is None and be.currency_id == usd_cur.id
        )
        home_eur = sum(
            float(be.amount)
            for be in balance_entries
            if be.bank_id is None and be.currency_id == eur_cur.id
        )
    except Exception:
        home_usd = home_eur = 0

    # Helper for side block styling with specific formatting
    def _style_side(r, c, val, fmt=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = _f(name="Arial")
        cell.alignment = _center()
        cell.border = _thin()
        if fmt:
            cell.number_format = fmt
        return cell

    # Row 5: Amounts (F5=EUR, G5=USD)
    _style_side(5, 6, home_eur, FMT_EUR)  # Euro format for F5
    _style_side(5, 7, home_usd, FMT_USD)  # Dollar format for G5
    _style_side(5, 8, "Total")

    # Row 6: Converted Values (EGP)
    _style_side(6, 6, f"=F5*B{eur_row}", FMT_EGP_RED)  # EGP format
    _style_side(6, 7, "=G5*B2", FMT_EGP_RED)  # EGP format
    _style_side(6, 8, "=F6+G6", FMT_EGP_RED)  # EGP format


def build_gold_price_sheet(ws, gold_qs, balance_entries):
    # Setup Columns
    ws.column_dimensions["A"].width = 18.6
    ws.column_dimensions["B"].width = 10.7
    ws.column_dimensions["C"].width = 10.7
    ws.column_dimensions["D"].width = 10.7
    ws.column_dimensions["E"].width = 10.7
    ws.column_dimensions["F"].width = 32.7
    ws.column_dimensions["G"].width = 15.8
    ws.column_dimensions["H"].width = 14.3

    FILL_BLACK = _fill("FF000000")
    latest = gold_qs.order_by("-fetched_at").first()

    # 1. Header
    for c, h in enumerate(["السعر", "شراء", "بيع", "المزيد", "الملاحظات"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial")
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    # 2. Main Gold Rows (2-6)
    carats = [
        ("جرام عيار 24", "carat_24k_buy", "carat_24k", "40 ج"),
        ("جرام عيار 22", "carat_22k_buy", "carat_22k", "37 ج"),
        ("جرام عيار 21", "carat_21k_buy", "carat_21k", "35 ج"),
        ("جرام عيار 18", "carat_18k_buy", "carat_18k", "30 ج"),
        ("جرام عيار 14 27 ج", "carat_18k_buy", "carat_18k", "27 ج"),
    ]
    for i, (label, bf, sf, karat) in enumerate(carats, 2):
        _apply_zebra_striping(ws, i, 5)
        for col, val in enumerate(
            [
                label,
                round(float(getattr(latest, bf, 0)), 0) if latest else 0,
                round(float(getattr(latest, sf, 0)), 0) if latest else 0,
                ">",
                karat,
            ],
            1,
        ):
            c = ws.cell(row=i, column=col, value=val)
            c.font = _f(name="Arial")
            c.alignment = _center()
            c.border = _thin()

    # 3. Dollar, Ounce, Gold Pound Rows (7-9)
    # Using the same zebra striping loop style for consistency
    rows_data = [
        [
            "الدولار 0 ج",
            float(latest.usd_to_egp) if latest else 0,
            float(latest.usd_to_egp) if latest else 0,
            None,
            "0 ج",
        ],
        [
            "الأونصة 0 $",
            float(latest.usd_per_oz) if latest else 0,
            float(latest.usd_per_oz) if latest else 0,
            None,
            "0 $",
        ],
        [
            "الجنيه الذهب 320 ج",
            round(float(latest.carat_21k_buy) * 8, 0) if latest else 0,
            round(float(latest.carat_21k) * 8, 0) if latest else 0,
            None,
            "320 ج",
        ],
    ]
    for i, row_data in enumerate(rows_data, 7):
        _apply_zebra_striping(ws, i, 5)
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = _f(name="Arial")
            c.alignment = _center()
            c.border = _thin()

    # 4. Side Table (Under the main table)
    from core.models import Currency

    grams = 0
    try:
        gold_cur = Currency.objects.get(code="Gold")
        grams = sum(
            float(be.amount)
            for be in balance_entries
            if be.bank_id is None and be.currency_id == gold_cur.id
        )
    except:
        pass

    # Merged header
    ws.merge_cells("G9:I9")
    c_title = ws.cell(row=9, column=7, value=f"{int(grams)} Grams")
    c_title.font = _f(name="Arial", bold=True)
    c_title.alignment = _center()
    c_title.border = _thin()
    ws.cell(row=9, column=8).border = _thin()
    ws.cell(row=9, column=9).border = _thin()

    # Titles
    for c, title in enumerate(["Now", "Paid", "Diff"], 7):
        c_head = ws.cell(row=10, column=c, value=title)
        c_head.font = _f(name="Arial", bold=True)
        c_head.alignment = _center()
        c_head.border = _thin()

    # Values
    vals = [f"=(C2+28.5)*(BALANCE!F2)", 897375, "=G11-H11"]
    for c, val in enumerate(vals, 7):
        c_val = ws.cell(row=11, column=c, value=val)
        c_val.font = _f(name="Arial")
        c_val.alignment = _center()
        c_val.border = _thin()
        # Apply EGP format to all three cells as they are all currency/financial results
        c_val.number_format = FMT_EGP_RED


def build_bank_certificates_sheet(ws, certs_qs):
    FILL_BLACK = _fill("FF000000")
    hdrs = [
        "Amount",
        "Interest Rate",
        "Interest Value",
        "Frequency",
        "Start Date",
        "End Date",
    ]

    # Header
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial")
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    for i, cert in enumerate(certs_qs.order_by("issue_date"), 2):
        _apply_zebra_striping(ws, i, 6)  # No is_dark argument needed
        ws.cell(row=i, column=1, value=float(cert.amount)).number_format = FMT_EGP_CERT
        ws.cell(row=i, column=2, value=float(cert.interest_rate)).number_format = (
            FMT_PCT
        )
        ws.cell(row=i, column=3, value=f"=(A{i}*B{i})/12").number_format = (
            FMT_EGP_CERT_R
        )
        ws.cell(row=i, column=4, value=cert.frequency)
        ws.cell(row=i, column=5, value=cert.issue_date).number_format = FMT_DATE
        ws.cell(row=i, column=6, value=cert.expiry_date).number_format = FMT_DATE


# ── BALANCE ───────────────────────────────────────────────────────────────────


def build_balance_sheet(ws, balance_entries, company_sheet_rows):

    ws.column_dimensions["A"].width = 26.5
    ws.column_dimensions["B"].width = 14.6
    ws.column_dimensions["C"].width = 12.5
    ws.column_dimensions["D"].width = 13.1
    ws.column_dimensions["E"].width = 13.1
    ws.column_dimensions["F"].width = 14.4
    ws.column_dimensions["G"].width = 14.4
    ws.column_dimensions["H"].width = 17.1
    ws.column_dimensions["I"].width = 12.7
    ws.column_dimensions["J"].width = 12.0
    ws.column_dimensions["K"].width = 29.7
    ws.row_dimensions[7].height = 18.0

    # Row 1 headers — bold, Arial, borders
    hdrs = [
        "Title",
        "EGP",
        "USD",
        "EUR",
        "SAR",
        "Gold",
        "Acct-Number",
        "Card-ID",
        "Swift-Code",
        "Customer-id",
        "Customer-Name",
    ]
    border_map = {
        1: _thin(),
        2: _thin(),
        3: _thin(),
        4: _thin_lr(),
        5: _thin_lr(),
        6: _thin_lr(),
        7: _thin_lr(),
        8: _thin_lr(),
        9: _thin(),
        10: _thin(),
        11: _thin(),
    }
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _f(bold=True, name="Arial")
        cell.border = border_map.get(c, _thin())

    from core.models import Currency, Bank as BankModel

    cur_map = {c.id: c.code for c in Currency.objects.all()}
    bank_map = {b.id: b for b in BankModel.objects.all()}

    # Row 2: Home Balance
    home = {
        cur_map.get(be.currency_id, "?"): float(be.amount)
        for be in balance_entries
        if be.bank_id is None and be.title == "Home Balance"
    }

    ws.cell(row=2, column=1, value="Home Balance").font = _f(bold=True, name="Arial")
    ws.cell(row=2, column=1).border = _thin()

    b2 = ws.cell(row=2, column=2, value=home.get("EGP", 0))
    b2.font = _f(bold=True, name="Arial")
    b2.border = _thin()
    b2.number_format = FMT_EGP_RED

    c2 = ws.cell(row=2, column=3, value=home.get("USD", 0))
    c2.font = _f(bold=True, name="Arial")
    c2.border = _thin()
    c2.number_format = FMT_USD

    d2 = ws.cell(row=2, column=4, value=home.get("EUR", 0))
    d2.font = _f(bold=True, name="Arial")
    d2.border = _thin()
    d2.number_format = FMT_EUR

    e2 = ws.cell(row=2, column=5, value=home.get("SAR", 0))
    e2.font = _f(bold=True, name="Arial")
    e2.border = _thin()
    e2.number_format = FMT_SAR

    f2 = ws.cell(row=2, column=6, value=home.get("Gold", 0))
    f2.font = _f(bold=True, name="Arial")
    f2.border = _thin()
    f2.number_format = FMT_GOLD

    # Bank rows
    excel_row = 3
    for be in sorted(balance_entries, key=lambda b: b.id):
        if be.title in ("Home Balance", "QNB Certificates Balance"):
            continue
        if cur_map.get(be.currency_id) != "EGP":
            continue

        a = ws.cell(row=excel_row, column=1, value=be.title)
        a.font = _f(bold=True, name="Arial")
        a.border = _thin()
        b = ws.cell(row=excel_row, column=2, value=float(be.amount))
        b.font = _f(bold=True, name="Arial")
        b.border = _thin()
        b.number_format = FMT_EGP_RED

        bank = bank_map.get(be.bank_id)
        if bank:
            for col, attr in [
                (7, "account_number"),
                (8, "card_number"),
                (9, "swift_code"),
                (10, "customer_id"),
                (11, "customer_name"),
            ]:
                v = getattr(bank, attr, "") or ""
                cell = ws.cell(row=excel_row, column=col, value=v)
                cell.font = _f(bold=True, name="Arial")
                cell.border = _thin()
                if col in (7, 8):
                    cell.number_format = FMT_INT
        excel_row += 1

    # QNB Certificates formula row
    from core.models import BankCertificate

    cert_count = BankCertificate.objects.count()
    cr = excel_row
    ws.cell(row=cr, column=1, value="QNB Certificates Balance").font = _f(
        bold=True, name="Arial"
    )
    ws.cell(row=cr, column=1).border = _thin()
    bc = ws.cell(
        row=cr, column=2, value=f"=SUM('Bank-Certificates'!A2:A{cert_count+1})"
    )
    bc.font = _f(bold=True, name="Arial")
    bc.border = _thin()
    bc.number_format = FMT_EGP_RED
    excel_row += 1

    # Total EGP
    ter = excel_row
    ws.cell(row=ter, column=1, value="Total EGP Balance").font = _f(
        bold=True, name="Arial"
    )
    ws.cell(row=ter, column=1).border = _thin()
    te = ws.cell(row=ter, column=2, value=f"=SUM(B2:B{cr})")
    te.font = _f(bold=True, name="Arial")
    te.border = _thin()
    te.number_format = FMT_EGP_RED
    excel_row += 1

    # Total all Balances — merged B:F, row height 18
    tar = excel_row
    ws.row_dimensions[tar].height = 18.0
    ws.merge_cells(f"B{tar}:F{tar}")

    ws.cell(row=tar, column=1, value="Total all Balances").font = _f(
        bold=True, name="Arial"
    )
    ws.cell(row=tar, column=1).border = _thin()

    formula = (
        f"=B{ter}"
        f"+(C2*('Exchange Rates'!B2))"
        f"+(D2*('Exchange Rates'!B3))"
        f"+(E2*('Exchange Rates'!B11))"
        f"+(F2*(('Gold Price'!C2)+28.5))"
    )

    # Target the merged cell (B:F)
    ta = ws.cell(row=tar, column=2, value=formula)

    # 1. Apply Centering
    ta.alignment = Alignment(horizontal="center", vertical="center")

    # 2. Apply Full Border Frame
    # Create a border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Apply to all cells in the merged range B to F
    for col in range(2, 7):  # B=2, C=3, D=4, E=5, F=6
        ws.cell(row=tar, column=col).border = thin_border

    ta.font = _f(bold=True, name="Arial")
    ta.number_format = FMT_EGP_RED
    excel_row += 1

    # Spacer rows then Total Pays / Total Work Months
    tpr = excel_row + 3
    tmr = tpr + 1

    # Companies that store bonus separately in col F
    BONUS_COMPANIES = set(company_sheet_rows.keys())

    pay_parts, month_parts = [], []
    for cname, (sname, srow) in company_sheet_rows.items():
        ref = f"'{sname}'!{{c}}{srow}" if " " in sname else f"{sname}!{{c}}{srow}"
        if cname in BONUS_COMPANIES:
            # Total pays = salary (D) + bonus (F)
            pay_parts.append(f"({ref.format(c='D')}+{ref.format(c='F')})")
        else:
            pay_parts.append(ref.format(c="D"))
        month_parts.append(ref.format(c="B"))

    # Total Pays
    label_tp = ws.cell(row=tpr, column=1, value="Total Pays")
    label_tp.font = _f(bold=True, name="Arial")
    label_tp.border = _thin()

    tp = ws.cell(row=tpr, column=2, value="=" + "+".join(pay_parts) if pay_parts else 0)
    tp.font = _f(bold=True, name="Arial")
    tp.border = _thin()
    tp.number_format = FMT_EGP_RED

    # Total Work Months
    label_tm = ws.cell(row=tmr, column=1, value="Total Work Months")
    label_tm.font = _f(bold=True, name="Arial")
    label_tm.border = _thin()

    tm = ws.cell(
        row=tmr, column=2, value="=" + "+".join(month_parts) if month_parts else 0
    )
    tm.font = _f(bold=True, name="Arial")
    tm.border = _thin()


# ── Expenses (new) ────────────────────────────────────────────────────────────


def build_expenses_sheet(ws, expenses_qs):
    hdrs = [
        "Date",
        "Year",
        "Month",
        "Category",
        "Sub-Category",
        "Description",
        "Amount",
        "Currency",
        "Payment Method",
        "Notes",
    ]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial")
        cell.fill = _fill(EXP_BG)
        cell.alignment = _center()
    widths = [14, 8, 12, 18, 20, 35, 14, 10, 16, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    expenses = list(
        expenses_qs.select_related("category", "subcategory", "currency").order_by(
            "year", "month", "date"
        )
    )
    row = 2
    year_total_rows = {}

    for year, yg in groupby(expenses, key=lambda e: e.year):
        year_entries = list(yg)
        year_start = row
        for month, mg in groupby(year_entries, key=lambda e: e.month):
            month_entries = list(mg)
            mname = MONTH_ORDER[month - 1] if 1 <= month <= 12 else str(month)
            mstart = row
            for exp in month_entries:
                ws.cell(row=row, column=1, value=exp.date).number_format = "YYYY-MM-DD"
                ws.cell(row=row, column=2, value=exp.year)
                ws.cell(row=row, column=3, value=mname)
                ws.cell(
                    row=row, column=4, value=exp.category.name if exp.category else ""
                )
                ws.cell(
                    row=row,
                    column=5,
                    value=exp.subcategory.name if exp.subcategory else "",
                )
                ws.cell(row=row, column=6, value=exp.description or "")
                ws.cell(row=row, column=7, value=float(exp.amount)).number_format = (
                    FMT_EGP_CERT
                )
                ws.cell(
                    row=row,
                    column=8,
                    value=exp.currency.code if exp.currency else "EGP",
                )
                ws.cell(row=row, column=9, value=exp.payment_method or "")
                ws.cell(row=row, column=10, value=exp.notes or "")
                row += 1
            mend = row - 1
            ws.cell(row=row, column=3, value=f"{mname} Total")
            ws.cell(row=row, column=7, value=f"=SUM(G{mstart}:G{mend})")
            for c in range(1, 11):
                ws.cell(row=row, column=c).font = _f(bold=True, name="Arial")
                ws.cell(row=row, column=c).fill = _fill(EXP_MONTH_BG)
            row += 1
        year_end = row - 1
        ws.cell(row=row, column=2, value=f"{year} Total")
        ws.cell(
            row=row,
            column=7,
            value=f"=SUMIF(B{year_start}:B{year_end},{year},G{year_start}:G{year_end})",
        )
        for c in range(1, 11):
            ws.cell(row=row, column=c).font = _f(bold=True, name="Arial")
            ws.cell(row=row, column=c).fill = _fill(EXP_YEAR_BG)
        year_total_rows[year] = row
        row += 2

    if year_total_rows:
        grand = "+".join(f"G{r}" for r in year_total_rows.values())
        ws.cell(row=row, column=1, value="Grand Total")
        ws.cell(row=row, column=7, value=f"={grand}")
        for c in range(1, 11):
            ws.cell(row=row, column=c).font = Font(bold=True, color=WHITE, name="Arial")
            ws.cell(row=row, column=c).fill = _fill(EXP_BG)


# ── Main ──────────────────────────────────────────────────────────────────────


def generate_excel(output_path=None):
    from core.models import (
        Company,
        BalanceEntry,
        BankCertificate,
        ExchangeRate,
        GoldPrice,
        Expense,
    )
    from django.db.models import Max

    wb = Workbook()
    wb.remove(wb.active)

    companies = list(Company.objects.all().order_by("order"))
    balance_entries = list(
        BalanceEntry.objects.select_related("currency", "bank").all()
    )

    # Exchange Rates
    ws_ex = wb.create_sheet("Exchange Rates")
    latest_ids = ExchangeRate.objects.values("currency_code").annotate(
        latest=Max("fetched_at")
    )
    rates = []
    for item in latest_ids:
        r = ExchangeRate.objects.filter(
            currency_code=item["currency_code"], fetched_at=item["latest"]
        ).first()
        if r:
            rates.append(r)
    build_exchange_rates_sheet(ws_ex, rates, balance_entries)
    auto_adjust_columns(ws_ex)

    # Gold Price
    ws_gold = wb.create_sheet("Gold Price")
    build_gold_price_sheet(ws_gold, GoldPrice.objects, balance_entries)
    auto_adjust_columns(ws_gold)

    # Salary sheets — capture summary rows
    company_sheet_rows = {}
    for company in companies:
        entries = list(company.salary_entries.all())
        ws_sal = wb.create_sheet(company.name)
        sr = build_salary_sheet(ws_sal, company, entries)
        # For single-company sheets the BALANCE references the outer Total row
        if company.name in ("ElSeweedy Technology", "Dedalus", "Globemed"):
            company_sheet_rows[company.name] = (company.name, sr)
        else:
            company_sheet_rows[company.name] = (company.name, sr)
    # auto_adjust_columns(ws_sal)

    # Bank-Certificates
    ws_cert = wb.create_sheet("Bank-Certificates")
    build_bank_certificates_sheet(ws_cert, BankCertificate.objects.all())
    auto_adjust_columns(ws_cert)

    # BALANCE
    ws_bal = wb.create_sheet("BALANCE")
    build_balance_sheet(ws_bal, balance_entries, company_sheet_rows)
    # auto_adjust_columns(ws_bal)
    # Expenses
    ws_exp = wb.create_sheet("Expenses")
    build_expenses_sheet(ws_exp, Expense.objects.all())
    # auto_adjust_columns(ws_exp)
    if output_path:
        wb.save(output_path)
        return output_path
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf