# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false

import json
from decimal import Decimal, InvalidOperation
from django.contrib.auth import authenticate, get_user_model, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import UserPassesTestMixin
from django.http import JsonResponse
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum, Count
from django.db.utils import OperationalError, ProgrammingError
from django.shortcuts import render, get_object_or_404, redirect
from django.core.exceptions import ValidationError
from .models import (
    Company,
    SalaryEntry,
    Bank,
    BalanceEntry,
    AppSettings,
    ExchangeRate,
    GoldPrice,
    GoldPriceHistory,
    Currency,
    ExpenseCategory,
    ExpenseSubcategory,
    Expense,
    BankCertificate,
    BankCertificateInterestHistory,
    _is_certificate_active,
    PagePermission,
    PAGE_PERMISSION_CHOICES,
    UserProfile,
    ReminderRule,
    CertificateStatus,
    ReminderLog,
    REMINDER_TYPE_CHOICES,
    SALARY_TRIGGER_CHOICES,
    FixedAsset,
    RealEstateDetails,
    VehicleDetails,
    GoldDetails,
    OtherAssetDetails,
    AssetRenovation,
    AssetMaintenance,
    AssetInsurance,
    AssetFurniture,
    AssetValuationHistory,
    AssetPurchasePayment,
    AssetSale,
    AssetPhoto,
    AssetMortgage,
    AssetRental,
    GoldTypeSetting,
    GoldPuritySetting,
    EmailTemplate,

)
from django.core.paginator import Paginator, EmptyPage
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import json as _json
import datetime
import os
import io
from django.conf import settings
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    HRFlowable,
    Image as RLImage,
    PageBreak,
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from arabic_reshaper import reshape
from bidi.algorithm import get_display
from django.views.decorators.http import require_http_methods
from core.services.net_worth_service import NetWorthService
from core.services.financial_sync_service import FinancialSyncService
from core.services.document_service import DocumentService
from core.services.exchange_rate_service import ExchangeRateService
from core.services.gold_valuation_service import GoldValuationService
from core.services.property_valuation_service import PropertyValuationService
from core.services.reminder_automation_service import ReminderAutomationService
from core.services.auth_workflow_service import AuthWorkflowService, EmailTemplateService
from core.services.cash_flow_forecast_service import CashFlowForecastService
from core.services.wealth_growth_forecast_service import WealthGrowthForecastService

@method_decorator(csrf_exempt, name="dispatch")
class ExportExcelWorkbookView(View):
    """
    Generates a multi-tab Excel Workbook from live DB data,
    matching the original Balance.xlsx format, styles, and formulas,
    with an added Expenses tab.
    """

    def get(self, request):
        return self.post(request)

    def post(self, request):
        from .excel_export import generate_excel
        from datetime import date

        buf = generate_excel()
        filename = f"Balance_Tracker_{date.today().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


User = get_user_model()

PAGE_PERMISSION_KEYS = [key for key, _ in PAGE_PERMISSION_CHOICES]

# Month sort order — ensures API returns months in calendar order, not alphabetically
MONTH_ORDER = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
    "Quarter-Bonuses",
]


def month_sort_key(entry_dict):
    try:
        return MONTH_ORDER.index(entry_dict.get("month", ""))
    except ValueError:
        return len(MONTH_ORDER)


def _run_certificate_interest_sync():
    from .services.certificate_interest_service import CertificateInterestService

    return CertificateInterestService().synchronize()


def _parse_iso_date(value):
    if not value:
        return None
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        try:
            return datetime.date.fromisoformat(value)
        except ValueError:
            return value
    return value


@login_required(login_url="/accounts/login/")
def index(request):
    return render(request, "index.html")


def _api_auth_required(request):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Authentication required"}, status=401)
    return None


class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_staff

    def handle_no_permission(self):
        if not self.request.user.is_authenticated:
            return JsonResponse({"error": "Authentication required"}, status=401)
        return JsonResponse({"error": "Admin access required"}, status=403)


def _build_user_dict(user):
    profile = AuthWorkflowService.get_profile(user)
    return {
        "id": user.id,
        "username": user.username,
        "email": user.email,
        "is_active": user.is_active,
        "is_staff": user.is_staff,
        "is_superuser": user.is_superuser,
        "email_verified": profile.email_verified,
        "account_status": profile.account_status,
    }


def _get_user_allowed_pages(user):
    if user.is_staff or user.is_superuser:
        return PAGE_PERMISSION_KEYS
    return [perm.page for perm in user.page_permissions.all()]


def _request_lang(request):
    return (
        request.POST.get("lang", "").strip()
        or request.GET.get("lang", "").strip()
        or request.COOKIES.get("wf_lang", "").strip()
        or AppSettings.get("active_language", "en")
        or "en"
    )


def _render_auth(request, template_name, extra_context=None):
    context = {"lang_code": _request_lang(request)}
    if extra_context:
        context.update(extra_context)
    response = render(request, template_name, context)
    response.set_cookie("wf_lang", context["lang_code"], max_age=31536000, samesite="Lax")
    return response


def _render_auth_status(request, *, title_key, message_key, tone="info", cta_href="", cta_key=""):
    return _render_auth(
        request,
        "auth_status.html",
        {
            "title_key": title_key,
            "message_key": message_key,
            "tone": tone,
            "cta_href": cta_href,
            "cta_key": cta_key,
        },
    )


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        lang = _request_lang(request)
        user_for_status = User.objects.filter(username=username).first()
        if user_for_status is not None:
            block_key = AuthWorkflowService.get_login_block(user_for_status)
            if block_key != "auth_error_invalid_login":
                return _render_auth(request, "login.html", {"error_key": block_key, "prefill_username": username})
        user = authenticate(request, username=username, password=password)
        if user is not None:
            profile = AuthWorkflowService.get_profile(user)
            profile.preferred_language = lang
            profile.save(update_fields=["preferred_language", "updated_at"])
            login(request, user)
            response = redirect("/")
            response.set_cookie("wf_lang", lang, max_age=31536000, samesite="Lax")
            return response
        return _render_auth(request, "login.html", {"error_key": "auth_error_invalid_login", "prefill_username": username})
    return _render_auth(request, "login.html")


def signup_view(request):
    if request.method == "POST":
        result = AuthWorkflowService.register_user(
            request,
            username=request.POST.get("username", ""),
            email=request.POST.get("email", ""),
            password=request.POST.get("password", ""),
            confirm_password=request.POST.get("confirm_password", ""),
            full_name=request.POST.get("full_name", ""),
            lang=_request_lang(request),
        )
        context = {
            "prefill_username": request.POST.get("username", "").strip(),
            "prefill_email": request.POST.get("email", "").strip(),
            "prefill_full_name": request.POST.get("full_name", "").strip(),
        }
        if result.ok:
            context["success_key"] = result.message_key
            return _render_auth(request, "signup.html", context)

        context["error_key"] = result.error_key
        if result.extra:
            context.update(result.extra)
        return _render_auth(request, "signup.html", context)

    return _render_auth(request, "signup.html")


def forgot_password_view(request):
    if request.method == "POST":
        if request.content_type == "application/json":
            data = json.loads(request.body.decode("utf-8") if isinstance(request.body, bytes) else request.body)
            identifier = data.get("email", "") or data.get("identifier", "")
        else:
            identifier = request.POST.get("email", "") or request.POST.get("identifier", "")
        result = AuthWorkflowService.request_password_reset(request, identifier=identifier, lang=_request_lang(request))
        if request.path.startswith("/api/"):
            status_code = 200 if result.ok else 400
            payload = {"message_key": result.message_key} if result.ok else {"error_key": result.error_key, "error": result.error_key}
            return JsonResponse(payload, status=status_code)
        return _render_auth(
            request,
            "forgot_password.html",
            {
                "success_key": result.message_key if result.ok else "",
                "error_key": result.error_key if not result.ok else "",
                "prefill_email": identifier.strip(),
            },
        )
    return _render_auth(request, "forgot_password.html", {"prefill_email": request.GET.get("email", "").strip()})


def reset_password_view(request, token):
    if request.method == "GET":
        resolved_token, error_key = AuthWorkflowService.resolve_token(token, "password_reset")
        if resolved_token is None:
            return _render_auth_status(
                request,
                title_key="auth_reset_password_heading",
                message_key=error_key,
                tone="danger",
                cta_href="/accounts/forgot-password/",
                cta_key="auth_forgot_password_button",
            )
        return _render_auth(request, "reset_password.html", {"reset_token": token})

    if request.method == "POST":
        result = AuthWorkflowService.reset_password(
            token,
            password=request.POST.get("password", ""),
            confirm_password=request.POST.get("confirm_password", ""),
        )
        if result.ok:
            return _render_auth_status(
                request,
                title_key="auth_reset_password_heading",
                message_key=result.message_key,
                tone="success",
                cta_href="/accounts/login/",
                cta_key="auth_login_button",
            )
        return _render_auth(request, "reset_password.html", {"error_key": result.error_key, "reset_token": token})
    return _render_auth(request, "reset_password.html", {"reset_token": token})


def verify_email_view(request, token):
    result = AuthWorkflowService.verify_email(request, token)
    if result.ok:
        return _render_auth_status(
            request,
            title_key="auth_verify_email_title",
            message_key=result.message_key,
            tone="success",
            cta_href="/accounts/pending-approval/",
            cta_key="auth_pending_approval_cta",
        )
    return _render_auth_status(
        request,
        title_key="auth_verify_email_title",
        message_key=result.error_key,
        tone="danger",
        cta_href="/accounts/login/",
        cta_key="auth_login_button",
    )


def pending_approval_view(request):
    return _render_auth_status(
        request,
        title_key="auth_pending_approval_title",
        message_key="auth_status_pending_admin_approval",
        tone="info",
        cta_href="/accounts/login/",
        cta_key="auth_login_button",
    )


def account_rejected_view(request):
    return _render_auth_status(
        request,
        title_key="auth_account_rejected_title",
        message_key="auth_status_rejected",
        tone="danger",
        cta_href="/accounts/forgot-password/",
        cta_key="auth_forgot_password_button",
    )


def account_disabled_view(request):
    return _render_auth_status(
        request,
        title_key="auth_account_disabled_title",
        message_key="auth_status_disabled",
        tone="danger",
        cta_href="/accounts/login/",
        cta_key="auth_login_button",
    )


def admin_approve_account_view(request, token):
    result = AuthWorkflowService.approve_user(token, actor=request.user if request.user.is_authenticated else None)
    return _render_auth_status(
        request,
        title_key="auth_admin_approval_title",
        message_key=result.message_key if result.ok else result.error_key,
        tone="success" if result.ok else "danger",
        cta_href="/accounts/login/",
        cta_key="auth_login_button",
    )


def admin_reject_account_view(request, token):
    result = AuthWorkflowService.reject_user(token, actor=request.user if request.user.is_authenticated else None)
    return _render_auth_status(
        request,
        title_key="auth_admin_rejection_title",
        message_key=result.message_key if result.ok else result.error_key,
        tone="danger" if result.ok else "danger",
        cta_href="/accounts/login/",
        cta_key="auth_login_button",
    )


def logout_view(request):
    logout(request)
    return redirect("/accounts/login/")


class LoginAPIView(View):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        data = json.loads(
            request.body.decode("utf-8")
            if isinstance(request.body, bytes)
            else request.body
        )
        username = data.get("username", "").strip()
        password = data.get("password", "")
        user_for_status = User.objects.filter(username=username).first()
        if user_for_status is not None:
            block_key = AuthWorkflowService.get_login_block(user_for_status)
            if block_key != "auth_error_invalid_login":
                return JsonResponse({"error_key": block_key, "error": block_key}, status=400)
        user = authenticate(request, username=username, password=password)
        if user is None:
            return JsonResponse({"error_key": "auth_error_invalid_login", "error": "auth_error_invalid_login"}, status=400)
        profile = AuthWorkflowService.get_profile(user)
        profile.preferred_language = str(data.get("lang", "") or profile.preferred_language or "en")
        profile.save(update_fields=["preferred_language", "updated_at"])
        login(request, user)
        return JsonResponse(
            {
                "user": _build_user_dict(user),
                "allowed_pages": _get_user_allowed_pages(user),
            }
        )


class SignupAPIView(View):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        data = json.loads(
            request.body.decode("utf-8")
            if isinstance(request.body, bytes)
            else request.body
        )
        result = AuthWorkflowService.register_user(
            request,
            username=data.get("username", ""),
            email=data.get("email", ""),
            password=data.get("password", ""),
            confirm_password=data.get("confirm_password", ""),
            full_name=data.get("full_name", ""),
            lang=str(data.get("lang", "") or "en"),
        )
        if not result.ok:
            payload = {"error_key": result.error_key, "error": result.error_key}
            if result.extra:
                payload.update(result.extra)
            return JsonResponse(payload, status=400)
        return JsonResponse({"message_key": result.message_key}, status=201)


class LogoutAPIView(View):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        logout(request)
        return JsonResponse({"success": True})


class CurrentUserView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({"user": None, "allowed_pages": []})
        return JsonResponse(
            {
                "user": _build_user_dict(request.user),
                "allowed_pages": _get_user_allowed_pages(request.user),
            }
        )


class UserListView(AdminRequiredMixin, View):
    def get(self, request):
        # support pagination and search: ?page=1&page_size=20&q=term
        q = request.GET.get("q", "").strip()
        page = int(request.GET.get("page", 1) or 1)
        page_size = int(request.GET.get("page_size", 20) or 20)

        qs = User.objects.order_by("username").all()
        if q:
            qs = qs.filter(Q(username__icontains=q) | Q(email__icontains=q))

        paginator = Paginator(qs, page_size)
        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        users = [_build_user_dict(u) for u in page_obj.object_list]
        return JsonResponse(
            {
                "users": users,
                "page": page_obj.number,
                "page_size": page_size,
                "total": paginator.count,
                "num_pages": paginator.num_pages,
            }
        )

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        data = json.loads(
            request.body.decode("utf-8")
            if isinstance(request.body, bytes)
            else request.body
        )
        username = data.get("username", "").strip()
        email = data.get("email", "").strip()
        password = data.get("password", "")
        if not username or not email or not password:
            return JsonResponse(
                {"error": "username, email and password are required"}, status=400
            )
        if User.objects.filter(username=username).exists():
            return JsonResponse({"error": "Username is already taken"}, status=400)
        user = User.objects.create_user(
            username=username, email=email, password=password
        )
        user.is_active = data.get("is_active", True)
        user.is_staff = data.get("is_staff", False)
        user.is_superuser = data.get("is_superuser", False)
        user.save()
        if user.is_active:
            AuthWorkflowService.enable_user(user, actor=request.user)
        else:
            AuthWorkflowService.disable_user(user, actor=request.user)
        return JsonResponse({"user": _build_user_dict(user)}, status=201)


class UserDetailView(AdminRequiredMixin, View):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        return JsonResponse({"user": _build_user_dict(user)})

    def put(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        original_is_active = user.is_active
        data = json.loads(
            request.body.decode("utf-8")
            if isinstance(request.body, bytes)
            else request.body
        )
        for field in ["email", "is_active", "is_staff", "is_superuser"]:
            if field in data:
                setattr(user, field, data[field])
        if data.get("password"):
            user.set_password(data["password"])
        user.save()
        if "is_active" in data and data["is_active"] != original_is_active:
            if data["is_active"]:
                AuthWorkflowService.enable_user(user, actor=request.user)
            else:
                AuthWorkflowService.disable_user(user, actor=request.user)
        return JsonResponse({"user": _build_user_dict(user)})

    def delete(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        user.delete()
        return JsonResponse({"deleted": pk})


class UserPermissionListView(AdminRequiredMixin, View):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        permissions = user.page_permissions.all()
        return JsonResponse(
            {
                "permissions": [perm.to_dict() for perm in permissions],
                "available_pages": PAGE_PERMISSION_CHOICES,
            }
        )

    def post(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        data = json.loads(
            request.body.decode("utf-8")
            if isinstance(request.body, bytes)
            else request.body
        )
        page = data.get("page")
        if page not in PAGE_PERMISSION_KEYS:
            return JsonResponse({"error": "Invalid page permission"}, status=400)
        perm, created = PagePermission.objects.get_or_create(user=user, page=page)
        return JsonResponse(
            {"permission": perm.to_dict()}, status=201 if created else 200
        )


class UserBulkActionView(AdminRequiredMixin, View):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        data = json.loads(
            request.body.decode("utf-8")
            if isinstance(request.body, bytes)
            else request.body
        )
        action = data.get("action")
        ids = data.get("ids") or []
        if not action or not isinstance(ids, list):
            return JsonResponse({"error": "action and ids required"}, status=400)

        users = User.objects.filter(id__in=ids)
        changed = 0
        if action == "delete":
            changed = users.count()
            users.delete()
        elif action == "activate":
            changed = users.count()
            for user in users:
                AuthWorkflowService.enable_user(user, actor=request.user)
        elif action == "deactivate":
            changed = users.count()
            for user in users:
                AuthWorkflowService.disable_user(user, actor=request.user)
        elif action == "set_staff":
            val = bool(data.get("value"))
            changed = users.update(is_staff=val)
        elif action == "set_superuser":
            val = bool(data.get("value"))
            changed = users.update(is_superuser=val)
        else:
            return JsonResponse({"error": "unknown action"}, status=400)

        return JsonResponse({"changed": changed})


class UserPermissionDetailView(AdminRequiredMixin, View):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def delete(self, request, pk):
        perm = get_object_or_404(PagePermission, pk=pk)
        perm.delete()
        return JsonResponse({"deleted": pk})


class PagePermissionChoicesView(AdminRequiredMixin, View):
    def get(self, request):
        return JsonResponse({"available_pages": PAGE_PERMISSION_CHOICES})


@login_required(login_url="/accounts/login/")
def user_management_page(request):
    # Only staff (admins) can access the management UI
    if not request.user.is_staff:
        return redirect("/")
    return render(request, "user_management.html")


@method_decorator(csrf_exempt, name="dispatch")
class CompanyListView(View):
    def get(self, request):
        companies = Company.objects.all().order_by("order")
        return JsonResponse({"companies": [c.to_dict() for c in companies]})

    def post(self, request):
        data = json.loads(request.body)
        company = Company.objects.create(
            name=data["name"],
            display_name=data.get("display_name", data["name"]),
            group_name=data.get("group_name", ""),
            color_hex=data.get("color_hex", "#0d6efd"),
            is_active=data.get("is_active", True),
            order=data.get("order", 0),
        )
        return JsonResponse(company.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class CompanyDetailView(View):
    def get(self, request, pk):
        c = get_object_or_404(Company, pk=pk)
        return JsonResponse(c.to_dict())

    def put(self, request, pk):
        c = get_object_or_404(Company, pk=pk)
        data = json.loads(request.body)
        for field in [
            "name",
            "display_name",
            "group_name",
            "color_hex",
            "is_active",
            "order",
        ]:
            if field in data:
                setattr(c, field, data[field])
        c.save()
        return JsonResponse(c.to_dict())

    def delete(self, request, pk):
        c = get_object_or_404(Company, pk=pk)
        c.delete()
        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class SalaryListView(View):
    def get(self, request):
        qs = SalaryEntry.objects.select_related("company").all()
        company_id = request.GET.get("company")
        year = request.GET.get("year")
        if company_id:
            qs = qs.filter(company_id=company_id)
        if year:
            qs = qs.filter(year=year)
        return JsonResponse(
            {"entries": sorted([e.to_dict() for e in qs], key=month_sort_key)}
        )

    def post(self, request):
        data = json.loads(request.body)
        entry = SalaryEntry.objects.create(
            company_id=data["company_id"],
            year=data["year"],
            month=data["month"],
            expected=data.get("expected", 0),
            paid=data.get("paid", 0),
            bonus=data.get("bonus", 0),
            notes=data.get("notes", ""),
        )
        return JsonResponse(entry.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class SalaryDetailView(View):
    def put(self, request, pk):
        entry = get_object_or_404(SalaryEntry, pk=pk)
        data = json.loads(request.body)
        for field in ["year", "month", "expected", "paid", "bonus", "notes"]:
            if field in data:
                setattr(entry, field, data[field])
        entry.save()
        return JsonResponse(entry.to_dict())

    def delete(self, request, pk):
        entry = get_object_or_404(SalaryEntry, pk=pk)
        entry.delete()
        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class SalarySummaryView(View):
    def get(self, request):
        companies = Company.objects.all().order_by("order")
        result = []
        grand = {
            "total_months": 0,
            "total_expected": 0.0,
            "total_paid": 0.0,
            "total_remaining": 0.0,
            "total_bonus": 0.0,
        }
        for c in companies:
            entries = c.salary_entries.all()
            agg = entries.aggregate(
                months=Count("id", filter=Q(paid__gt=0)),
                exp=Sum("expected"),
                paid=Sum("paid"),
                bonus=Sum("bonus"),
            )
            exp = float(agg["exp"] or 0)
            paid = float(agg["paid"] or 0)
            bonus = float(agg["bonus"] or 0)
            # Calculate company total
            company_total_paid = paid + bonus
            company_total_exp = exp + bonus
            company_remaining = max(0.0, exp - paid)  # Remaining = Expected - Base Paid
            result.append(
                {
                    "id": c.id,
                    "name": c.name,
                    "display_name": c.display_name,
                    "group_name": c.group_name,
                    "color_hex": c.color_hex,
                    "total_months": agg["months"],
                    "total_expected": company_total_exp,  # Corrected to include bonus in total expected
                    "total_paid": company_total_paid,  # Corrected to include bonus in total paid
                    "total_remaining": company_remaining,  # Remaining is still based on expected vs base paid
                    "total_bonus": bonus,
                    "years": list(
                        entries.values_list("year", flat=True)
                        .distinct()
                        .order_by("year")
                    ),
                }
            )
            grand["total_months"] += agg["months"]
            grand["total_expected"] += exp
            grand["total_paid"] += company_total_paid  # ADDED BONUS HERE
            # This ensures the grand total is the sum of the individual rows
            grand["total_remaining"] += company_remaining
            grand["total_bonus"] += bonus
        return JsonResponse({"companies": result, "grand_total": grand})


@method_decorator(csrf_exempt, name="dispatch")
class BankListView(View):
    def get(self, request):
        return JsonResponse({"banks": [b.to_dict() for b in Bank.objects.all()]})

    def post(self, request):
        data = json.loads(request.body)
        bank = Bank.objects.create(
            name=data["name"],
            account_number=data.get("account_number", ""),
            card_id=data.get("card_id", ""),
            swift_code=data.get("swift_code", ""),
            customer_id=data.get("customer_id", ""),
            customer_name=data.get("customer_name", ""),
            is_active=data.get("is_active", True),
            order=data.get("order", 0),
        )
        return JsonResponse(bank.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class BankDetailView(View):
    def put(self, request, pk):
        bank = get_object_or_404(Bank, pk=pk)
        data = json.loads(request.body)
        for field in [
            "name",
            "account_number",
            "card_id",
            "swift_code",
            "customer_id",
            "customer_name",
            "is_active",
            "order",
        ]:
            if field in data:
                setattr(bank, field, data[field])
        bank.save()
        return JsonResponse(bank.to_dict())

    def delete(self, request, pk):
        bank = get_object_or_404(Bank, pk=pk)
        bank.delete()
        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class BankCertificateListView(View):
    def get(self, request):
        certificates = BankCertificate.objects.select_related("bank", "currency").all()
        return JsonResponse({"certificates": [c.to_dict() for c in certificates]})

    def post(self, request):
        data = json.loads(request.body)
        certificate = BankCertificate.objects.create(
            bank_id=data["bank_id"],
            currency_id=data.get("currency_id"),
            issue_date=data.get("issue_date") or None,
            expiry_date=data.get("expiry_date") or None,
            amount=data.get("amount", 0),
            interest_rate=data.get("interest_rate", 0),
            interest_value=data.get("interest_value", 0),
            frequency=data.get("frequency", ""),
            status=data.get("status", "Active"),
            notes=data.get("notes", ""),
        )
        return JsonResponse(certificate.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class BankCertificateDetailView(View):
    def get(self, request, pk):
        certificate = get_object_or_404(BankCertificate, pk=pk)
        return JsonResponse(certificate.to_dict())

    def put(self, request, pk):
        certificate = get_object_or_404(BankCertificate, pk=pk)
        data = json.loads(request.body)
        for field in [
            "bank_id",
            "currency_id",
            "issue_date",
            "expiry_date",
            "amount",
            "interest_rate",
            "interest_value",
            "frequency",
            "status",
            "notes",
        ]:
            if field in data:
                setattr(certificate, field, data[field])
        certificate.save()
        return JsonResponse(certificate.to_dict())

    def delete(self, request, pk):
        certificate = get_object_or_404(BankCertificate, pk=pk)
        certificate.delete()
        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class BankCertificateInterestHistoryView(View):
    def get(self, request, certificate_id):
        certificate = get_object_or_404(BankCertificate, pk=certificate_id)
        rows = (
            BankCertificateInterestHistory.objects.select_related("bank", "currency")
            .filter(certificate_id=certificate_id)
            .order_by("-posting_date", "-id")
        )

        start = request.GET.get("start")
        end = request.GET.get("end")
        if start:
            rows = rows.filter(posting_date__gte=start)
        if end:
            rows = rows.filter(posting_date__lte=end)

        return JsonResponse(
            {
                "certificate": certificate.to_dict(),
                "items": [row.to_dict() for row in rows],
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class CurrencyListView(View):
    def get(self, request):
        currencies = Currency.objects.all().order_by("order")
        return JsonResponse({"currencies": [c.to_dict() for c in currencies]})

    def post(self, request):
        data = json.loads(request.body)
        currency = Currency.objects.create(
            code=data["code"],
            symbol=data.get("symbol", ""),
            flag=data.get("flag", "💱"),
            name=data.get("name", data["code"]),
            order=data.get("order", 0),
        )
        return JsonResponse(currency.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class CurrencyDetailView(View):
    def get(self, request, pk):
        c = get_object_or_404(Currency, pk=pk)
        return JsonResponse(c.to_dict())

    def put(self, request, pk):
        c = get_object_or_404(Currency, pk=pk)
        data = json.loads(request.body)
        for field in ["code", "symbol", "flag", "name", "order"]:
            if field in data:
                setattr(c, field, data[field])
        c.save()
        return JsonResponse(c.to_dict())

    def delete(self, request, pk):
        c = get_object_or_404(Currency, pk=pk)
        c.delete()
        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class BalanceListView(View):
    def _normalize_purity_key(self, purity_value):
        text = str(purity_value or "").strip().lower()
        if "24" in text or "999" in text:
            return "24k"
        if "22" in text or "916" in text:
            return "22k"
        if "21" in text or "875" in text:
            return "21k"
        if "18" in text or "750" in text:
            return "18k"
        return "24k"

    def _cashback_per_gram_for_purity(self, purity_value):
        key = self._normalize_purity_key(purity_value)
        setting = GoldPuritySetting.objects.filter(key=key, is_active=True).first()
        return float(setting.cashback_per_gram) if setting else 0.0

    def _sell_per_gram_for_purity(self, latest_gold, purity_value):
        if not latest_gold:
            return 0.0
        key = self._normalize_purity_key(purity_value)
        if key == "22k":
            return float(latest_gold.carat_22k or 0)
        if key == "21k":
            return float(latest_gold.carat_21k or 0)
        if key == "18k":
            return float(latest_gold.carat_18k or 0)
        return float(latest_gold.carat_24k or 0)

    def get(self, request):
        _run_certificate_interest_sync()
        return JsonResponse(NetWorthService().balance_payload())

    def post(self, request):
        data = json.loads(request.body)
        balance_type = data["balance_type"]
        purity = data.get("purity", "")
        if balance_type == BalanceEntry.BalanceType.GOLD:
            purity = _normalize_gold_purity(purity)
        else:
            purity = ""

        entry = BalanceEntry.objects.create(
            title=data["title"],
            balance_type=balance_type,
            bank_id=data.get("bank_id"),
            currency_id=data.get("currency_id", 1),
            purity=purity,
            amount=data.get("amount", 0),
            notes=data.get("notes", ""),
        )
        return JsonResponse(entry.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class BalanceDetailView(View):
    def put(self, request, pk):
        entry = get_object_or_404(BalanceEntry, pk=pk)
        data = json.loads(request.body)
        for field in [
            "title",
            "balance_type",
            "bank_id",
            "currency_id",
            "amount",
            "notes",
            "purity",
        ]:
            if field in data:
                setattr(entry, field, data[field])

        if entry.balance_type == BalanceEntry.BalanceType.GOLD:
            entry.purity = _normalize_gold_purity(entry.purity)
        else:
            entry.purity = ""

        entry.save()
        return JsonResponse(entry.to_dict())

    def delete(self, request, pk):
        entry = get_object_or_404(BalanceEntry, pk=pk)
        entry.delete()
        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class SettingsView(View):
    def get(self, request):
        settings = AppSettings.objects.all()
        return JsonResponse({"settings": {s.key: s.value for s in settings}})

    def post(self, request):
        data = json.loads(request.body)
        obj = AppSettings.set(data["key"], data["value"])
        return JsonResponse({"key": obj.key, "value": obj.value})


@method_decorator(csrf_exempt, name="dispatch")
class EmailTemplateListView(AdminRequiredMixin, View):
    def get(self, request):
        lang = request.GET.get("lang", "en")
        return JsonResponse({"items": EmailTemplateService.list_templates(lang)})


@method_decorator(csrf_exempt, name="dispatch")
class EmailTemplateDetailView(AdminRequiredMixin, View):
    def get(self, request, pk):
        lang = request.GET.get("lang", "en")
        template = get_object_or_404(EmailTemplate, pk=pk)
        EmailTemplateService.ensure_defaults()
        return JsonResponse(template.to_dict(lang))

    def put(self, request, pk):
        template = get_object_or_404(EmailTemplate, pk=pk)
        data = json.loads(request.body)
        lang = str(data.get("lang", "en") or "en")
        updated = EmailTemplateService.update_template(
            template,
            lang=lang,
            subject=(data.get("subject") or "").strip(),
            body=(data.get("body") or "").strip(),
        )
        return JsonResponse(updated.to_dict(lang))


@method_decorator(csrf_exempt, name="dispatch")
class EmailSettingsTestView(AdminRequiredMixin, View):
    def post(self, request):
        data = json.loads(request.body or "{}")
        recipient = (data.get("to_email") or "").strip()
        if not recipient:
            recipient = (
                AppSettings.get("administrator_notification_email", "").strip()
                or AppSettings.get("sender_email", "").strip()
            )

        ok, message_key = AuthWorkflowService.send_smtp_test_email(to_email=recipient)
        return JsonResponse(
            {
                "ok": ok,
                "message_key": message_key,
            },
            status=200 if ok else 400,
        )


def _seed_gold_settings_defaults():
    default_types = [
        ("Coins", 1),
        ("Bars", 2),
        ("Jewelry", 3),
    ]
    for name, order in default_types:
        GoldTypeSetting.objects.get_or_create(
            name=name,
            defaults={"is_active": True, "order": order},
        )

    default_purities = [
        ("24k", "24K", 0),
        ("22k", "22K", 0),
        ("21k", "21K", 0),
        ("18k", "18K", 0),
    ]
    for key, label, order in default_purities:
        GoldPuritySetting.objects.get_or_create(
            key=key,
            defaults={
                "label": label,
                "cashback_per_gram": 0,
                "is_active": True,
                "order": order,
            },
        )


@method_decorator(csrf_exempt, name="dispatch")
class GoldTypeSettingsListView(View):
    def get(self, request):
        _seed_gold_settings_defaults()
        rows = GoldTypeSetting.objects.all()
        return JsonResponse({"items": [row.to_dict() for row in rows]})

    def post(self, request):
        data = json.loads(request.body)
        item = GoldTypeSetting.objects.create(
            name=(data.get("name") or "").strip(),
            is_active=bool(data.get("is_active", True)),
            order=int(data.get("order", 0) or 0),
        )
        return JsonResponse(item.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class GoldTypeSettingsDetailView(View):
    def put(self, request, pk):
        item = get_object_or_404(GoldTypeSetting, pk=pk)
        data = json.loads(request.body)
        for field in ["name", "is_active", "order"]:
            if field in data:
                setattr(item, field, data[field])
        item.save()
        return JsonResponse(item.to_dict())

    def delete(self, request, pk):
        item = get_object_or_404(GoldTypeSetting, pk=pk)
        item.is_active = False
        item.save(update_fields=["is_active", "updated_at"])
        return JsonResponse({"disabled": pk})


@method_decorator(csrf_exempt, name="dispatch")
class GoldPuritySettingsListView(View):
    def get(self, request):
        _seed_gold_settings_defaults()
        rows = GoldPuritySetting.objects.all()
        return JsonResponse({"items": [row.to_dict() for row in rows]})

    def post(self, request):
        data = json.loads(request.body)
        key = str(data.get("key") or "").strip().lower()
        if key and not key.endswith("k"):
            key = f"{key}k"
        item = GoldPuritySetting.objects.create(
            key=key,
            label=(data.get("label") or "").strip() or key.upper(),
            cashback_per_gram=Decimal(str(data.get("cashback_per_gram", 0) or 0)),
            is_active=bool(data.get("is_active", True)),
            order=int(data.get("order", 0) or 0),
        )
        return JsonResponse(item.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class GoldPuritySettingsDetailView(View):
    def put(self, request, pk):
        item = get_object_or_404(GoldPuritySetting, pk=pk)
        data = json.loads(request.body)

        if "key" in data:
            key = str(data.get("key") or "").strip().lower()
            if key and not key.endswith("k"):
                key = f"{key}k"
            item.key = key

        if "label" in data:
            item.label = (data.get("label") or "").strip()

        if "cashback_per_gram" in data:
            item.cashback_per_gram = Decimal(str(data.get("cashback_per_gram") or 0))

        if "is_active" in data:
            item.is_active = bool(data.get("is_active"))

        if "order" in data:
            item.order = int(data.get("order") or 0)

        item.save()
        return JsonResponse(item.to_dict())

    def delete(self, request, pk):
        item = get_object_or_404(GoldPuritySetting, pk=pk)
        item.is_active = False
        item.save(update_fields=["is_active", "updated_at"])
        return JsonResponse({"disabled": pk})


# ── Exchange Rates views ──────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class ExchangeRateListView(View):
    """GET  /api/rates/          → latest rate per currency
    POST /api/rates/refresh/  → fetch from internet and save"""

    def get(self, request):
        """Return the single most-recent row per currency code."""
        from django.db.models import Max

        latest_ids = (
            ExchangeRate.objects.values("currency_code")
            .annotate(max_id=Max("id"))
            .values_list("max_id", flat=True)
        )
        rates = ExchangeRate.objects.filter(id__in=latest_ids).order_by("currency_code")
        last = ExchangeRate.objects.order_by("-fetched_at").first()
        return JsonResponse(
            {
                "rates": [r.to_dict() for r in rates],
                "fetched_at": (
                    last.fetched_at.strftime("%Y-%m-%d %H:%M") if last else None
                ),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class ExchangeRateRefreshView(View):
    """Calls open.er-api.com and saves latest rates to DB."""

    def post(self, request):
        try:
            result = ExchangeRateService().refresh_latest_rates().to_dict()
            return JsonResponse({**result, "message": f"Fetched {result['saved']} currencies"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=502)


# ── Gold Price views ──────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class GoldPriceListView(View):
    """GET /api/gold/ → latest gold price"""

    def get(self, request):
        latest = GoldPrice.objects.order_by("-fetched_at").first()
        if not latest:
            return JsonResponse(
                {"gold": None, "message": "No data yet. Click Refresh."}
            )
        return JsonResponse({"gold": latest.to_dict()})


@method_decorator(csrf_exempt, name="dispatch")
class GoldPriceRefreshView(View):
    """Fetches EGP gold prices from goldbullioneg.com and USD/EGP from open.er-api.com."""

    def get(self, request):
        return self.post(request)

    def post(self, request):
        try:
            result = GoldValuationService().refresh_latest_prices().to_dict()
            latest = GoldPrice.objects.order_by("-fetched_at").first()
            return JsonResponse({**result, "gold": latest.to_dict() if latest else None})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=502)


# ══════════════════════════════════════════════════════════════
# EXPENSE VIEWS
# ══════════════════════════════════════════════════════════════


@method_decorator(csrf_exempt, name="dispatch")
class ExpenseCategoryListView(View):
    def get(self, request):
        cats = ExpenseCategory.objects.prefetch_related("subcategories").all()
        data = []
        for c in cats:
            d = c.to_dict()
            d["subcategories"] = [s.to_dict() for s in c.subcategories.all()]
            data.append(d)
        return JsonResponse({"categories": data})

    def post(self, request):
        data = json.loads(request.body)
        cat = ExpenseCategory.objects.create(
            name=data["name"],
            icon=data.get("icon", "💰"),
            color_hex=data.get("color_hex", "#0d6efd"),
            order=data.get("order", 0),
        )
        return JsonResponse(cat.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class ExpenseCategoryDetailView(View):
    def put(self, request, pk):
        cat = get_object_or_404(ExpenseCategory, pk=pk)
        data = json.loads(request.body)
        for f in ["name", "icon", "color_hex", "order"]:
            if f in data:
                setattr(cat, f, data[f])
        cat.save()
        return JsonResponse(cat.to_dict())

    def delete(self, request, pk):
        cat = get_object_or_404(ExpenseCategory, pk=pk)
        cat.delete()
        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class ExpenseSubcategoryListView(View):
    def post(self, request):
        data = json.loads(request.body)
        sub = ExpenseSubcategory.objects.create(
            category_id=data["category_id"],
            name=data["name"],
            order=data.get("order", 0),
        )
        return JsonResponse(sub.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class ExpenseSubcategoryDetailView(View):
    def put(self, request, pk):
        sub = get_object_or_404(ExpenseSubcategory, pk=pk)
        data = json.loads(request.body)
        for f in ["name", "order"]:
            if f in data:
                setattr(sub, f, data[f])
        sub.save()
        return JsonResponse(sub.to_dict())

    def delete(self, request, pk):
        sub = get_object_or_404(ExpenseSubcategory, pk=pk)
        sub.delete()
        return JsonResponse({"deleted": pk})


def _normalize_expense_payment_method(method_value):
    method = str(method_value or "").strip().lower()
    if method == "cash":
        return "cash"
    if method in {"bank", "bank transfer", "bank_transfer"}:
        return "bank"
    if method == "card":
        return "card"
    return method


def _expense_requires_bank(method_value):
    return _normalize_expense_payment_method(method_value) in {"bank", "card"}


def _expense_affects_balance(method_value):
    return _normalize_expense_payment_method(method_value) in {"cash", "bank", "card"}


def _get_target_cash_balance_entry(payment_method, bank_id):
    qs = BalanceEntry.objects.select_for_update().filter(
        balance_type=BalanceEntry.BalanceType.CASH,
    )
    egp_or_cash_qs = qs.filter(
        Q(currency__code__iexact="EGP")
        | Q(currency__code__iexact="CASH")
        | Q(currency__name__iexact="Cash")
    )
    if egp_or_cash_qs.exists():
        qs = egp_or_cash_qs

    normalized_method = _normalize_expense_payment_method(payment_method)
    if normalized_method == "cash":
        qs = qs.filter(bank__isnull=True)
    else:
        qs = qs.filter(bank_id=bank_id)

    return qs.order_by("id").first()


def _apply_expense_balance_delta(payment_method, bank_id, amount_delta):
    if not _expense_affects_balance(payment_method):
        return

    delta = Decimal(str(amount_delta or 0))
    if delta == 0:
        return

    entry = _get_target_cash_balance_entry(payment_method, bank_id)
    if not entry:
        raise ValueError("matching_balance_entry_not_found")

    if delta < 0 and (Decimal(entry.amount or 0) + delta) < 0:
        raise ValueError("insufficient_balance")

    entry.amount = Decimal(entry.amount or 0) + delta
    entry.save(update_fields=["amount"])


@method_decorator(csrf_exempt, name="dispatch")
class ExpenseListView(View):
    def get(self, request):
        qs = Expense.objects.select_related("category", "subcategory", "currency", "bank").all()

        year = request.GET.get("year")
        month = request.GET.get("month")
        cat_id = request.GET.get("category")
        search = request.GET.get("search", "").strip()

        start_date = request.GET.get("start")
        end_date = request.GET.get("end")

        if start_date and end_date:
            qs = qs.filter(date__gte=start_date, date__lte=end_date)
        else:
            if year:
                qs = qs.filter(year=int(year))

            if month:
                qs = qs.filter(month=int(month))

        if cat_id:
            qs = qs.filter(category_id=int(cat_id))

        if search:
            qs = qs.filter(description__icontains=search) | qs.filter(
                notes__icontains=search
            )

        entries = [e.to_dict() for e in qs]

        total = sum(float(e["amount"] or 0) for e in entries)

        return JsonResponse({"entries": entries, "total": total})

    def post(self, request):
        data = json.loads(request.body)
        from datetime import date as _date

        payment_method = data.get("payment_method", "Cash")
        bank_id = data.get("bank_id")
        if _expense_requires_bank(payment_method) and not bank_id:
            return JsonResponse(
                {
                    "error": "Bank account is required for this payment method",
                    "error_key": "bank_account_required",
                },
                status=400,
            )
        if _normalize_expense_payment_method(payment_method) == "cash":
            bank_id = None

        d = _date.fromisoformat(data["date"])
        try:
            with transaction.atomic():
                amount_value = Decimal(str(data.get("amount", 0) or 0))
                if _expense_affects_balance(payment_method):
                    target_entry = _get_target_cash_balance_entry(payment_method, bank_id)
                    if not target_entry:
                        raise ValueError("matching_balance_entry_not_found")
                    if amount_value > Decimal(target_entry.amount or 0):
                        raise ValueError("insufficient_balance")

                exp = Expense.objects.create(
                    date=d,
                    year=d.year,
                    month=d.month,
                    category_id=data.get("category_id"),
                    subcategory_id=data.get("subcategory_id"),
                    description=data.get("description", ""),
                    amount=data.get("amount", 0),
                    currency_id=data.get("currency_id"),
                    bank_id=bank_id,
                    payment_method=payment_method,
                    notes=data.get("notes", ""),
                )
                _apply_expense_balance_delta(exp.payment_method, exp.bank_id, -exp.amount)
        except ValueError as exc:
            if str(exc) == "matching_balance_entry_not_found":
                return JsonResponse(
                    {
                        "error": "Matching balance entry not found",
                        "error_key": "matching_balance_not_found",
                    },
                    status=400,
                )
            if str(exc) == "insufficient_balance":
                return JsonResponse(
                    {
                        "error": "insufficient_balance",
                        "error_key": "insufficient_balance",
                    },
                    status=400,
                )
            raise

        return JsonResponse(exp.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class ExpenseDetailView(View):
    def put(self, request, pk):
        exp = get_object_or_404(Expense, pk=pk)
        data = json.loads(request.body)

        previous_amount = Decimal(exp.amount or 0)
        previous_method = exp.payment_method
        previous_bank_id = exp.bank_id

        if "date" in data:
            from datetime import date as _date

            d = _date.fromisoformat(data["date"])
            exp.date = d
            exp.year = d.year
            exp.month = d.month

        next_method = data.get("payment_method", exp.payment_method)
        next_bank_id = data.get("bank_id", exp.bank_id)
        if _expense_requires_bank(next_method) and not next_bank_id:
            return JsonResponse(
                {
                    "error": "Bank account is required for this payment method",
                    "error_key": "bank_account_required",
                },
                status=400,
            )

        for f in [
            "category_id",
            "subcategory_id",
            "description",
            "amount",
            "currency_id",
            "bank_id",
            "payment_method",
            "notes",
        ]:
            if f in data:
                setattr(exp, f, data[f])

        if _normalize_expense_payment_method(exp.payment_method) == "cash":
            exp.bank_id = None

        try:
            with transaction.atomic():
                next_amount = Decimal(exp.amount or 0)
                if _expense_affects_balance(exp.payment_method):
                    next_target = _get_target_cash_balance_entry(exp.payment_method, exp.bank_id)
                    if not next_target:
                        raise ValueError("matching_balance_entry_not_found")

                    available_balance = Decimal(next_target.amount or 0)
                    if _expense_affects_balance(previous_method):
                        previous_target = _get_target_cash_balance_entry(previous_method, previous_bank_id)
                        if previous_target and previous_target.id == next_target.id:
                            available_balance += previous_amount

                    if next_amount > available_balance:
                        raise ValueError("insufficient_balance")

                _apply_expense_balance_delta(previous_method, previous_bank_id, previous_amount)
                exp.save()
                _apply_expense_balance_delta(exp.payment_method, exp.bank_id, -Decimal(exp.amount or 0))
        except ValueError as exc:
            if str(exc) == "matching_balance_entry_not_found":
                return JsonResponse(
                    {
                        "error": "Matching balance entry not found",
                        "error_key": "matching_balance_not_found",
                    },
                    status=400,
                )
            if str(exc) == "insufficient_balance":
                return JsonResponse(
                    {
                        "error": "insufficient_balance",
                        "error_key": "insufficient_balance",
                    },
                    status=400,
                )
            raise

        return JsonResponse(exp.to_dict())

    def delete(self, request, pk):
        exp = get_object_or_404(Expense, pk=pk)

        try:
            with transaction.atomic():
                _apply_expense_balance_delta(exp.payment_method, exp.bank_id, Decimal(exp.amount or 0))
                exp.delete()
        except ValueError as exc:
            if str(exc) == "matching_balance_entry_not_found":
                return JsonResponse(
                    {
                        "error": "Matching balance entry not found",
                        "error_key": "matching_balance_not_found",
                    },
                    status=400,
                )
            raise

        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class ExpenseSummaryView(View):
    """Returns monthly totals + category breakdown for charts."""

    def get(self, request):
        from django.db.models import Sum
        import calendar
        import datetime

        year = request.GET.get("year")
        month = request.GET.get("month")
        qs = Expense.objects.all()
        if year:
            qs = qs.filter(year=int(year))
        if month:
            qs = qs.filter(month=int(month))

        # By category
        by_cat = {}
        for e in qs.select_related("category"):
            name = e.category.name if e.category else "Uncategorised"
            icon = e.category.icon if e.category else "💰"
            color = e.category.color_hex if e.category else "#6c757d"
            key = name
            if key not in by_cat:
                by_cat[key] = {"name": name, "icon": icon, "color": color, "total": 0}
            by_cat[key]["total"] += float(e.amount)

        # Monthly trend (last 12 months)
        monthly = []
        for m in range(1, 13):
            y = int(year) if year else datetime.date.today().year
            total = (
                Expense.objects.filter(year=y, month=m).aggregate(t=Sum("amount"))["t"]
                or 0
            )
            monthly.append({"month": m, "total": float(total)})

        grand_total = sum(v["total"] for v in by_cat.values())

        # Income summary for reports (salary + certificate interest)
        month_names = [
            "january",
            "february",
            "march",
            "april",
            "may",
            "june",
            "july",
            "august",
            "september",
            "october",
            "november",
            "december",
        ]

        def _month_bounds(y, m):
            start = datetime.date(y, m, 1)
            end = datetime.date(y, m, calendar.monthrange(y, m)[1])
            return start, end

        def _year_bounds(y):
            return datetime.date(y, 1, 1), datetime.date(y, 12, 31)

        def _includes_in_period(certificate, start_date, end_date):
            if not certificate:
                return False
            issue_date = certificate.issue_date
            expiry_date = certificate.expiry_date
            if issue_date and issue_date > end_date:
                return False
            if expiry_date and expiry_date < start_date:
                return False
            return True

        salary_amount = 0.0
        total_interest = 0.0
        rental_income = 0.0

        if month:
            target_year = int(year) if year else datetime.date.today().year
            target_month = int(month)
            prev_month = target_month - 1
            prev_year = target_year
            if prev_month == 0:
                prev_month = 12
                prev_year -= 1

            prev_month_name = month_names[prev_month - 1]
            salary_qs = SalaryEntry.objects.filter(
                year=prev_year,
                month__iexact=prev_month_name,
            )
            salary_amount = float(
                salary_qs.aggregate(total=Sum("paid"))["total"] or 0
            )

            start_date, end_date = _month_bounds(target_year, target_month)
            certs = BankCertificate.objects.all()
            total_interest = sum(
                float(c.interest_value or 0)
                for c in certs
                if _includes_in_period(c, start_date, end_date)
            )
            rental_income = float(FinancialSyncService().period_rental_income_total("month"))
        elif year:
            salary_amount = float(
                SalaryEntry.objects.filter(year=int(year)).aggregate(total=Sum("paid"))["total"]
                or 0
            )
            start_date, end_date = _year_bounds(int(year))
            certs = BankCertificate.objects.all()
            total_interest = sum(
                float(c.interest_value or 0)
                for c in certs
                if _includes_in_period(c, start_date, end_date)
            )
            rental_income = float(FinancialSyncService().period_rental_income_total("year"))
        else:
            today = datetime.date.today()
            start_date, end_date = _month_bounds(today.year, today.month)
            salary_amount = 0.0
            total_interest = sum(
                float(c.interest_value or 0)
                for c in BankCertificate.objects.all()
                if _includes_in_period(c, start_date, end_date)
            )
            rental_income = float(FinancialSyncService().period_rental_income_total("month"))

        total_income = salary_amount + total_interest + rental_income

        return JsonResponse(
            {
                "by_category": list(by_cat.values()),
                "monthly_trend": monthly,
                "grand_total": grand_total,
                "income_summary": {
                    "total_income": total_income,
                    "total_salary": salary_amount,
                    "total_interest": total_interest,
                    "total_rental_income": rental_income,
                },
            }
        )


# ══════════════════════════════════════════════════════════════
# PDF REPORT VIEW
# ══════════════════════════════════════════════════════════════
# Helper to load translations
def get_translations(lang):
    path = os.path.join(settings.BASE_DIR, "static", "i18n", f"{lang}.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}


def format_arabic(text):
    return get_display(reshape(str(text)))


def get_text(key, lang, t, default=""):
    text = t.get(key, default)
    return format_arabic(text) if lang == "ar" else text


@method_decorator(csrf_exempt, name="dispatch")
class GenerateReportView(View):
    """
    POST /api/reports/generate/
    body: { type: "monthly"|"yearly"|"custom",
            year: 2026, month: 5,       # for monthly
            start_date: "2026-01-01",   # for custom
            end_date:   "2026-05-31" }
    Returns: PDF file
    """

    def post(self, request):
        import json as _json, datetime
        from django.http import HttpResponse, JsonResponse
        from django.db.models import Sum

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate,
                Paragraph,
                Spacer,
                Table,
                TableStyle,
                HRFlowable,
            )
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            import io
        except ImportError:
            return JsonResponse(
                {"error": "reportlab not installed. Run: pip install reportlab"},
                status=500,
            )

        data = _json.loads(request.body)
        lang = data.get("lang", "en")
        t = get_translations(lang)

        # Register Arabic-compatible font if the file exists
        font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "arial.ttf")
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont("ArabicFont", font_path))

        # Decide which font style template to use
        pdf_font = "ArabicFont" if lang == "ar" else "Helvetica-Bold"

        rtype = data.get("type", "monthly")
        year = int(data.get("year", datetime.date.today().year))
        month = int(data.get("month", datetime.date.today().month))

        # Accept both parameter styles (with or without suffix) to be fully secure
        start_date = data.get("start_date") or data.get("start")
        end_date = data.get("end_date") or data.get("end")

        # ── Filter expenses safely across all field schemas ──
        qs = Expense.objects.select_related("category", "subcategory").all()
        if rtype == "monthly":
            qs = qs.filter(year=year, month=month)
            month_name = datetime.date(year, month, 1).strftime("%B")
            json_month_key = f"month_{month_name.lower()}"
            translated_month = t.get(json_month_key)

            if not translated_month:
                if lang == "ar":
                    ARABIC_MONTHS = {
                        "January": "يناير",
                        "February": "فبراير",
                        "March": "مارس",
                        "April": "أبريل",
                        "May": "مايو",
                        "June": "يونيو",
                        "July": "يوليو",
                        "August": "أغسطس",
                        "September": "سبتمبر",
                        "October": "أكتوبر",
                        "November": "نوفمبر",
                        "December": "ديسمبر",
                    }
                    translated_month = ARABIC_MONTHS.get(month_name, month_name)
                else:
                    translated_month = month_name

            # FIXED: Changed long em-dash (—) to standard universal hyphen (-)
            title_str = f"{t.get('monthly_report', 'Monthly Report')} - {translated_month} {year}"
            filename = f"report_{year}_{month:02d}.pdf"
        elif rtype == "yearly":
            qs = qs.filter(year=year)
            # FIXED: Changed to standard hyphen
            title_str = f"{t.get('yearly_report', 'Yearly Report')} - {year}"
            filename = f"report_{year}.pdf"
        else:
            from datetime import date as _date

            sd = _date.fromisoformat(start_date)
            ed = _date.fromisoformat(end_date)
            qs = qs.filter(date__gte=sd, date__lte=ed)

            title_str = f"{t.get('report', 'Report')} {start_date} {t.get('to', 'to')} {end_date}"
            filename = f"report_{start_date}_{end_date}.pdf"

        if lang == "ar":
            title_str = format_arabic(title_str)

        expenses = list(qs)
        total_exp = sum(float(e.amount) for e in expenses)

        # Income for period (salary paid amounts)
        total_inc = 0
        if rtype == "monthly":
            # Target the previous month relative to the report month
            curr_date = datetime.date(year, month, 1)
            prev_date = curr_date - datetime.timedelta(days=1)
            sal_qs = SalaryEntry.objects.filter(
                year=prev_date.year, month=prev_date.strftime("%B")
            )
        elif rtype == "yearly":
            sal_qs = SalaryEntry.objects.filter(year=year)
        else:
            from datetime import date as _date

            sd = _date.fromisoformat(start_date)
            ed = _date.fromisoformat(end_date)

            MONTHS = [
                "January",
                "February",
                "March",
                "April",
                "May",
                "June",
                "July",
                "August",
                "September",
                "October",
                "November",
                "December",
            ]

            sal_qs = SalaryEntry.objects.none()

            for year_num in range(sd.year, ed.year + 1):

                year_entries = SalaryEntry.objects.filter(year=year_num)

                for entry in year_entries:

                    try:

                        month_index = MONTHS.index(entry.month) + 1

                        entry_date = _date(year_num, month_index, 1)

                        if sd <= entry_date <= ed:

                            sal_qs |= SalaryEntry.objects.filter(pk=entry.pk)

                    except Exception:
                        pass

        total_inc += sum(float(s.paid or 0) for s in sal_qs)

        # 2. Add Bank Interest (Summing all certificates)
        total_interest = sum(
            float(c.interest_value or 0) for c in BankCertificate.objects.all()
        )
        # total_interest = 0
        total_inc += total_interest

        # 3. Final Calculations
        net_sav = total_inc - total_exp
        sav_rate = (net_sav / total_inc * 100) if total_inc > 0 else 0

        # Category breakdown
        cat_totals = {}
        for e in expenses:
            cname = e.category.name if e.category else "Uncategorised"
            cat_totals[cname] = cat_totals.get(cname, 0) + float(e.amount)

        # ── Build PDF ──
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        navy = colors.HexColor("#000080")
        blue = colors.HexColor("#1a6ef5")
        green = colors.HexColor("#00d68f")
        red = colors.HexColor("#ff4d6d")
        yellow = colors.HexColor("#ffd166")
        grey = colors.HexColor("#7b97cc")

        H1 = ParagraphStyle(
            "H1",
            fontSize=22,
            textColor=blue,
            spaceAfter=15,
            alignment=TA_CENTER,
            fontName=pdf_font,
        )
        H11 = ParagraphStyle(
            "H11",
            fontSize=18,
            textColor=navy,
            spaceAfter=15,
            alignment=TA_CENTER,
            fontName=pdf_font,
        )
        H2 = ParagraphStyle(
            "H2",
            fontSize=14,
            textColor=navy,
            spaceAfter=4,
            spaceBefore=12,
            fontName=pdf_font,
        )
        BODY = ParagraphStyle("BODY", fontSize=10, textColor=navy, spaceAfter=4)
        SUB = ParagraphStyle("SUB", fontSize=9, textColor=grey, spaceAfter=2)

        story = []

        # Cover
        story.append(Spacer(1, 1 * cm))

        # Fetch the clean, localized text without unstable emojis
        report_text = get_text("financial_report", lang, t, "Financial Report")

        # Append the titles cleanly to the story
        story.append(Paragraph(report_text, H1))
        story.append(Paragraph(title_str, H11))
        story.append(HRFlowable(width="100%", thickness=1, color=blue))
        story.append(Spacer(1, 0.5 * cm))
        # Dynamically set table title alignments based on the document language
        table_title_style = ParagraphStyle(
            "TableTitle", parent=H2, alignment=TA_RIGHT if lang == "ar" else TA_LEFT
        )
        # Summary KPIs
        story.append(
            Paragraph(get_text("summary", lang, t, "Summary"), table_title_style)
        )

        # Define explicit Paragraph styles for table cells to handle Arabic layout flawlessly
        cell_L = ParagraphStyle(
            "CellL", fontName=pdf_font, fontSize=10, textColor=navy, alignment=TA_LEFT
        )
        cell_R = ParagraphStyle(
            "CellR", fontName=pdf_font, fontSize=10, textColor=navy, alignment=TA_RIGHT
        )
        cell_HL = ParagraphStyle(
            "CellHL",
            fontName=pdf_font,
            fontSize=10,
            textColor=colors.white,
            alignment=TA_LEFT,
        )
        cell_HR = ParagraphStyle(
            "CellHR",
            fontName=pdf_font,
            fontSize=10,
            textColor=colors.white,
            alignment=TA_RIGHT,
        )

        kpi_data = [
            [
                Paragraph(get_text("metric", lang, t, "Metric"), cell_HL),
                Paragraph(get_text("amount", lang, t, "Amount (EGP)"), cell_HR),
            ],
            [
                Paragraph(get_text("total_income", lang, t, "Total Income"), cell_L),
                Paragraph(f"{total_inc:,.2f}", cell_R),
            ],
            [
                Paragraph(
                    get_text("total_expenses", lang, t, "Total Expenses"), cell_L
                ),
                Paragraph(f"{total_exp:,.2f}", cell_R),
            ],
            [
                Paragraph(get_text("net_savings", lang, t, "Net Savings"), cell_L),
                Paragraph(
                    f"{net_sav:,.2f}",
                    ParagraphStyle(
                        "NetSavR",
                        parent=cell_R,
                        textColor=green if net_sav >= 0 else red,
                    ),
                ),
            ],
            [
                Paragraph(get_text("savings_rate", lang, t, "Savings Rate"), cell_L),
                Paragraph(f"{sav_rate:.1f}%", cell_R),
            ],
        ]

        kpi_table = Table(kpi_data, colWidths=[9 * cm, 7 * cm])
        kpi_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), blue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, -1), pdf_font),  # Applied globally
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.HexColor("#f0f4ff"), colors.white],
                    ),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1e3a6e")),
                    ("PADDING", (0, 0), (-1, -1), 7),
                ]
            )
        )
        story.append(kpi_table)
        story.append(Spacer(1, 0.5 * cm))

        # Category breakdown
        if cat_totals:
            story.append(
                Paragraph(
                    get_text("cat_breakdown", lang, t, "Expense Breakdown by Category"),
                    table_title_style,
                )
            )

            cell_L9 = ParagraphStyle(
                "CellL9",
                fontName=pdf_font,
                fontSize=9,
                textColor=navy,
                alignment=TA_LEFT,
            )
            cell_R9 = ParagraphStyle(
                "CellR9",
                fontName=pdf_font,
                fontSize=9,
                textColor=navy,
                alignment=TA_RIGHT,
            )
            cell_HL9 = ParagraphStyle(
                "CellHL9",
                fontName=pdf_font,
                fontSize=9,
                textColor=colors.white,
                alignment=TA_LEFT,
            )
            cell_HR9 = ParagraphStyle(
                "CellHR9",
                fontName=pdf_font,
                fontSize=9,
                textColor=colors.white,
                alignment=TA_RIGHT,
            )

            cat_data = [
                [
                    Paragraph(get_text("category", lang, t, "Category"), cell_HL9),
                    Paragraph(get_text("amount", lang, t, "Amount (EGP)"), cell_HR9),
                    Paragraph(get_text("pct", lang, t, "% of Total"), cell_HR9),
                ]
            ]

            for cname, ctotal in sorted(cat_totals.items(), key=lambda x: -x[1]):
                pct = (ctotal / total_exp * 100) if total_exp > 0 else 0
                # Reshape dynamic database category names if language is Arabic
                display_cname = format_arabic(cname) if lang == "ar" else cname

                cat_data.append(
                    [
                        Paragraph(display_cname, cell_L9),
                        Paragraph(f"{ctotal:,.2f}", cell_R9),
                        Paragraph(f"{pct:.1f}%", cell_R9),
                    ]
                )

            cat_data.append(
                [
                    Paragraph(get_text("total", lang, t, "TOTAL"), cell_L9),
                    Paragraph(f"{total_exp:,.2f}", cell_R9),
                    Paragraph("100%", cell_R9),
                ]
            )

            cat_table = Table(cat_data, colWidths=[9 * cm, 5 * cm, 3 * cm])
            cat_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), blue),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        (
                            "FONTNAME",
                            (0, 0),
                            (-1, -1),
                            pdf_font,
                        ),  # Fixed range to cover all cells
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8f0fe")),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -2),
                            [colors.HexColor("#f0f4ff"), colors.white],
                        ),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1e3a6e")),
                        ("PADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            story.append(cat_table)
            story.append(Spacer(1, 0.5 * cm))

        # Detailed expense entries
        if expenses:
            story.append(
                Paragraph(
                    get_text("expense_entries", lang, t, "Expense Entries"),
                    table_title_style,
                )
            )

            cell_L8 = ParagraphStyle(
                "CellL8",
                fontName=pdf_font,
                fontSize=8,
                textColor=navy,
                alignment=TA_LEFT,
            )
            cell_R8 = ParagraphStyle(
                "CellR8",
                fontName=pdf_font,
                fontSize=8,
                textColor=navy,
                alignment=TA_RIGHT,
            )
            cell_HL8 = ParagraphStyle(
                "CellHL8",
                fontName=pdf_font,
                fontSize=8,
                textColor=colors.white,
                alignment=TA_LEFT,
            )
            cell_HR8 = ParagraphStyle(
                "CellHR8",
                fontName=pdf_font,
                fontSize=8,
                textColor=colors.white,
                alignment=TA_RIGHT,
            )

            exp_data = [
                [
                    Paragraph(get_text("date", lang, t, "Date"), cell_HL8),
                    Paragraph(get_text("category", lang, t, "Category"), cell_HL8),
                    Paragraph(
                        get_text("description", lang, t, "Description"), cell_HL8
                    ),
                    Paragraph(get_text("method", lang, t, "Method"), cell_HL8),
                    Paragraph(get_text("amount", lang, t, "Amount"), cell_HR8),
                ]
            ]

            for e in sorted(expenses, key=lambda x: x.date):
                cname = e.category.name if e.category else "—"
                desc = e.description or "—"
                method = e.payment_method or "—"

                # Reshape dynamic Arabic inputs from the database if active language is Arabic
                if lang == "ar":
                    cname = format_arabic(cname)
                    desc = format_arabic(desc[:40])
                    method = format_arabic(method)
                else:
                    desc = desc[:40]

                exp_data.append(
                    [
                        Paragraph(e.date.strftime("%d/%m/%Y"), cell_L8),
                        Paragraph(cname, cell_L8),
                        Paragraph(desc, cell_L8),
                        Paragraph(method, cell_L8),
                        Paragraph(f"{float(e.amount):,.2f}", cell_R8),
                    ]
                )

            exp_table = Table(
                exp_data, colWidths=[2.5 * cm, 3.5 * cm, 6 * cm, 3 * cm, 3 * cm]
            )
            exp_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), blue),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        (
                            "FONTNAME",
                            (0, 0),
                            (-1, -1),
                            pdf_font,
                        ),  # Fixed range to cover all cells
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.HexColor("#f0f4ff"), colors.white],
                        ),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#1e3a6e")),
                        ("PADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            story.append(exp_table)

        # Footer
        story.append(Spacer(1, 1 * cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=grey))

        # 1. Extract current date components
        today = datetime.date.today()
        f_day = today.day
        f_year = today.year
        f_month_name = today.strftime("%B")  # e.g., "June"

        # 2. Look up the translation key across ALL languages matching the header logic
        f_json_key = f"month_{f_month_name.lower()}"
        f_translated_month = t.get(f_json_key)

        # Fallback handling if a language JSON file is missing the specific month key
        if not f_translated_month:
            if lang == "ar":
                ARABIC_MONTHS = {
                    "January": "يناير",
                    "February": "فبراير",
                    "March": "مارس",
                    "April": "أبريل",
                    "May": "مايو",
                    "June": "يونيو",
                    "July": "يوليو",
                    "August": "أغسطس",
                    "September": "سبتمبر",
                    "October": "أكتوبر",
                    "November": "نوفمبر",
                    "December": "ديسمبر",
                }
                f_translated_month = ARABIC_MONTHS.get(f_month_name, f_month_name)
            else:
                f_translated_month = f_month_name

        # 3. Pull the "generated_by" label from the translation context
        raw_label = t.get("generated_by", "Generated by WealthFlow")

        # 4. Construct layout string based on text direction rules
        if lang == "ar":
            # Combined with a standard hyphen, then reshaped once safely
            raw_footer = f"{raw_label} - {f_day} {f_translated_month} {f_year}"
            footer_text = format_arabic(raw_footer)
        else:
            # Handles English, French, and all other LTR languages uniformly
            footer_text = f"{raw_label} - {f_day} {f_translated_month} {f_year}"

        # 5. Append the translated paragraph to the layout story
        story.append(
            Paragraph(
                footer_text,
                ParagraphStyle(
                    "F",
                    fontSize=8,
                    textColor=grey,
                    alignment=TA_CENTER,
                    fontName=pdf_font,
                ),
            )
        )

        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # ── Profile update + avatar upload ───────────────────────────


from django.db.models.signals import post_save
from django.dispatch import receiver


@receiver(post_save, sender=User)
def create_user_profile(sender, instance, created, **kwargs):
    """Auto-create a UserProfile whenever a new User is created."""
    if created:
        UserProfile.objects.get_or_create(user=instance)


@method_decorator(csrf_exempt, name="dispatch")
class UpdateProfileView(View):
    """
    GET  /api/auth/profile/          — get current user profile
    POST /api/auth/profile/          — update full_name / bio / birthday
    POST /api/auth/profile/avatar/   — upload profile picture (multipart)
    """

    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({"error": "Not authenticated"}, status=401)
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        return JsonResponse(
            {"profile": profile.to_dict(), "user": _build_user_dict(request.user)}
        )

    def post(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({"error": "Not authenticated"}, status=401)
        profile, _ = UserProfile.objects.get_or_create(user=request.user)

        # Handle avatar upload — store as base64 in DB (no file system)
        if request.FILES.get("avatar"):
            import base64 as _b64

            f = request.FILES["avatar"]
            mime_type = f.content_type or "image/jpeg"
            raw_bytes = f.read()
            # Resize to max 256x256 to keep DB size reasonable
            try:
                from PIL import Image
                import io as _io

                img = Image.open(_io.BytesIO(raw_bytes))
                img.thumbnail((256, 256), Image.LANCZOS)
                buf = _io.BytesIO()
                fmt = "JPEG" if "jpeg" in mime_type or "jpg" in mime_type else "PNG"
                img.save(buf, format=fmt, quality=85)
                raw_bytes = buf.getvalue()
                mime_type = "image/jpeg" if fmt == "JPEG" else "image/png"
            except Exception:
                pass  # If Pillow not available, store full image
            b64_str = _b64.b64encode(raw_bytes).decode("utf-8")
            profile.avatar_b64 = f"data:{mime_type};base64,{b64_str}"
            profile.save()
            return JsonResponse(
                {"avatar_url": profile.avatar_url(), "message": "Avatar updated"}
            )

        # Handle JSON profile update
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        if "full_name" in data:
            profile.full_name = data["full_name"].strip()
            # Also update Django User first/last name
            parts = profile.full_name.split(" ", 1)
            request.user.first_name = parts[0]
            request.user.last_name = parts[1] if len(parts) > 1 else ""
            request.user.save(update_fields=["first_name", "last_name"])
        if "bio" in data:
            profile.bio = data["bio"]
        if "birthday" in data:
            raw_birthday = data.get("birthday")
            if raw_birthday in (None, ""):
                profile.birthday = None
            elif isinstance(raw_birthday, str):
                try:
                    parsed_birthday = datetime.date.fromisoformat(raw_birthday.strip())
                except ValueError:
                    return JsonResponse({"error": "Invalid birthday format. Use YYYY-MM-DD."}, status=400)
                if parsed_birthday > timezone.localdate():
                    return JsonResponse({"error": "Birthday cannot be in the future."}, status=400)
                profile.birthday = parsed_birthday
            else:
                return JsonResponse({"error": "Invalid birthday format. Use YYYY-MM-DD."}, status=400)

        profile.save()
        return JsonResponse(
            {"profile": profile.to_dict(), "user": _build_user_dict(request.user)}
        )


# ── Excel Export View ──────────────────────────────────────────────────────────


@login_required
def export_excel(request):
    """Generate and download the full Balance tracker Excel workbook."""
    from .excel_export import generate_excel
    from datetime import date

    buf = generate_excel()
    filename = f"Balance_Tracker_{date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ════════════════════════════════════════════════════════════════════════════
# REMINDER ENGINE VIEWS
# ════════════════════════════════════════════════════════════════════════════

from .models import (
    ReminderRule,
    CertificateStatus,
    ReminderLog,
    REMINDER_TYPE_CHOICES,
    SALARY_TRIGGER_CHOICES,
)


@method_decorator(csrf_exempt, name="dispatch")
class ReminderRuleListView(View):
    def get(self, request):
        rules = ReminderRule.objects.all()
        return JsonResponse(
            {
                "rules": [r.to_dict() for r in rules],
                "rule_types": [
                    {"value": v, "label": l} for v, l in REMINDER_TYPE_CHOICES
                ],
                "salary_triggers": [
                    {"value": v, "label": l} for v, l in SALARY_TRIGGER_CHOICES
                ],
            }
        )

    def post(self, request):
        data = json.loads(request.body)
        rule = ReminderRule.objects.create(
            name=data["name"],
            rule_type=data.get("rule_type", "cert_maturity"),
            is_active=data.get("is_active", True),
            days_before=int(data.get("days_before", 30)),
            salary_trigger=data.get("salary_trigger", "day_of_month"),
            salary_day=int(data.get("salary_day", 25)),
            salary_message=data.get("salary_message", ""),
        )
        return JsonResponse({"rule": rule.to_dict()}, status=201)


@method_decorator(csrf_exempt, name="dispatch")
class ReminderRuleDetailView(View):
    def put(self, request, pk):
        rule = get_object_or_404(ReminderRule, pk=pk)
        data = json.loads(request.body)
        rule.name = data.get("name", rule.name)
        rule.rule_type = data.get("rule_type", rule.rule_type)
        rule.is_active = data.get("is_active", rule.is_active)
        rule.days_before = int(data.get("days_before", rule.days_before))
        rule.salary_trigger = data.get("salary_trigger", rule.salary_trigger)
        rule.salary_day = int(data.get("salary_day", rule.salary_day))
        rule.salary_message = data.get("salary_message", rule.salary_message)
        rule.save()
        return JsonResponse({"rule": rule.to_dict()})

    def delete(self, request, pk):
        rule = get_object_or_404(ReminderRule, pk=pk)
        rule.delete()
        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class ReminderCheckView(View):
    """Called on page load — evaluates all active rules and returns due reminders."""

    def get(self, request):
        result = ReminderAutomationService().evaluate(today=timezone.localdate()).to_dict()
        return JsonResponse(result)


@method_decorator(csrf_exempt, name="dispatch")
class FixedAssetValuationRefreshView(View):
    def post(self, request, pk):
        asset = get_object_or_404(FixedAsset, pk=pk)
        updated, provider_name = PropertyValuationService().refresh_asset(asset, today=timezone.localdate())
        asset.refresh_from_db()
        return JsonResponse(
            {
                "updated": updated,
                "provider": provider_name,
                "asset": asset.to_dict(),
            }
        )


def _salary_trigger_day(rule, today):
    """Compute the calendar day this rule fires on for the given month."""
    import calendar as cal

    last_day = cal.monthrange(today.year, today.month)[1]
    if rule.salary_trigger == "day_of_month":
        return min(rule.salary_day, last_day)
    elif rule.salary_trigger == "days_before_eom":
        return max(1, last_day - rule.salary_day)
    elif rule.salary_trigger == "days_after_som":
        return min(rule.salary_day + 1, last_day)
    return rule.salary_day


@method_decorator(csrf_exempt, name="dispatch")
class ReminderLogListView(View):
    """Return recent reminder log entries."""

    def get(self, request):
        limit = int(request.GET.get("limit", 30))
        logs = ReminderLog.objects.select_related("rule").all()[:limit]
        return JsonResponse({"logs": [l.to_dict() for l in logs]})

    def delete(self, request):
        """Clear all log entries (reset fired state)."""
        ReminderLog.objects.all().delete()
        return JsonResponse({"cleared": True})


# ════════════════════════════════════════════════════════════════════════════
# CERTIFICATE STATUS VIEWS
# ════════════════════════════════════════════════════════════════════════════


@method_decorator(csrf_exempt, name="dispatch")
class CertificateStatusListView(View):
    def get(self, request):
        statuses = CertificateStatus.objects.all()
        return JsonResponse({"statuses": [s.to_dict() for s in statuses]})

    def post(self, request):
        data = json.loads(request.body)
        # If new status is default, unset any existing default
        if data.get("is_default"):
            CertificateStatus.objects.filter(is_default=True).update(is_default=False)
        s = CertificateStatus.objects.create(
            name=data["name"],
            color_hex=data.get("color_hex", "#1a6ef5"),
            is_default=data.get("is_default", False),
            is_terminal=data.get("is_terminal", False),
            order=int(data.get("order", 0)),
        )
        return JsonResponse({"status": s.to_dict()}, status=201)


@method_decorator(csrf_exempt, name="dispatch")
class CertificateStatusDetailView(View):
    def put(self, request, pk):
        s = get_object_or_404(CertificateStatus, pk=pk)
        data = json.loads(request.body)
        if data.get("is_default") and not s.is_default:
            CertificateStatus.objects.filter(is_default=True).update(is_default=False)
        s.name = data.get("name", s.name)
        s.color_hex = data.get("color_hex", s.color_hex)
        s.is_default = data.get("is_default", s.is_default)
        s.is_terminal = data.get("is_terminal", s.is_terminal)
        s.order = int(data.get("order", s.order))
        s.save()
        return JsonResponse({"status": s.to_dict()})

    def delete(self, request, pk):
        s = get_object_or_404(CertificateStatus, pk=pk)
        s.delete()
        return JsonResponse({"deleted": pk})


# ════════════════════════════════════════════════════════════════════════════
# ADVANCED REPORTS VIEWS
# ════════════════════════════════════════════════════════════════════════════


@method_decorator(csrf_exempt, name="dispatch")
class SalaryReportView(View):
    """Salary + bonus analytics by year and company."""

    def get(self, request):
        from django.db.models import Sum, Count, Q

        year = request.GET.get("year")
        company_id = request.GET.get("company_id")

        qs = SalaryEntry.objects.all()
        if year:
            qs = qs.filter(year=int(year))
        if company_id:
            qs = qs.filter(company_id=int(company_id))

        # By year
        by_year = list(
            qs.values("year")
            .annotate(
                total_paid=Sum("paid"),
                total_bonus=Sum("bonus"),
                total_expected=Sum("expected"),
                paid_months=Count("id", filter=Q(paid__gt=0)),
            )
            .order_by("year")
        )

        # By company
        by_company = []
        for c in Company.objects.all().order_by("order"):
            cqs = qs.filter(company=c)
            agg = cqs.aggregate(
                total_paid=Sum("paid"),
                total_bonus=Sum("bonus"),
                total_expected=Sum("expected"),
                paid_months=Count("id", filter=Q(paid__gt=0)),
            )
            if agg["paid_months"]:
                by_company.append(
                    {
                        "company_id": c.id,
                        "company_name": c.display_name or c.name,
                        "color_hex": c.color_hex,
                        "total_paid": float(agg["total_paid"] or 0),
                        "total_bonus": float(agg["total_bonus"] or 0),
                        "total_expected": float(agg["total_expected"] or 0),
                        "paid_months": agg["paid_months"] or 0,
                    }
                )

        # Grand totals
        grand = qs.aggregate(
            total_paid=Sum("paid"),
            total_bonus=Sum("bonus"),
            total_expected=Sum("expected"),
            paid_months=Count("id", filter=Q(paid__gt=0)),
        )

        # Available years
        years = list(
            SalaryEntry.objects.values_list("year", flat=True)
            .distinct()
            .order_by("year")
        )
        companies = [
            {"id": c.id, "name": c.display_name or c.name}
            for c in Company.objects.all().order_by("order")
        ]

        return JsonResponse(
            {
                "by_year": by_year,
                "by_company": by_company,
                "grand": {
                    "total_paid": float(grand["total_paid"] or 0),
                    "total_bonus": float(grand["total_bonus"] or 0),
                    "total_expected": float(grand["total_expected"] or 0),
                    "paid_months": grand["paid_months"] or 0,
                },
                "years": years,
                "companies": companies,
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class BalanceReportView(View):
    """Balance summary across banks and currencies."""

    def get(self, request):
        _run_certificate_interest_sync()
        from django.db.models import Sum

        entries = BalanceEntry.objects.select_related("bank", "currency").all()
        banks = Bank.objects.all()

        # Group by bank
        by_bank = []
        for bank in banks:
            bank_entries = entries.filter(bank=bank)
            total_egp = float(
                bank_entries.filter(currency__code="EGP").aggregate(s=Sum("amount"))[
                    "s"
                ]
                or 0
            )
            by_bank.append(
                {
                    "bank_id": bank.id,
                    "bank_name": bank.name,
                    "total_egp": total_egp,
                    "entries": [e.to_dict() for e in bank_entries],
                }
            )

        # Unbanked entries (cash / home)
        home = entries.filter(bank__isnull=True)
        by_currency = []
        for e in home:
            by_currency.append(e.to_dict())

        net_worth_data = NetWorthService().portfolio_components()
        cert_total = float(net_worth_data["certificate_total_egp"])
        cert_interest_total = float(net_worth_data["certificate_interest_total_egp"])

        cert_monthly_interest = cert_interest_total if cert_interest_total else 0.0

        return JsonResponse(
            {
                "by_bank": by_bank,
                "home_entries": by_currency,
                "cert_total": cert_total,
                "cert_interest": cert_monthly_interest,
                "cert_interest_total": cert_interest_total,
                "fixed_assets_total": float(net_worth_data["fixed_assets_total_egp"]),
                "net_worth": float(net_worth_data["net_worth_egp"]),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class CertificateReportView(View):
    """Certificate maturity and analytics report."""

    def get(self, request):
        from datetime import date, timedelta
        from django.db.models import Sum, Count

        today = date.today()
        active_certs = BankCertificate.objects.select_related("bank", "currency").filter(
            status__iexact="active"
        )

        agg = active_certs.aggregate(
            total_count=Count("id"),
            total_amount=Sum("amount"),
            total_interest=Sum("interest_value"),
        )

        # Maturity buckets (configurable label from settings, days from AppSettings)
        bucket_days = [
            ("overdue", 0, -1),
            ("30_days", 0, 30),
            ("90_days", 31, 90),
            ("180_days", 91, 180),
            ("later", 181, 9999),
        ]

        buckets = {}
        for label, low, high in bucket_days:
            if label == "overdue":
                buckets[label] = [
                    c.to_dict()
                    for c in active_certs.filter(expiry_date__lt=today)
                ]
            else:
                buckets[label] = [
                    c.to_dict()
                    for c in active_certs.filter(
                        expiry_date__gte=today + timedelta(days=low),
                        expiry_date__lte=today + timedelta(days=high),
                    )
                ]

        # By status (active certificates only)
        by_status = {}
        for c in active_certs:
            by_status[c.status] = by_status.get(c.status, {"count": 0, "total": 0})
            by_status[c.status]["count"] += 1
            by_status[c.status]["total"] += float(c.amount)

        # Monthly interest cashflow (next 12 months) for active certificates only
        monthly_cf = []
        for i in range(12):
            m_start = today.replace(day=1) + timedelta(days=32 * i)
            m_start = m_start.replace(day=1)
            m_certs = active_certs.filter(
                expiry_date__year=m_start.year,
                expiry_date__month=m_start.month,
            )
            monthly_cf.append(
                {
                    "month": m_start.strftime("%b %Y"),
                    "count": m_certs.count(),
                    "amount": float(m_certs.aggregate(s=Sum("amount"))["s"] or 0),
                }
            )

        total_interest = float(agg["total_interest"] or 0)
        # Keep monthly interest aligned with historical report expectations.
        # VS Code, don't this again to (total_interest / 12.0)
        monthly_interest = (total_interest) if total_interest else 0.0

        return JsonResponse(
            {
                "summary": {
                    "total_count": agg["total_count"] or 0,
                    "total_amount": float(agg["total_amount"] or 0),
                    "total_interest": total_interest,
                    "monthly_interest": monthly_interest,
                },
                "buckets": buckets,
                "by_status": by_status,
                "monthly_cf": monthly_cf,
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DashboardSummaryView(View):
    """Enhanced dashboard summary — salary KPIs + cert maturity + balance."""

    def get(self, request):
        _run_certificate_interest_sync()
        from datetime import date, timedelta
        from django.db.models import Sum, Count, Q

        today = date.today()

        # Salary grand totals
        sal_agg = SalaryEntry.objects.aggregate(
            total_paid=Sum("paid"),
            total_bonus=Sum("bonus"),
            total_expected=Sum("expected"),
            paid_months=Count("id", filter=Q(paid__gt=0)),
        )

        # Certificates
        certs = BankCertificate.objects.select_related("bank").all()
        cert_agg = certs.aggregate(
            total=Sum("amount"),
            total_interest=Sum("interest_value"),
        )
        expiring_soon_days = int(AppSettings.get("cert_expiry_warning_days", "30"))
        expiring_soon = list(
            certs.filter(
                expiry_date__gte=today,
                expiry_date__lte=today + timedelta(days=expiring_soon_days),
            )
            .order_by("expiry_date")
            .values("id", "expiry_date", "amount", "status", "bank__name")
        )
        for e in expiring_soon:
            e["days_left"] = (e["expiry_date"] - today).days
            e["expiry_date"] = e["expiry_date"].isoformat()

        # Active reminders (due today)
        active_reminders = []
        for rule in ReminderRule.objects.filter(
            is_active=True, rule_type="cert_maturity"
        ):
            logs = ReminderLog.objects.filter(rule=rule, fired_on=today).values(
                "message", "related_id", "related_model"
            )
            for l in logs:
                active_reminders.append({"rule": rule.name, "message": l["message"]})

        # Balance
        bal_entries = BalanceEntry.objects.select_related("currency").all()
        egp_balance = float(
            bal_entries.filter(currency__code="EGP").aggregate(s=Sum("amount"))["s"]
            or 0
        )

        net_worth = NetWorthService().portfolio_components()

        return JsonResponse(
            {
                "salary": {
                    "total_paid": float(sal_agg["total_paid"] or 0),
                    "total_bonus": float(sal_agg["total_bonus"] or 0),
                    "total_expected": float(sal_agg["total_expected"] or 0),
                    "paid_months": sal_agg["paid_months"] or 0,
                },
                "certificates": {
                    "total_amount": float(cert_agg["total"] or 0),
                    "total_interest": float(cert_agg["total_interest"] or 0),
                    "monthly_interest": float(cert_agg["total_interest"] or 0),
                    "count": certs.count(),
                },
                "expiring_soon": expiring_soon,
                "active_reminders": active_reminders,
                "egp_balance": egp_balance,
                "expiry_warning_days": expiring_soon_days,
                "net_worth": {
                    "total": float(net_worth["net_worth_egp"]),
                    "liquid_assets": float(net_worth["liquid_assets_total_egp"]),
                    "fixed_assets": float(net_worth["fixed_assets_total_egp"]),
                },
            }
        )


from datetime import date, timedelta
from core.models import SalaryEntry


@method_decorator(csrf_exempt, name="dispatch")
class CertificateForecastView(View):
    def get(self, request):
        _run_certificate_interest_sync()
        return JsonResponse(NetWorthService().certificate_forecast_payload(today=date.today()))


@method_decorator(csrf_exempt, name="dispatch")
class CashFlowForecastView(View):
    def get(self, request):
        auth_error = _api_auth_required(request)
        if auth_error:
            return auth_error

        _run_certificate_interest_sync()
        payload = CashFlowForecastService(today=date.today()).payload()
        return JsonResponse(payload)


@method_decorator(csrf_exempt, name="dispatch")
class WealthGrowthForecastView(View):
    def get(self, request):
        auth_error = _api_auth_required(request)
        if auth_error:
            return auth_error

        _run_certificate_interest_sync()
        payload = WealthGrowthForecastService(today=date.today()).payload()
        return JsonResponse(payload)

# ============================================================
# Fixed Assets APIs
# ============================================================

REAL_ESTATE_ASSET_TYPES = {"Real Estate"}
VEHICLE_ASSET_TYPES = {"Vehicles"}
GOLD_ASSET_TYPES = {"Gold"}
OTHER_ASSET_TYPES = {"Other Assets"}

ASSET_PAYMENT_METHOD_CASH = "Cash"
ASSET_PAYMENT_METHOD_CARD = "Card"
ASSET_PAYMENT_METHOD_BANK = "Bank"
ASSET_PAYMENT_METHOD_BANK_TRANSFER = "Bank Transfer"

ASSET_PAYMENT_METHOD_NORMALIZED = {
    "cash": ASSET_PAYMENT_METHOD_CASH,
    "card": ASSET_PAYMENT_METHOD_CARD,
    "bank": ASSET_PAYMENT_METHOD_BANK,
    "bank transfer": ASSET_PAYMENT_METHOD_BANK_TRANSFER,
    "bank_transfer": ASSET_PAYMENT_METHOD_BANK_TRANSFER,
}

GOLD_UNIT_TO_GRAMS = {
    "g": Decimal("1"),
    "gm": Decimal("1"),
    "gram": Decimal("1"),
    "grams": Decimal("1"),
    "kg": Decimal("1000"),
    "kilogram": Decimal("1000"),
    "kilograms": Decimal("1000"),
    "oz": Decimal("31.1034768"),
    "ounce": Decimal("31.1034768"),
    "ounces": Decimal("31.1034768"),
    "tola": Decimal("11.6638038"),
}


def _to_decimal(value, default="0"):
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def _gold_unit_factor(unit_value):
    normalized = str(unit_value or "gram").strip().lower()
    return GOLD_UNIT_TO_GRAMS.get(normalized, Decimal("1"))


def _gold_weight_in_grams(weight_value, unit_value):
    return _to_decimal(weight_value) * _gold_unit_factor(unit_value)


def _normalize_gold_purity(purity_value):
    text = str(purity_value or "").strip().lower()
    if "24" in text or "999" in text:
        return "24k"
    if "22" in text or "916" in text:
        return "22k"
    if "21" in text or "875" in text:
        return "21k"
    if "18" in text or "750" in text:
        return "18k"
    return "24k"


def _gold_sell_price_per_gram(latest_gold_price, purity_key):
    price_map = {
        "24k": _to_decimal(latest_gold_price.carat_24k),
        "22k": _to_decimal(latest_gold_price.carat_22k),
        "21k": _to_decimal(latest_gold_price.carat_21k),
        "18k": _to_decimal(latest_gold_price.carat_18k),
    }
    return price_map.get(purity_key, price_map["24k"])


def _gold_cashback_per_gram(purity_value):
    key = _normalize_gold_purity(purity_value)
    setting = GoldPuritySetting.objects.filter(key=key, is_active=True).first()
    if not setting:
        return Decimal("0")
    return _to_decimal(setting.cashback_per_gram)


def _latest_gold_price():
    return GoldPrice.objects.order_by("-fetched_at").first()


def _refresh_gold_asset_pricing(asset, gold_details=None, latest_gold_price=None):
    if asset.asset_type not in GOLD_ASSET_TYPES:
        return

    details = gold_details
    if details is None:
        details = getattr(asset, "gold_details", None)
    if details is None:
        return

    latest_gold = latest_gold_price or _latest_gold_price()
    if latest_gold is None:
        return

    usd_to_egp = _to_decimal(latest_gold.usd_to_egp)
    if usd_to_egp > 0:
        asset.purchase_usd_rate = usd_to_egp
        asset.purchase_price_usd = _to_decimal(asset.purchase_price) / usd_to_egp

    purity_key = _normalize_gold_purity(details.purity)
    sell_price_per_gram = _gold_sell_price_per_gram(latest_gold, purity_key)
    unit_factor = _gold_unit_factor(details.unit)
    details.market_price = sell_price_per_gram * unit_factor

    cashback_per_gram = _gold_cashback_per_gram(details.purity)
    details.cashback_per_gram = cashback_per_gram
    total_weight_grams = _gold_weight_in_grams(details.weight, details.unit)
    asset.current_market_value = total_weight_grams * (sell_price_per_gram + cashback_per_gram)
    asset.valuation_source = "Automatic"
    asset.last_valuation_date = timezone.now().date()

    details.save(update_fields=["market_price", "updated_at"])
    asset.save(
        update_fields=[
            "purchase_usd_rate",
            "purchase_price_usd",
            "current_market_value",
            "valuation_source",
            "last_valuation_date",
            "updated_at",
        ]
    )


def _sync_gold_balance_from_assets():
    gold_currency = Currency.objects.filter(code__iexact="gold").first()
    if not gold_currency:
        return

    gold_assets = (
        FixedAsset.objects.filter(asset_type__in=GOLD_ASSET_TYPES, status="Owned")
        .select_related("gold_details")
        .order_by("id")
    )

    grams_by_purity = {}
    for asset in gold_assets:
        details = getattr(asset, "gold_details", None)
        if details is None:
            continue
        grams = _gold_weight_in_grams(details.weight, details.unit)
        purity_key = _normalize_gold_purity(details.purity)
        grams_by_purity[purity_key] = grams_by_purity.get(purity_key, Decimal("0")) + grams

    balance_qs = BalanceEntry.objects.filter(
        balance_type=BalanceEntry.BalanceType.GOLD,
        currency_id=gold_currency.id,
    ).order_by("id")

    if not grams_by_purity:
        balance_qs.delete()
        return

    existing_by_purity = {str(e.purity or "").lower(): e for e in balance_qs}
    used_ids = []
    for purity_key, grams in grams_by_purity.items():
        entry = existing_by_purity.get(purity_key)
        title = f"{gold_currency.name or 'Gold'} {purity_key.upper()}"
        amount = grams.quantize(Decimal("0.01"))

        if entry:
            entry.title = title
            entry.bank = None
            entry.amount = amount
            entry.notes = ""
            entry.purity = purity_key
            entry.save()
            used_ids.append(entry.id)
        else:
            created = BalanceEntry.objects.create(
                title=title,
                balance_type=BalanceEntry.BalanceType.GOLD,
                bank=None,
                currency_id=gold_currency.id,
                purity=purity_key,
                amount=amount,
                notes="",
            )
            used_ids.append(created.id)

    balance_qs.exclude(id__in=used_ids).delete()


def _refresh_all_gold_assets_from_live_prices():
    latest_gold = _latest_gold_price()
    if latest_gold is None:
        return

    gold_assets = FixedAsset.objects.filter(asset_type__in=GOLD_ASSET_TYPES).select_related("gold_details")
    for asset in gold_assets:
        details = getattr(asset, "gold_details", None)
        if details is None:
            continue
        _refresh_gold_asset_pricing(asset, details, latest_gold)

    _sync_gold_balance_from_assets()


def _clear_non_selected_asset_details(asset):
    if asset.asset_type not in REAL_ESTATE_ASSET_TYPES and hasattr(asset, "real_estate"):
        asset.real_estate.delete()

    if asset.asset_type not in VEHICLE_ASSET_TYPES and hasattr(asset, "vehicle_details"):
        asset.vehicle_details.delete()

    if asset.asset_type not in GOLD_ASSET_TYPES and hasattr(asset, "gold_details"):
        asset.gold_details.delete()

    if asset.asset_type not in OTHER_ASSET_TYPES and hasattr(asset, "other_asset_details"):
        asset.other_asset_details.delete()


def _sync_vehicle_details(asset, details_data):
    if asset.asset_type not in VEHICLE_ASSET_TYPES or not details_data:
        if hasattr(asset, "vehicle_details"):
            asset.vehicle_details.delete()
        return

    VehicleDetails.objects.update_or_create(
        asset=asset,
        defaults={
            "brand": details_data.get("brand", ""),
            "model": details_data.get("model", ""),
            "year": details_data.get("year") or None,
            "vin": details_data.get("vin", ""),
            "engine": details_data.get("engine", ""),
            "transmission": details_data.get("transmission", ""),
            "fuel_type": details_data.get("fuel_type", ""),
            "mileage": details_data.get("mileage", 0),
            "plate_number": details_data.get("plate_number", ""),
            "license_expiry_date": _parse_iso_date(details_data.get("license_expiry_date")),
            "color": details_data.get("color", ""),
        },
    )


def _sync_gold_details(asset, details_data):
    if asset.asset_type not in GOLD_ASSET_TYPES or not details_data:
        if hasattr(asset, "gold_details"):
            asset.gold_details.delete()
        return

    details_obj, _ = GoldDetails.objects.update_or_create(
        asset=asset,
        defaults={
            "gold_type": details_data.get("gold_type", ""),
            "purity": _normalize_gold_purity(details_data.get("purity", "")),
            "weight": details_data.get("weight", 0),
            "unit": details_data.get("unit", "gram"),
            "cashback_per_gram": _gold_cashback_per_gram(details_data.get("purity", "")),
            "purchase_weight": details_data.get("purchase_weight", 0),
        },
    )

    _refresh_gold_asset_pricing(asset, details_obj)


def _sync_other_asset_details(asset, details_data):
    if asset.asset_type not in OTHER_ASSET_TYPES or not details_data:
        if hasattr(asset, "other_asset_details"):
            asset.other_asset_details.delete()
        return

    OtherAssetDetails.objects.update_or_create(
        asset=asset,
        defaults={
            "category": details_data.get("category", ""),
            "manufacturer": details_data.get("manufacturer", ""),
            "model": details_data.get("model", ""),
            "serial_number": details_data.get("serial_number", ""),
            "description": details_data.get("description", ""),
            "warranty_expiry": details_data.get("warranty_expiry") or None,
            "notes": details_data.get("notes", ""),
        },
    )


def _sync_asset_maintenance(asset, items):
    AssetMaintenance.objects.filter(asset=asset).delete()
    if asset.asset_type not in VEHICLE_ASSET_TYPES:
        return

    for item in items or []:
        if not item.get("date"):
            continue
        AssetMaintenance.objects.create(
            asset=asset,
            date=item.get("date"),
            maintenance_type=item.get("type", ""),
            cost=item.get("cost", 0),
            notes=item.get("notes", ""),
        )


def _sync_asset_insurance(asset, items):
    AssetInsurance.objects.filter(asset=asset).delete()
    if asset.asset_type not in VEHICLE_ASSET_TYPES:
        return

    for item in items or []:
        if not item.get("company"):
            continue
        AssetInsurance.objects.create(
            asset=asset,
            company=item.get("company", ""),
            policy_number=item.get("policy_number", ""),
            expiry_date=item.get("expiry_date") or None,
            premium=item.get("premium", 0),
        )


def _sync_asset_mortgage(asset, mortgage_data):
    if asset.asset_type not in REAL_ESTATE_ASSET_TYPES or not mortgage_data:
        if hasattr(asset, "mortgage"):
            asset.mortgage.delete()
        return

    has_values = any(
        mortgage_data.get(key) not in (None, "", 0, 0.0)
        for key in [
            "loan_amount",
            "remaining_balance",
            "monthly_installment",
            "interest_rate",
            "start_date",
            "end_date",
        ]
    )

    if not has_values:
        if hasattr(asset, "mortgage"):
            asset.mortgage.delete()
        return

    AssetMortgage.objects.update_or_create(
        asset=asset,
        defaults={
            "loan_amount": mortgage_data.get("loan_amount", 0),
            "remaining_balance": mortgage_data.get("remaining_balance", 0),
            "monthly_installment": mortgage_data.get("monthly_installment", 0),
            "interest_rate": mortgage_data.get("interest_rate", 0),
            "start_date": _parse_iso_date(mortgage_data.get("start_date")),
            "end_date": _parse_iso_date(mortgage_data.get("end_date")),
        },
    )


def _sync_asset_rental(asset, rental_data):
    if asset.asset_type not in REAL_ESTATE_ASSET_TYPES or not rental_data:
        if hasattr(asset, "rental"):
            asset.rental.delete()
        return

    has_values = any(
        rental_data.get(key) not in (None, "", 0, 0.0)
        for key in [
            "monthly_rent",
            "occupancy_rate",
            "tenant_name",
            "contract_start",
            "contract_end",
            "notes",
        ]
    )

    if not has_values:
        if hasattr(asset, "rental"):
            asset.rental.delete()
        return

    AssetRental.objects.update_or_create(
        asset=asset,
        defaults={
            "monthly_rent": rental_data.get("monthly_rent", 0),
            "occupancy_rate": rental_data.get("occupancy_rate", 0),
            "tenant_name": rental_data.get("tenant_name", ""),
            "contract_start": _parse_iso_date(rental_data.get("contract_start")),
            "contract_end": _parse_iso_date(rental_data.get("contract_end")),
            "notes": rental_data.get("notes", ""),
        },
    )


def _sync_asset_furniture(asset, items):
    if asset.asset_type not in REAL_ESTATE_ASSET_TYPES:
        AssetFurniture.objects.filter(asset=asset).delete()
        return

    AssetFurniture.objects.filter(asset=asset).delete()
    for item in items or []:
        if not item.get("name"):
            continue
        AssetFurniture.objects.create(
            asset=asset,
            name=item.get("name", ""),
            category=item.get("category", ""),
            purchase_date=item.get("purchase_date") or None,
            amount_egp=item.get("amount_egp", 0),
            usd_rate=item.get("usd_rate", 0),
            amount_usd=item.get("amount_usd", 0),
            quantity=item.get("quantity", 1),
            notes=item.get("notes", ""),
        )


def _sync_asset_valuation_history(asset, items):
    if asset.asset_type not in REAL_ESTATE_ASSET_TYPES and asset.asset_type not in VEHICLE_ASSET_TYPES and asset.asset_type not in OTHER_ASSET_TYPES:
        AssetValuationHistory.objects.filter(asset=asset).delete()
        return

    AssetValuationHistory.objects.filter(asset=asset).delete()
    created_items = []
    for item in items or []:
        if not item.get("valuation_date"):
            continue
        created_items.append(
            AssetValuationHistory.objects.create(
                asset=asset,
                valuation_date=item.get("valuation_date"),
                market_value=item.get("market_value", 0),
                valuation_source=item.get("valuation_source", "Manual"),
                notes=item.get("notes", ""),
            )
        )

    if created_items:
        latest_item = max(created_items, key=lambda value: value.valuation_date)
        asset.current_market_value = latest_item.market_value
        asset.last_valuation_date = latest_item.valuation_date
        asset.valuation_source = latest_item.valuation_source
        asset.save()


def _normalize_asset_payment_method(method_value):
    normalized = str(method_value or "").strip().lower()
    return ASSET_PAYMENT_METHOD_NORMALIZED.get(normalized, ASSET_PAYMENT_METHOD_CASH)


def _asset_payment_requires_bank(method_value):
    return _normalize_asset_payment_method(method_value) != ASSET_PAYMENT_METHOD_CASH


def _asset_payment_currency_required(currency_id):
    return currency_id is not None and str(currency_id).strip() != ""


def _default_egp_currency_id():
    currency = Currency.objects.filter(code__iexact="EGP").order_by("id").first()
    return currency.id if currency else None


def _normalize_purchase_payments_payload(rows, purchase_price, purchase_currency_id=None, allow_empty=False):
    normalized_rows = []
    running_total = Decimal("0")
    resolved_currency_id = purchase_currency_id

    if rows is None:
        rows = []

    if not isinstance(rows, list):
        raise ValueError("purchase_payments_invalid")

    for row in rows:
        if not isinstance(row, dict):
            raise ValueError("purchase_payments_invalid")

        row_currency_id = row.get("currency_id")
        if not _asset_payment_currency_required(resolved_currency_id) and _asset_payment_currency_required(row_currency_id):
            resolved_currency_id = row_currency_id

        payment_method = _normalize_asset_payment_method(row.get("payment_method"))
        bank_id = row.get("bank_id")
        if _asset_payment_requires_bank(payment_method) and not bank_id:
            raise ValueError("bank_account_required")
        if payment_method == ASSET_PAYMENT_METHOD_CASH:
            bank_id = None

        amount = _to_decimal(row.get("amount"), default="0")
        if amount <= 0:
            raise ValueError("amount_required")

        normalized_rows.append(
            {
                "currency_id": None,
                "payment_method": payment_method,
                "bank_id": int(bank_id) if bank_id else None,
                "amount": amount,
            }
        )
        running_total += amount

    if not _asset_payment_currency_required(resolved_currency_id):
        if allow_empty and not normalized_rows:
            return []
        resolved_currency_id = _default_egp_currency_id()
        if not resolved_currency_id:
            raise ValueError("currency_required")

    resolved_currency_id = int(resolved_currency_id)
    for row in normalized_rows:
        row["currency_id"] = resolved_currency_id

    if not normalized_rows:
        if allow_empty:
            return []

        return [
            {
                "currency_id": resolved_currency_id,
                "payment_method": ASSET_PAYMENT_METHOD_CASH,
                "bank_id": None,
                "amount": _to_decimal(purchase_price),
            }
        ]

    target_total = _to_decimal(purchase_price)
    if running_total.quantize(Decimal("0.01")) != target_total.quantize(Decimal("0.01")):
        raise ValueError("purchase_payment_total_mismatch")

    return normalized_rows


def _get_asset_cash_balance_entry(currency_id, bank_id):
    qs = BalanceEntry.objects.select_for_update().filter(
        balance_type=BalanceEntry.BalanceType.CASH,
        currency_id=currency_id,
    )
    if bank_id:
        qs = qs.filter(bank_id=bank_id)
    else:
        qs = qs.filter(bank__isnull=True)
    return qs.order_by("id").first()


def _apply_asset_balance_delta(currency_id, payment_method, bank_id, amount_delta):
    delta = _to_decimal(amount_delta)
    if delta == 0:
        return

    resolved_method = _normalize_asset_payment_method(payment_method)
    resolved_bank_id = bank_id if _asset_payment_requires_bank(resolved_method) else None

    entry = _get_asset_cash_balance_entry(currency_id, resolved_bank_id)
    if not entry:
        raise ValueError("matching_balance_entry_not_found")

    next_amount = _to_decimal(entry.amount) + delta
    if next_amount < 0:
        raise ValueError("insufficient_balance")

    entry.amount = next_amount
    entry.save(update_fields=["amount"])


def _apply_asset_purchase_rows_delta(rows, sign):
    sign_multiplier = Decimal("1") if sign >= 0 else Decimal("-1")
    for row in rows:
        _apply_asset_balance_delta(
            currency_id=row.get("currency_id"),
            payment_method=row.get("payment_method"),
            bank_id=row.get("bank_id"),
            amount_delta=sign_multiplier * _to_decimal(row.get("amount")),
        )


def _purchase_rows_from_instances(instances):
    return [
        {
            "currency_id": item.currency_id,
            "payment_method": item.payment_method,
            "bank_id": item.bank_id,
            "amount": _to_decimal(item.amount),
        }
        for item in instances
    ]


def _sync_asset_purchase_payments(asset, rows):
    AssetPurchasePayment.objects.filter(asset=asset).delete()
    for row in rows:
        AssetPurchasePayment.objects.create(
            asset=asset,
            currency_id=row["currency_id"],
            payment_method=row["payment_method"],
            bank_id=row.get("bank_id"),
            amount=row["amount"],
        )


def _resolve_sale_deposit_values(data, existing_sale=None):
    fallback_currency_id = _default_egp_currency_id()

    existing_currency_id = existing_sale.deposit_currency_id if existing_sale else None
    currency_id = data.get("deposit_currency_id", existing_currency_id or fallback_currency_id)
    if not _asset_payment_currency_required(currency_id):
        raise ValueError("currency_required")

    existing_method = existing_sale.deposit_method if existing_sale else ASSET_PAYMENT_METHOD_CASH
    method = _normalize_asset_payment_method(data.get("deposit_method", existing_method))

    existing_bank_id = existing_sale.deposit_bank_id if existing_sale else None
    bank_id = data.get("deposit_bank_id", existing_bank_id)
    if _asset_payment_requires_bank(method) and not bank_id:
        raise ValueError("bank_account_required")
    if method == ASSET_PAYMENT_METHOD_CASH:
        bank_id = None

    return {
        "deposit_currency_id": int(currency_id),
        "deposit_method": method,
        "deposit_bank_id": int(bank_id) if bank_id else None,
    }


def _sale_payment_row(sale):
    currency_id = sale.deposit_currency_id or _default_egp_currency_id()
    method = _normalize_asset_payment_method(sale.deposit_method or ASSET_PAYMENT_METHOD_CASH)
    bank_id = sale.deposit_bank_id if _asset_payment_requires_bank(method) else None
    return {
        "currency_id": currency_id,
        "payment_method": method,
        "bank_id": bank_id,
        "amount": _to_decimal(sale.net_sale_amount),
    }


@method_decorator(csrf_exempt, name="dispatch")
class FixedAssetListView(View):

    def get(self, request):
        qs = FixedAsset.objects.all().order_by("name")

        asset_type = request.GET.get("asset_type")
        status = request.GET.get("status")

        if asset_type:
            qs = qs.filter(asset_type=asset_type)

        if status:
            qs = qs.filter(status=status)

        service = NetWorthService()
        return JsonResponse(
            {
                "assets": [a.to_dict() for a in qs],
                "portfolio_snapshot": service.fixed_assets_snapshot(),
            }
        )

    def post(self, request):
        data = json.loads(request.body)
        re = data.get("real_estate_details")
        vehicle_details = data.get("vehicle_details")
        gold_details = data.get("gold_details")
        other_asset_details = data.get("other_asset_details")

        purchase_rows_raw = data.get("purchase_payments", [])

        try:
            with transaction.atomic():
                purchase_rows = _normalize_purchase_payments_payload(
                    purchase_rows_raw,
                    data.get("purchase_price", 0),
                    purchase_currency_id=data.get("purchase_currency_id"),
                )

                asset = FixedAsset.objects.create(
                    name=data["name"],
                    asset_type=data["asset_type"],
                    status=data.get("status", "Owned"),
                    purchase_date=data["purchase_date"],
                    purchase_price=data.get("purchase_price", 0),
                    purchase_usd_rate=data.get("purchase_usd_rate", 0),
                    purchase_price_usd=data.get("purchase_price_usd", 0),
                    current_market_value=data.get("current_market_value", 0),
                    valuation_source=data.get("valuation_source", "Manual"),
                    last_valuation_date=data.get("last_valuation_date") or None,
                    notes=data.get("notes", ""),
                )

                if re:
                    RealEstateDetails.objects.create(
                        asset=asset,
                        country=re.get("country", "Egypt"),
                        governorate=re.get("governorate", ""),
                        city=re.get("city", ""),
                        district=re.get("district", ""),
                        full_address=re.get("address", ""),
                        area_m2=re.get("apartment_area", 0),
                        bedrooms=re.get("rooms", 0),
                        bathrooms=re.get("bathrooms", 0),
                        floor_number=re.get("floor", 0),
                        building_floors=re.get("building_floors", 0),
                        build_year=re.get("building_year") or None,
                        facing=re.get("facades", ""),
                        finishing_level=re.get("finishing_level", ""),
                        electricity_meter_private=re.get("electricity", False),
                        water_meter_private=re.get("water", False),
                        has_gas=re.get("gas", False),
                        has_elevator=re.get("elevator", False),
                        has_garage=re.get("garage", False),
                        has_land_share=re.get("has_land_share", False),
                        land_share_ratio=re.get("land_share", ""),
                        land_share_sqm=float(re.get("land_share_sqm") or 0),
                        latitude=re.get("latitude") or None,
                        longitude=re.get("longitude") or None,
                        licensed=re.get("licensed", False),
                        description=re.get("description", ""),
                    )

                _sync_vehicle_details(asset, vehicle_details)
                _sync_gold_details(asset, gold_details)
                _sync_other_asset_details(asset, other_asset_details)

                _sync_asset_mortgage(asset, data.get("mortgage_details"))
                _sync_asset_rental(asset, data.get("rental_details"))

                for item in data.get("renovations", []):
                    if asset.asset_type not in REAL_ESTATE_ASSET_TYPES:
                        break

                    AssetRenovation.objects.create(
                        asset=asset,
                        date=item.get("date") or None,
                        category=item.get("category", ""),
                        description=item.get("description", ""),
                        amount_egp=item.get("amount_egp", 0),
                        usd_rate=item.get("usd_rate", 0),
                        amount_usd=item.get("amount_usd", 0),
                        notes=item.get("notes", ""),
                    )

                _sync_asset_maintenance(asset, data.get("maintenance", []))
                _sync_asset_insurance(asset, data.get("insurance", []))
                _sync_asset_furniture(asset, data.get("furniture", []))
                _sync_asset_valuation_history(asset, data.get("valuation_history", []))
                _clear_non_selected_asset_details(asset)

                _apply_asset_purchase_rows_delta(purchase_rows, sign=-1)
                _sync_asset_purchase_payments(asset, purchase_rows)

                _sync_gold_balance_from_assets()

        except ValueError as exc:
            return JsonResponse(
                {
                    "error": str(exc),
                    "error_key": str(exc),
                },
                status=400,
            )

        return JsonResponse(asset.to_dict(), status=201)
    
@method_decorator(csrf_exempt, name="dispatch")
class FixedAssetDetailView(View):

    def get(self, request, pk):
        asset = get_object_or_404(FixedAsset, pk=pk)
        return JsonResponse(asset.to_dict())

    def put(self, request, pk):
        asset = get_object_or_404(FixedAsset, pk=pk)

        data = json.loads(request.body)
        vehicle_details = data.get("vehicle_details")
        gold_details = data.get("gold_details")
        other_asset_details = data.get("other_asset_details")

        fields = [
            "name",
            "asset_type",
            "status",
            "purchase_date",
            "purchase_price",
            "purchase_usd_rate",
            "purchase_price_usd",
            "current_market_value",
            "valuation_source",
            "last_valuation_date",
            "notes",
        ]

        previous_rows = _purchase_rows_from_instances(
            AssetPurchasePayment.objects.filter(asset=asset).order_by("id")
        )

        purchase_rows_payload_present = "purchase_payments" in data
        purchase_rows_raw = data.get("purchase_payments", [])

        try:
            with transaction.atomic():
                if purchase_rows_payload_present:
                    allow_empty = len(previous_rows) == 0
                    purchase_rows = _normalize_purchase_payments_payload(
                        purchase_rows_raw,
                        data.get("purchase_price", asset.purchase_price),
                        purchase_currency_id=data.get("purchase_currency_id"),
                        allow_empty=allow_empty,
                    )
                else:
                    purchase_rows = previous_rows

                if previous_rows:
                    _apply_asset_purchase_rows_delta(previous_rows, sign=1)

                for field in fields:
                    if field in data:
                        setattr(asset, field, data[field])

                asset.save()

                re = data.get("real_estate_details")

                if re:
                    obj, _ = RealEstateDetails.objects.get_or_create(asset=asset)

                    obj.country = re.get("country", "Egypt")
                    obj.governorate = re.get("governorate", "")
                    obj.city = re.get("city", "")
                    obj.district = re.get("district", "")
                    obj.full_address = re.get("address", "")

                    obj.area_m2 = re.get("apartment_area", 0)

                    obj.bedrooms = re.get("rooms", 0)
                    obj.bathrooms = re.get("bathrooms", 0)

                    obj.floor_number = re.get("floor", 0)
                    obj.building_floors = re.get("building_floors", 0)
                    obj.build_year = re.get("building_year") or None

                    obj.facing = re.get("facades", "")

                    obj.furnished_status = re.get("furnished_status", "Unfurnished")
                    obj.finishing_level = re.get("finishing_level", "")

                    obj.electricity_meter_private = re.get("electricity", False)
                    obj.water_meter_private = re.get("water", False)
                    obj.has_gas = re.get("gas", False)

                    obj.has_elevator = re.get("elevator", False)
                    obj.has_garage = re.get("garage", False)
                    obj.has_land_share = re.get("has_land_share", False)
                    obj.land_share_ratio = re.get("land_share", "")
                    obj.land_share_sqm = float(re.get("land_share_sqm") or 0)
                    obj.latitude = re.get("latitude") or None
                    obj.longitude = re.get("longitude") or None
                    obj.licensed = re.get("licensed", False)
                    obj.description = re.get("description", "")

                    obj.save()
                elif asset.asset_type not in REAL_ESTATE_ASSET_TYPES and hasattr(asset, "real_estate"):
                    asset.real_estate.delete()

                _sync_vehicle_details(asset, vehicle_details)
                _sync_gold_details(asset, gold_details)
                _sync_other_asset_details(asset, other_asset_details)

                AssetRenovation.objects.filter(asset=asset).delete()

                for item in data.get("renovations", []):
                    if asset.asset_type not in REAL_ESTATE_ASSET_TYPES:
                        break

                    AssetRenovation.objects.create(
                        asset=asset,
                        date=item.get("date") or None,
                        category=item.get("category", ""),
                        description=item.get("description", ""),
                        amount_egp=item.get("amount_egp", 0),
                        usd_rate=item.get("usd_rate", 0),
                        amount_usd=item.get("amount_usd", 0),
                        notes=item.get("notes", ""),
                    )

                _sync_asset_maintenance(asset, data.get("maintenance", []))
                _sync_asset_insurance(asset, data.get("insurance", []))
                _sync_asset_mortgage(asset, data.get("mortgage_details"))
                _sync_asset_rental(asset, data.get("rental_details"))
                _sync_asset_furniture(asset, data.get("furniture", []))
                _sync_asset_valuation_history(asset, data.get("valuation_history", []))
                _clear_non_selected_asset_details(asset)

                if purchase_rows_payload_present:
                    if purchase_rows:
                        _apply_asset_purchase_rows_delta(purchase_rows, sign=-1)
                        _sync_asset_purchase_payments(asset, purchase_rows)
                    else:
                        AssetPurchasePayment.objects.filter(asset=asset).delete()
                elif previous_rows:
                    _apply_asset_purchase_rows_delta(previous_rows, sign=-1)

                _sync_gold_balance_from_assets()

        except ValueError as exc:
            return JsonResponse(
                {
                    "error": str(exc),
                    "error_key": str(exc),
                },
                status=400,
            )

        return JsonResponse(asset.to_dict())

    def delete(self, request, pk):
        asset = get_object_or_404(FixedAsset, pk=pk)

        purchase_rows = _purchase_rows_from_instances(
            AssetPurchasePayment.objects.filter(asset=asset).order_by("id")
        )

        try:
            with transaction.atomic():
                # Reverse only when this asset has explicit payment-source rows.
                if purchase_rows:
                    _apply_asset_purchase_rows_delta(purchase_rows, sign=1)

                asset.delete()
                _sync_gold_balance_from_assets()
        except ValueError as exc:
            return JsonResponse(
                {
                    "error": str(exc),
                    "error_key": str(exc),
                },
                status=400,
            )

        return JsonResponse({"deleted": pk})

@method_decorator(csrf_exempt, name="dispatch")
class FixedAssetPhotoView(View):

    def post(self, request, pk):

        asset = get_object_or_404(FixedAsset, pk=pk)

        files = request.FILES.getlist("photos")

        if not files:
            return JsonResponse(
                {"error": "No photos uploaded"},
                status=400,
            )

        uploaded = []

        for file in files:

            photo = AssetPhoto.objects.create(
                asset=asset,
                image_data=file.read(),
                filename=file.name,
                mime_type=file.content_type,
            )

            uploaded.append(photo.to_dict())

        return JsonResponse(uploaded, safe=False)

    def delete(self, request, pk, photo_id):

        photo = get_object_or_404(
            AssetPhoto,
            pk=photo_id,
            asset_id=pk,
        )

        photo.delete()

        return JsonResponse({"deleted": True})

class AssetPhotoView(View):

    def get(self, request, photo_id):

        photo = get_object_or_404(
            AssetPhoto,
            pk=photo_id,
        )

        return HttpResponse(
            photo.image_data,
            content_type=photo.mime_type,
        )


def _document_validation_error_response(exc):
    if hasattr(exc, "messages") and exc.messages:
        message = str(exc.messages[0])
    else:
        message = str(exc)
    return JsonResponse({"error": message}, status=400)


def _document_database_error_response(exc):
    message = str(exc or "")
    if "no such table" in message.lower() and "core_document" in message.lower():
        return JsonResponse(
            {"error": "documents_schema_missing", "detail": "Run database migrations to enable document management."},
            status=503,
        )
    return JsonResponse({"error": "documents_unavailable"}, status=503)


@method_decorator(csrf_exempt, name="dispatch")
class DocumentListUploadView(View):
    service = DocumentService()

    def get(self, request, parent_type, parent_id):
        try:
            docs = self.service.list_documents(parent_type, parent_id)
            return JsonResponse({"documents": docs})
        except ValidationError as exc:
            return _document_validation_error_response(exc)
        except (OperationalError, ProgrammingError) as exc:
            return _document_database_error_response(exc)

    def post(self, request, parent_type, parent_id):
        uploaded_file = request.FILES.get("file")
        category = request.POST.get("document_category")
        notes = request.POST.get("notes", "")

        try:
            item = self.service.upload_document(
                parent_type=parent_type,
                parent_id=parent_id,
                uploaded_file=uploaded_file,
                uploaded_by=request.user,
                category=category,
                notes=notes,
            )
            return JsonResponse(item, status=201)
        except ValidationError as exc:
            return _document_validation_error_response(exc)
        except (OperationalError, ProgrammingError) as exc:
            return _document_database_error_response(exc)


@method_decorator(csrf_exempt, name="dispatch")
class DocumentFileView(View):
    service = DocumentService()

    def get(self, request, document_id):
        try:
            metadata, content = self.service.get_document_content(document_id)
        except (OperationalError, ProgrammingError) as exc:
            return _document_database_error_response(exc)
        if metadata is None:
            return JsonResponse({"error": "document_not_found"}, status=404)

        disposition = request.GET.get("disposition", "inline").strip().lower()
        if disposition not in {"inline", "attachment"}:
            disposition = "inline"

        response = HttpResponse(content, content_type=metadata.get("mime_type") or "application/octet-stream")
        response["Content-Disposition"] = f'{disposition}; filename="{metadata.get("original_file_name", "document")}"'
        response["Content-Length"] = str(metadata.get("file_size", len(content)))
        return response

    def put(self, request, document_id):
        doc = self.service.get_document(document_id)
        if doc is None:
            return JsonResponse({"error": "document_not_found"}, status=404)

        uploaded_file = request.FILES.get("file")
        category = request.POST.get("document_category") if "document_category" in request.POST else None
        notes = request.POST.get("notes") if "notes" in request.POST else None

        try:
            item = self.service.replace_document(
                document_id=document_id,
                uploaded_file=uploaded_file,
                uploaded_by=request.user,
                category=category,
                notes=notes,
            )
            return JsonResponse(item)
        except ValidationError as exc:
            return _document_validation_error_response(exc)
        except (OperationalError, ProgrammingError) as exc:
            return _document_database_error_response(exc)

    def post(self, request, document_id):
        return self.put(request, document_id)

    def delete(self, request, document_id):
        try:
            deleted = self.service.delete_document(document_id)
        except (OperationalError, ProgrammingError) as exc:
            return _document_database_error_response(exc)
        if not deleted:
            return JsonResponse({"error": "document_not_found"}, status=404)
        return JsonResponse({"deleted": True})


@method_decorator(csrf_exempt, name="dispatch")
class DocumentCategoriesView(View):
    service = DocumentService()

    def get(self, request):
        parent_type = request.GET.get("parent_type", "")
        try:
            categories = self.service.categories_for_parent(parent_type)
            return JsonResponse({"categories": categories})
        except ValidationError as exc:
            return _document_validation_error_response(exc)
        except (OperationalError, ProgrammingError) as exc:
            return _document_database_error_response(exc)
    
@method_decorator(csrf_exempt, name="dispatch")
class AssetRenovationListView(View):

    def get(self, request):
        asset_id = request.GET.get("asset")

        qs = AssetRenovation.objects.all().order_by("date", "id")

        if asset_id:
            qs = qs.filter(asset_id=asset_id)

        return JsonResponse({
            "renovations": [r.to_dict() for r in qs]
        })

    def post(self, request):
        data = json.loads(request.body)

        item = AssetRenovation.objects.create(
            asset_id=data["asset_id"],
            date=data["date"],
            category=data["category"],
            description=data.get("description", ""),
            amount_egp=data.get("amount_egp", 0),
            usd_rate=data.get("usd_rate", 0),
            amount_usd=data.get("amount_usd", 0),
            notes=data.get("notes", ""),
        )

        return JsonResponse(item.to_dict(), status=201)

@method_decorator(csrf_exempt, name="dispatch")
class AssetRenovationDetailView(View):

    def put(self, request, pk):
        item = get_object_or_404(AssetRenovation, pk=pk)

        data = json.loads(request.body)

        fields = [
            "date",
            "category",
            "description",
            "amount_egp",
            "usd_rate",
            "amount_usd",
            "notes",
        ]

        for field in fields:
            if field in data:
                setattr(item, field, data[field])

        item.save()

        return JsonResponse(item.to_dict())

    def delete(self, request, pk):
        item = get_object_or_404(AssetRenovation, pk=pk)
        item.delete()

        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class AssetMaintenanceListView(View):

    def get(self, request):
        asset_id = request.GET.get("asset")

        qs = AssetMaintenance.objects.all().order_by("date", "id")

        if asset_id:
            qs = qs.filter(asset_id=asset_id)

        return JsonResponse({
            "maintenance": [m.to_dict() for m in qs]
        })

    def post(self, request):
        data = json.loads(request.body)

        item = AssetMaintenance.objects.create(
            asset_id=data["asset_id"],
            date=data["date"],
            maintenance_type=data["maintenance_type"],
            cost=data.get("cost", 0),
            notes=data.get("notes", ""),
        )

        return JsonResponse(item.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class AssetMaintenanceDetailView(View):

    def put(self, request, pk):
        item = get_object_or_404(AssetMaintenance, pk=pk)

        data = json.loads(request.body)

        fields = ["date", "maintenance_type", "cost", "notes"]

        for field in fields:
            if field in data:
                setattr(item, field, data[field])

        item.save()

        return JsonResponse(item.to_dict())

    def delete(self, request, pk):
        item = get_object_or_404(AssetMaintenance, pk=pk)
        item.delete()

        return JsonResponse({"deleted": pk})


@method_decorator(csrf_exempt, name="dispatch")
class AssetInsuranceListView(View):

    def get(self, request):
        asset_id = request.GET.get("asset")

        qs = AssetInsurance.objects.all().order_by("expiry_date", "id")

        if asset_id:
            qs = qs.filter(asset_id=asset_id)

        return JsonResponse({
            "insurance": [i.to_dict() for i in qs]
        })

    def post(self, request):
        data = json.loads(request.body)

        item = AssetInsurance.objects.create(
            asset_id=data["asset_id"],
            company=data["company"],
            policy_number=data.get("policy_number", ""),
            expiry_date=data.get("expiry_date") or None,
            premium=data.get("premium", 0),
        )

        return JsonResponse(item.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class AssetInsuranceDetailView(View):

    def put(self, request, pk):
        item = get_object_or_404(AssetInsurance, pk=pk)

        data = json.loads(request.body)

        fields = ["company", "policy_number", "expiry_date", "premium"]

        for field in fields:
            if field in data:
                setattr(item, field, data[field])

        item.save()

        return JsonResponse(item.to_dict())

    def delete(self, request, pk):
        item = get_object_or_404(AssetInsurance, pk=pk)
        item.delete()

        return JsonResponse({"deleted": pk})
    
@method_decorator(csrf_exempt, name="dispatch")
class AssetFurnitureListView(View):

    def get(self, request):
        asset_id = request.GET.get("asset")

        qs = AssetFurniture.objects.all().order_by("name")

        if asset_id:
            qs = qs.filter(asset_id=asset_id)

        return JsonResponse({
            "furniture": [f.to_dict() for f in qs]
        })

    def post(self, request):
        data = json.loads(request.body)

        item = AssetFurniture.objects.create(
            asset_id=data["asset_id"],
            name=data["name"],
            category=data.get("category", ""),
            purchase_date=data.get("purchase_date") or None,
            amount_egp=data.get("amount_egp", 0),
            usd_rate=data.get("usd_rate", 0),
            amount_usd=data.get("amount_usd", 0),
            quantity=data.get("quantity", 1),
            notes=data.get("notes", ""),
        )

        return JsonResponse(item.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class AssetFurnitureDetailView(View):

    def put(self, request, pk):
        item = get_object_or_404(AssetFurniture, pk=pk)

        data = json.loads(request.body)

        fields = [
            "name",
            "category",
            "purchase_date",
            "amount_egp",
            "usd_rate",
            "amount_usd",
            "quantity",
            "notes",
        ]

        for field in fields:
            if field in data:
                setattr(item, field, data[field])

        item.save()

        return JsonResponse(item.to_dict())

    def delete(self, request, pk):
        item = get_object_or_404(AssetFurniture, pk=pk)
        item.delete()

        return JsonResponse({"deleted": pk})

@method_decorator(csrf_exempt, name="dispatch")
class AssetValuationHistoryListView(View):

    def get(self, request):
        asset_id = request.GET.get("asset")

        qs = AssetValuationHistory.objects.all().order_by(
            "-valuation_date",
            "-id",
        )

        if asset_id:
            qs = qs.filter(asset_id=asset_id)

        return JsonResponse({
            "valuation_history": [
                v.to_dict() for v in qs
            ]
        })

    def post(self, request):
        data = json.loads(request.body)

        item = AssetValuationHistory.objects.create(
            asset_id=data["asset_id"],
            valuation_date=data["valuation_date"],
            market_value=data["market_value"],
            valuation_source=data.get(
                "valuation_source",
                "Manual",
            ),
            notes=data.get("notes", ""),
        )

        asset = item.asset
        asset.current_market_value = item.market_value
        asset.last_valuation_date = item.valuation_date
        asset.valuation_source = item.valuation_source
        asset.save()

        return JsonResponse(item.to_dict(), status=201)


@method_decorator(csrf_exempt, name="dispatch")
class AssetValuationHistoryDetailView(View):

    def put(self, request, pk):
        item = get_object_or_404(
            AssetValuationHistory,
            pk=pk,
        )

        data = json.loads(request.body)

        fields = [
            "valuation_date",
            "market_value",
            "valuation_source",
            "notes",
        ]

        for field in fields:
            if field in data:
                setattr(item, field, data[field])

        item.save()

        asset = item.asset
        asset.current_market_value = item.market_value
        asset.last_valuation_date = item.valuation_date
        asset.valuation_source = item.valuation_source
        asset.save()

        return JsonResponse(item.to_dict())

    def delete(self, request, pk):
        item = get_object_or_404(
            AssetValuationHistory,
            pk=pk,
        )
        item.delete()

        return JsonResponse({"deleted": pk})

@method_decorator(csrf_exempt, name="dispatch")
class AssetSaleView(View):

    def get(self, request, asset_id):
        asset = get_object_or_404(FixedAsset, pk=asset_id)

        if hasattr(asset, "sale"):
            return JsonResponse(asset.sale.to_dict())

        return JsonResponse({}, status=404)

    def post(self, request, asset_id):
        asset = get_object_or_404(FixedAsset, pk=asset_id)

        data = json.loads(request.body)
        sale_date_value = data.get("sale_date")
        if isinstance(sale_date_value, str):
            try:
                sale_date_value = datetime.date.fromisoformat(sale_date_value)
            except ValueError:
                pass

        existing_sale = getattr(asset, "sale", None)

        try:
            with transaction.atomic():
                if existing_sale is not None:
                    previous_row = _sale_payment_row(existing_sale)
                    _apply_asset_balance_delta(
                        currency_id=previous_row["currency_id"],
                        payment_method=previous_row["payment_method"],
                        bank_id=previous_row["bank_id"],
                        amount_delta=-_to_decimal(previous_row["amount"]),
                    )

                deposit_values = _resolve_sale_deposit_values(data, existing_sale=existing_sale)

                sale, created = AssetSale.objects.update_or_create(
                    asset=asset,
                    defaults={
                        "sale_date": sale_date_value,
                        "sale_price": data["sale_price"],
                        "selling_expenses": data.get("selling_expenses", 0),
                        "net_sale_amount": data["net_sale_amount"],
                        "deposit_balance_id": data.get("deposit_balance_id"),
                        "deposit_currency_id": deposit_values["deposit_currency_id"],
                        "deposit_method": deposit_values["deposit_method"],
                        "deposit_bank_id": deposit_values["deposit_bank_id"],
                        "notes": data.get("notes", ""),
                    },
                )

                current_row = _sale_payment_row(sale)
                _apply_asset_balance_delta(
                    currency_id=current_row["currency_id"],
                    payment_method=current_row["payment_method"],
                    bank_id=current_row["bank_id"],
                    amount_delta=_to_decimal(current_row["amount"]),
                )

                asset.status = "Sold"
                asset.save()
                _sync_gold_balance_from_assets()

        except ValueError as exc:
            return JsonResponse(
                {
                    "error": str(exc),
                    "error_key": str(exc),
                },
                status=400,
            )

        return JsonResponse(sale.to_dict(), status=201 if created else 200)

    def delete(self, request, asset_id):
        asset = get_object_or_404(FixedAsset, pk=asset_id)

        if not hasattr(asset, "sale"):
            return JsonResponse({}, status=404)

        try:
            with transaction.atomic():
                sale_row = _sale_payment_row(asset.sale)
                _apply_asset_balance_delta(
                    currency_id=sale_row["currency_id"],
                    payment_method=sale_row["payment_method"],
                    bank_id=sale_row["bank_id"],
                    amount_delta=-_to_decimal(sale_row["amount"]),
                )

                asset.sale.delete()

                if asset.status == "Sold":
                    asset.status = "Owned"
                    asset.save()

                _sync_gold_balance_from_assets()

        except ValueError as exc:
            return JsonResponse(
                {
                    "error": str(exc),
                    "error_key": str(exc),
                },
                status=400,
            )

        return JsonResponse({"deleted": True})


def _fixed_asset_report_queryset():
    return (
        FixedAsset.objects.select_related(
            "real_estate",
            "vehicle_details",
            "gold_details",
            "other_asset_details",
            "sale",
        )
        .prefetch_related(
            "photos",
            "renovations",
            "maintenance",
            "insurance",
            "furniture",
            "valuation_history",
        )
        .order_by("name")
    )


def _fixed_asset_report_context(request):
    scope = request.GET.get("scope", "single")
    asset_id = request.GET.get("asset_id")
    lang = request.GET.get("lang", "en")
    t = get_translations(lang)

    queryset = _fixed_asset_report_queryset()

    if scope == "single":
        if not asset_id:
            raise ValueError("asset_id is required")
        queryset = queryset.filter(pk=asset_id)

    assets = list(queryset)
    if not assets:
        raise FixedAsset.DoesNotExist()

    return {
        "scope": scope,
        "asset_id": asset_id,
        "lang": lang,
        "t": t,
        "assets": assets,
        "portfolio_snapshot": NetWorthService().fixed_assets_snapshot(),
    }


def _fixed_asset_display_value(value):
    if value in (None, "", []):
        return "-"
    return str(value)


def _fixed_asset_report_label(t, lang, key, default):
    return get_text(key, lang, t, default)


def _fixed_asset_user_text(value, lang):
    if value in (None, ""):
        return "-"
    text = str(value)
    has_arabic = any("\u0600" <= ch <= "\u06FF" for ch in text)
    return format_arabic(text) if lang == "ar" or has_arabic else text


def _fixed_asset_pdf_table(rows, col_widths, font_name):
    table = Table(rows, colWidths=col_widths, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9e1f2")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1f2937")),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#94a3b8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _build_fixed_asset_pdf_story(asset, lang, t, styles, title_style, heading_style, body_style, font_name):
    data = asset.to_dict()
    story = []

    asset_name = _fixed_asset_user_text(asset.name, lang)
    story.append(Paragraph(asset_name, title_style))
    story.append(Spacer(1, 0.25 * cm))

    gain_amount = float(data.get("current_market_value") or 0) - float(data.get("purchase_price") or 0)
    general_rows = [
        [
            _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
            asset_name,
        ],
        [
            _fixed_asset_report_label(t, lang, "asset_type", "Asset Type"),
            _fixed_asset_report_label(
                t,
                lang,
                f"type_{str(data.get('asset_type') or 'other').lower()}",
                data.get("asset_type") or "-",
            ),
        ],
        [
            _fixed_asset_report_label(t, lang, "status", "Status"),
            _fixed_asset_report_label(
                t,
                lang,
                str(data.get("status") or "owned").lower(),
                data.get("status") or "-",
            ),
        ],
        [
            _fixed_asset_report_label(t, lang, "purchase_date", "Purchase Date"),
            _fixed_asset_display_value(data.get("purchase_date")),
        ],
        [
            _fixed_asset_report_label(t, lang, "purchase_price_egp", "Purchase Price (EGP)"),
            f"{float(data.get('purchase_price') or 0):,.2f}",
        ],
        [
            _fixed_asset_report_label(t, lang, "current_market_value", "Current Market Value"),
            f"{float(data.get('current_market_value') or 0):,.2f}",
        ],
        [
            _fixed_asset_report_label(t, lang, "gain_amount", "Gain Amount"),
            f"{gain_amount:,.2f}",
        ],
        [
            _fixed_asset_report_label(t, lang, "notes", "Notes"),
            _fixed_asset_user_text(data.get("notes"), lang),
        ],
    ]

    story.append(Paragraph(_fixed_asset_report_label(t, lang, "general_information", "General Information"), heading_style))
    story.append(_fixed_asset_pdf_table(general_rows, [5 * cm, 10.5 * cm], font_name))
    story.append(Spacer(1, 0.3 * cm))

    real_estate = data.get("real_estate") or {}
    if real_estate:
        property_rows = [
            [_fixed_asset_report_label(t, lang, "country", "Country"), _fixed_asset_user_text(real_estate.get("country"), lang)],
            [_fixed_asset_report_label(t, lang, "governorate", "Governorate"), _fixed_asset_user_text(real_estate.get("governorate"), lang)],
            [_fixed_asset_report_label(t, lang, "city", "City"), _fixed_asset_user_text(real_estate.get("city"), lang)],
            [_fixed_asset_report_label(t, lang, "district", "District"), _fixed_asset_user_text(real_estate.get("district"), lang)],
            [_fixed_asset_report_label(t, lang, "address", "Address"), _fixed_asset_user_text(real_estate.get("address"), lang)],
            [_fixed_asset_report_label(t, lang, "apt_area", "Property Area (Sqm)"), _fixed_asset_display_value(real_estate.get("apartment_area"))],
            [_fixed_asset_report_label(t, lang, "land_area", "Land Area"), _fixed_asset_display_value(real_estate.get("land_area"))],
            [_fixed_asset_report_label(t, lang, "rooms", "Bedrooms"), _fixed_asset_display_value(real_estate.get("rooms"))],
            [_fixed_asset_report_label(t, lang, "bathrooms", "Bathrooms"), _fixed_asset_display_value(real_estate.get("bathrooms"))],
            [_fixed_asset_report_label(t, lang, "description", "Description"), _fixed_asset_user_text(real_estate.get("description"), lang)],
        ]
        story.append(Paragraph(_fixed_asset_report_label(t, lang, "property_details", "Property Details"), heading_style))
        story.append(_fixed_asset_pdf_table(property_rows, [5 * cm, 10.5 * cm], font_name))
        story.append(Spacer(1, 0.3 * cm))

    photos = list(asset.photos.all())
    if photos:
        story.append(Paragraph(_fixed_asset_report_label(t, lang, "photos", "Photos"), heading_style))
        image_rows = []
        current_row = []
        for photo in photos[:4]:
            try:
                img = RLImage(io.BytesIO(photo.image_data), width=6 * cm, height=4.5 * cm)
                current_row.append(img)
            except Exception:
                current_row.append(Paragraph(_fixed_asset_user_text(photo.filename or photo.title or photo.id, lang), body_style))
            if len(current_row) == 2:
                image_rows.append(current_row)
                current_row = []
        if current_row:
            while len(current_row) < 2:
                current_row.append(Paragraph("", body_style))
            image_rows.append(current_row)
        story.append(Table(image_rows, colWidths=[7.7 * cm, 7.7 * cm], hAlign="LEFT"))
        story.append(Spacer(1, 0.3 * cm))

    def build_collection_section(title_key, title_default, items, headers, value_rows):
        if not items:
            return
        story.append(Paragraph(_fixed_asset_report_label(t, lang, title_key, title_default), heading_style))
        rows = [[_fixed_asset_report_label(t, lang, key, default) for key, default in headers]]
        rows.extend(value_rows(item) for item in items)
        story.append(_fixed_asset_pdf_table(rows, [4 * cm, 4 * cm, 3.5 * cm, 4 * cm], font_name))
        story.append(Spacer(1, 0.3 * cm))

    build_collection_section(
        "renovations",
        "Renovations",
        data.get("renovations") or [],
        [("date", "Date"), ("category", "Category"), ("amount_egp", "Amount EGP"), ("notes", "Notes")],
        lambda item: [
            _fixed_asset_display_value(item.get("date")),
            _fixed_asset_user_text(item.get("category"), lang),
            f"{float(item.get('amount_egp') or 0):,.2f}",
            _fixed_asset_user_text(item.get("notes"), lang),
        ],
    )

    build_collection_section(
        "furniture",
        "Furniture",
        data.get("furniture") or [],
        [("asset_name", "Name"), ("category", "Category"), ("amount_egp", "Amount EGP"), ("notes", "Notes")],
        lambda item: [
            _fixed_asset_user_text(item.get("name"), lang),
            _fixed_asset_user_text(item.get("category"), lang),
            f"{float(item.get('amount_egp') or 0):,.2f}",
            _fixed_asset_user_text(item.get("notes"), lang),
        ],
    )

    build_collection_section(
        "valuation_history",
        "Valuation History",
        data.get("valuation_history") or [],
        [("date", "Date"), ("current_market_value", "Market Value"), ("valuation_source", "Valuation Source"), ("notes", "Notes")],
        lambda item: [
            _fixed_asset_display_value(item.get("valuation_date")),
            f"{float(item.get('market_value') or 0):,.2f}",
            _fixed_asset_user_text(item.get("valuation_source"), lang),
            _fixed_asset_user_text(item.get("notes"), lang),
        ],
    )

    sale = data.get("sale") or None
    if sale:
        sale_rows = [
            [_fixed_asset_report_label(t, lang, "sale_date", "Sale Date"), _fixed_asset_display_value(sale.get("sale_date"))],
            [_fixed_asset_report_label(t, lang, "sale_price_egp", "Sale Price (EGP)"), f"{float(sale.get('sale_price') or 0):,.2f}"],
            [_fixed_asset_report_label(t, lang, "selling_expenses_egp", "Selling Expenses (EGP)"), f"{float(sale.get('selling_expenses') or 0):,.2f}"],
            [_fixed_asset_report_label(t, lang, "net_sale_amount", "Net Sale Amount"), f"{float(sale.get('net_sale_amount') or 0):,.2f}"],
            [_fixed_asset_report_label(t, lang, "deposit_balance", "Deposit Balance"), _fixed_asset_display_value(sale.get("deposit_balance_id"))],
            [_fixed_asset_report_label(t, lang, "notes", "Notes"), _fixed_asset_user_text(sale.get("notes"), lang)],
        ]
        story.append(Paragraph(_fixed_asset_report_label(t, lang, "sale_information", "Sale Information"), heading_style))
        story.append(_fixed_asset_pdf_table(sale_rows, [5 * cm, 10.5 * cm], font_name))
        story.append(Spacer(1, 0.3 * cm))

    return story


@method_decorator(csrf_exempt, name="dispatch")
class FixedAssetPdfReportView(View):

    def get(self, request):
        try:
            context = _fixed_asset_report_context(request)
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        except FixedAsset.DoesNotExist:
            return JsonResponse({"error": "No fixed assets found"}, status=404)

        lang = context["lang"]
        t = context["t"]
        assets = context["assets"]
        scope = context["scope"]
        portfolio_snapshot = context.get("portfolio_snapshot") or {}

        font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "arial.ttf")
        font_exists = os.path.exists(font_path)
        if font_exists:
            pdfmetrics.registerFont(TTFont("ArabicFont", font_path))

        # Always prefer Arabic-capable font when available so mixed-language
        # content (e.g., Arabic names in EN report) renders correctly.
        font_name = "ArabicFont" if font_exists else "Helvetica"
        font_name_bold = "ArabicFont" if font_name == "ArabicFont" else "Helvetica-Bold"

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "FixedAssetTitle",
            parent=styles["Heading1"],
            fontName=font_name_bold,
            fontSize=16,
            textColor=colors.HexColor("#1f2937"),
            spaceAfter=8,
        )
        heading_style = ParagraphStyle(
            "FixedAssetHeading",
            parent=styles["Heading2"],
            fontName=font_name_bold,
            fontSize=12,
            textColor=colors.HexColor("#1a6ef5"),
            spaceBefore=4,
            spaceAfter=6,
        )
        body_style = ParagraphStyle(
            "FixedAssetBody",
            parent=styles["BodyText"],
            fontName=font_name,
            fontSize=9,
            textColor=colors.HexColor("#1f2937"),
            leading=11,
        )

        report_title = _fixed_asset_report_label(
            t,
            lang,
            "fixed_assets_report_title",
            "Fixed Assets Report",
        )
        if scope == "single":
            report_title = f"{report_title} - {_fixed_asset_user_text(assets[0].name, lang)}"

        story = [Paragraph(report_title, title_style), Spacer(1, 0.35 * cm)]

        if scope == "portfolio":
            summary_rows = [
                [
                    _fixed_asset_report_label(t, lang, "total_fixed_assets_value", "Total Fixed Assets"),
                    f"{float(portfolio_snapshot.get('total_fixed_assets_value') or 0):,.2f}",
                ],
                [
                    _fixed_asset_report_label(t, lang, "net_worth", "Net Worth"),
                    f"{float(portfolio_snapshot.get('total_net_worth') or 0):,.2f}",
                ],
                [
                    _fixed_asset_report_label(t, lang, "net_worth_contribution", "Net Worth Contribution"),
                    f"{float(portfolio_snapshot.get('net_worth_contribution') or 0):,.2f}%",
                ],
            ]
            story.append(Paragraph(_fixed_asset_report_label(t, lang, "portfolio_distribution", "Portfolio Distribution"), heading_style))
            story.append(_fixed_asset_pdf_table(summary_rows, [7 * cm, 8.5 * cm], font_name))
            story.append(Spacer(1, 0.35 * cm))

        for index, asset in enumerate(assets):
            story.extend(
                _build_fixed_asset_pdf_story(
                    asset,
                    lang,
                    t,
                    styles,
                    title_style,
                    heading_style,
                    body_style,
                    font_name,
                )
            )
            if index < len(assets) - 1:
                story.append(PageBreak())

        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        filename = (
            f"fixed_asset_{assets[0].id}_report.pdf"
            if scope == "single"
            else "fixed_assets_portfolio_report.pdf"
        )
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


@method_decorator(csrf_exempt, name="dispatch")
class FixedAssetExcelReportView(View):

    def get(self, request):
        try:
            context = _fixed_asset_report_context(request)
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        except FixedAsset.DoesNotExist:
            return JsonResponse({"error": "No fixed assets found"}, status=404)

        lang = context["lang"]
        t = context["t"]
        assets = context["assets"]
        scope = context["scope"]
        portfolio_snapshot = context.get("portfolio_snapshot") or {}

        wb = openpyxl.Workbook()
        summary_ws = wb.active
        if summary_ws is None:
            summary_ws = wb.create_sheet(title="Summary")
        else:
            summary_ws.title = "Summary"

        header_font = Font(bold=True)

        summary_headers = [
            _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
            _fixed_asset_report_label(t, lang, "asset_type", "Asset Type"),
            _fixed_asset_report_label(t, lang, "status", "Status"),
            _fixed_asset_report_label(t, lang, "purchase_date", "Purchase Date"),
            _fixed_asset_report_label(t, lang, "purchase_price_egp", "Purchase Price (EGP)"),
            _fixed_asset_report_label(t, lang, "current_market_value", "Current Market Value"),
            _fixed_asset_report_label(t, lang, "country", "Country"),
            _fixed_asset_report_label(t, lang, "city", "City"),
            _fixed_asset_report_label(t, lang, "address", "Address"),
            _fixed_asset_report_label(t, lang, "sale_date", "Sale Date"),
            _fixed_asset_report_label(t, lang, "net_sale_amount", "Net Sale Amount"),
            _fixed_asset_report_label(t, lang, "notes", "Notes"),
        ]
        summary_ws.append(summary_headers)
        for cell in summary_ws[1]:
            cell.font = header_font

        for asset in assets:
            data = asset.to_dict()
            real_estate = data.get("real_estate") or {}
            sale = data.get("sale") or {}
            summary_ws.append(
                [
                    data.get("name"),
                    data.get("asset_type"),
                    data.get("status"),
                    data.get("purchase_date"),
                    float(data.get("purchase_price") or 0),
                    float(data.get("current_market_value") or 0),
                    real_estate.get("country"),
                    real_estate.get("city"),
                    real_estate.get("address"),
                    sale.get("sale_date"),
                    float(sale.get("net_sale_amount") or 0),
                    data.get("notes"),
                ]
            )

        if scope == "portfolio":
            summary_ws.append([])
            summary_ws.append([
                _fixed_asset_report_label(t, lang, "total_fixed_assets_value", "Total Fixed Assets"),
                float(portfolio_snapshot.get("total_fixed_assets_value") or 0),
            ])
            summary_ws.append([
                _fixed_asset_report_label(t, lang, "net_worth", "Net Worth"),
                float(portfolio_snapshot.get("total_net_worth") or 0),
            ])
            summary_ws.append([
                _fixed_asset_report_label(t, lang, "net_worth_contribution", "Net Worth Contribution"),
                float(portfolio_snapshot.get("net_worth_contribution") or 0),
            ])

        collections = [
            (
                "Renovations",
                _fixed_asset_report_label(t, lang, "renovations", "Renovations"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "date", "Date"),
                    _fixed_asset_report_label(t, lang, "category", "Category"),
                    _fixed_asset_report_label(t, lang, "amount_egp", "Amount EGP"),
                    _fixed_asset_report_label(t, lang, "notes", "Notes"),
                ],
                lambda asset_data, item: [
                    asset_data.get("name"),
                    item.get("date"),
                    item.get("category"),
                    float(item.get("amount_egp") or 0),
                    item.get("notes"),
                ],
                lambda asset_data: asset_data.get("renovations") or [],
            ),
            (
                "Furniture",
                _fixed_asset_report_label(t, lang, "furniture", "Furniture"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "category", "Category"),
                    _fixed_asset_report_label(t, lang, "purchase_date", "Purchase Date"),
                    _fixed_asset_report_label(t, lang, "amount_egp", "Amount EGP"),
                    _fixed_asset_report_label(t, lang, "notes", "Notes"),
                ],
                lambda asset_data, item: [
                    item.get("name"),
                    item.get("category"),
                    item.get("purchase_date"),
                    float(item.get("amount_egp") or 0),
                    item.get("notes"),
                ],
                lambda asset_data: asset_data.get("furniture") or [],
            ),
            (
                "Valuations",
                _fixed_asset_report_label(t, lang, "valuation_history", "Valuation History"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "date", "Date"),
                    _fixed_asset_report_label(t, lang, "current_market_value", "Market Value"),
                    _fixed_asset_report_label(t, lang, "valuation_source", "Valuation Source"),
                    _fixed_asset_report_label(t, lang, "notes", "Notes"),
                ],
                lambda asset_data, item: [
                    asset_data.get("name"),
                    item.get("valuation_date"),
                    float(item.get("market_value") or 0),
                    item.get("valuation_source"),
                    item.get("notes"),
                ],
                lambda asset_data: asset_data.get("valuation_history") or [],
            ),
            (
                "Photos",
                _fixed_asset_report_label(t, lang, "photos", "Photos"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "description", "Description"),
                    _fixed_asset_report_label(t, lang, "notes", "Filename"),
                    "URL",
                ],
                lambda asset_data, item: [
                    asset_data.get("name"),
                    item.get("title"),
                    item.get("filename"),
                    item.get("url"),
                ],
                lambda asset_data: asset_data.get("photos") or [],
            ),
        ]

        sale_ws = wb.create_sheet(title="Sale")
        sale_headers = [
            _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
            _fixed_asset_report_label(t, lang, "sale_date", "Sale Date"),
            _fixed_asset_report_label(t, lang, "sale_price_egp", "Sale Price (EGP)"),
            _fixed_asset_report_label(t, lang, "selling_expenses_egp", "Selling Expenses (EGP)"),
            _fixed_asset_report_label(t, lang, "net_sale_amount", "Net Sale Amount"),
            _fixed_asset_report_label(t, lang, "deposit_balance", "Deposit Balance"),
            _fixed_asset_report_label(t, lang, "notes", "Notes"),
        ]
        sale_ws.append(sale_headers)
        for cell in sale_ws[1]:
            cell.font = header_font

        for asset in assets:
            asset_data = asset.to_dict()
            sale = asset_data.get("sale")
            if not sale:
                continue
            sale_ws.append(
                [
                    asset_data.get("name"),
                    sale.get("sale_date"),
                    float(sale.get("sale_price") or 0),
                    float(sale.get("selling_expenses") or 0),
                    float(sale.get("net_sale_amount") or 0),
                    sale.get("deposit_balance_id"),
                    sale.get("notes"),
                ]
            )

        for sheet_name, title, headers, row_builder, collection_getter in collections:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
            for asset in assets:
                asset_data = asset.to_dict()
                for item in collection_getter(asset_data):
                    ws.append(row_builder(asset_data, item))

        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = (
            f"fixed_asset_{assets[0].id}_report.xlsx"
            if scope == "single"
            else "fixed_assets_portfolio_report.xlsx"
        )
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response