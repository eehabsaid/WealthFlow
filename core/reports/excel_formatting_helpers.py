# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Exact formats from original ───────────────────────────────────────────────
FMT_EGP = "[$ج.م.\u200f-C01]\\ #,##0.00_-"
FMT_EGP_RED = "[$ج.م.\u200f-C01]\\ #,##0.00;[Red][$ج.م.\u200f-C01]\\ #,##0.00"
FMT_USD = '"$"#,##0.00;[Red]"$"#,##0.00'
FMT_EUR = "[$EUR]\\ #,##0.00;[Red][$EUR]\\ #,##0.00"
FMT_SAR = "[$SAR]\\ #,##0.00;[Red][$SAR]\\ #,##0.00"
FMT_GOLD = '0\\ "Grams"'
FMT_EGP_CERT = "[$EGP]\\ #,##0.00"
FMT_EGP_CERT_R = "[$EGP]\\ #,##0.00;[Red][$EGP]\\ #,##0.00"
FMT_PCT = "0.00%"
FMT_DATE = "dd-mmm-yyyy"
FMT_INT = "0"

GREY = "FF7F7F7F"
RED = "FFFF0000"
WHITE = "FFFFFFFF"
EXP_BG = "FF203864"
EXP_MONTH_BG = "FFD9E1F2"
EXP_YEAR_BG = "FFBDD7EE"

MONTH_ORDER = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
    "Quarter-Bonuses",
]


def auto_adjust_columns(ws):
    """
    Adjusts column widths based on content, using a larger buffer
    specifically for date columns.
    """
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        is_date_col = False

        for cell in col:
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                is_date_col = True

            fmt = str(cell.number_format).lower()
            if any(x in fmt for x in ["yyyy", "mmmm", "dddd"]):
                is_date_col = True

            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_length:
                max_length = len(val)

        if is_date_col:
            ws.column_dimensions[column].width = max_length + 20
        else:
            ws.column_dimensions[column].width = max_length + 10


def _f(bold=False, size=11, color=None, name="Arial"):
    kw = dict(bold=bold, size=size, name=name)
    if color:
        kw["color"] = color
    return Font(**kw)


def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _thin_lr():
    s = Side(style="thin")
    return Border(left=s, right=s)


def _thin_tb():
    s = Side(style="thin")
    return Border(top=s, bottom=s)


def _thin_b():
    return Border(bottom=Side(style="thin"))


def _thin_t():
    return Border(top=Side(style="thin"))


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _align(h="general", v="bottom", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _center():
    return Alignment(horizontal="center", vertical="bottom")


def _msort(m):
    m = str(m).strip()
    for i, n in enumerate(MONTH_ORDER):
        if n.lower().startswith(m.lower()[:3]):
            return i
    return 99


def _apply_zebra_striping(ws, row, col_count):
    fill = _fill("FFF2F2F2") if (row % 2 == 0) else _fill("FFFFFFFF")
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).fill = fill
