# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
import io
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse, JsonResponse
from core.models import FixedAsset
from core.reports.fixed_asset_report_helpers import (
    fixed_asset_report_context as _fixed_asset_report_context,
    fixed_asset_report_label as _fixed_asset_report_label,
)

class FixedAssetExcelReportGenerator(object):

    def get(self, request):
        try:
            context = _fixed_asset_report_context(request)
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        except FixedAsset.DoesNotExist:
            return JsonResponse({"error": "No fixed assets found"}, status=404)

        lang = context["lang"]
        t = context["t"]
        assets = context["assets"]
        scope = context["scope"]
        portfolio_snapshot = context.get("portfolio_snapshot") or {}

        wb = openpyxl.Workbook()
        summary_ws = wb.active
        if summary_ws is None:
            summary_ws = wb.create_sheet(title="Summary")
        else:
            summary_ws.title = "Summary"

        header_font = Font(bold=True)

        summary_headers = [
            _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
            _fixed_asset_report_label(t, lang, "asset_type", "Asset Type"),
            _fixed_asset_report_label(t, lang, "status", "Status"),
            _fixed_asset_report_label(t, lang, "purchase_date", "Purchase Date"),
            _fixed_asset_report_label(t, lang, "purchase_price_egp", "Purchase Price (EGP)"),
            _fixed_asset_report_label(t, lang, "total_investment_egp", "Total Investment (EGP)"),
            _fixed_asset_report_label(t, lang, "current_market_value", "Current Market Value"),
            _fixed_asset_report_label(t, lang, "gain_loss", "Gain / Loss"),
            _fixed_asset_report_label(t, lang, "country", "Country"),
            _fixed_asset_report_label(t, lang, "city", "City"),
            _fixed_asset_report_label(t, lang, "address", "Address"),
            _fixed_asset_report_label(t, lang, "sale_date", "Sale Date"),
            _fixed_asset_report_label(t, lang, "net_sale_amount", "Net Sale Amount"),
            _fixed_asset_report_label(t, lang, "notes", "Notes"),
        ]
        summary_ws.append(summary_headers)
        for cell in summary_ws[1]:
            cell.font = header_font

        for asset in assets:
            data = asset.to_dict()
            real_estate = data.get("real_estate") or {}
            sale = data.get("sale") or {}
            summary_ws.append(
                [
                    data.get("name"),
                    data.get("asset_type"),
                    data.get("status"),
                    data.get("purchase_date"),
                    float(data.get("purchase_price") or 0),
                    float(data.get("total_investment") or data.get("purchase_price") or 0),
                    float(data.get("current_market_value") or 0),
                    float(data.get("gain_loss") or 0),
                    real_estate.get("country"),
                    real_estate.get("city"),
                    real_estate.get("address"),
                    sale.get("sale_date"),
                    float(sale.get("net_sale_amount") or 0),
                    data.get("notes"),
                ]
            )

        if scope == "portfolio":
            summary_ws.append([])
            summary_ws.append([
                _fixed_asset_report_label(t, lang, "total_fixed_assets_value", "Total Fixed Assets"),
                float(portfolio_snapshot.get("total_fixed_assets_value") or 0),
            ])
            summary_ws.append([
                _fixed_asset_report_label(t, lang, "net_worth", "Net Worth"),
                float(portfolio_snapshot.get("total_net_worth") or 0),
            ])
            summary_ws.append([
                _fixed_asset_report_label(t, lang, "net_worth_contribution", "Net Worth Contribution"),
                float(portfolio_snapshot.get("net_worth_contribution") or 0),
            ])

        collections = [
            (
                "Acquisition Costs",
                _fixed_asset_report_label(t, lang, "acquisition_costs", "Acquisition Costs"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "date", "Date"),
                    _fixed_asset_report_label(t, lang, "category", "Category"),
                    _fixed_asset_report_label(t, lang, "amount_egp", "Amount EGP"),
                    _fixed_asset_report_label(t, lang, "notes", "Notes"),
                ],
                lambda asset_data, item: [
                    asset_data.get("name"),
                    item.get("date"),
                    item.get("category"),
                    float(item.get("amount_egp") or 0),
                    item.get("notes"),
                ],
                lambda asset_data: asset_data.get("acquisition_costs") or [],
            ),
            (
                "Renovations",
                _fixed_asset_report_label(t, lang, "renovations", "Renovations"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "date", "Date"),
                    _fixed_asset_report_label(t, lang, "category", "Category"),
                    _fixed_asset_report_label(t, lang, "amount_egp", "Amount EGP"),
                    _fixed_asset_report_label(t, lang, "notes", "Notes"),
                ],
                lambda asset_data, item: [
                    asset_data.get("name"),
                    item.get("date"),
                    item.get("category"),
                    float(item.get("amount_egp") or 0),
                    item.get("notes"),
                ],
                lambda asset_data: asset_data.get("renovations") or [],
            ),
            (
                "Furniture",
                _fixed_asset_report_label(t, lang, "furniture", "Furniture"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "category", "Category"),
                    _fixed_asset_report_label(t, lang, "purchase_date", "Purchase Date"),
                    _fixed_asset_report_label(t, lang, "amount_egp", "Amount EGP"),
                    _fixed_asset_report_label(t, lang, "notes", "Notes"),
                ],
                lambda asset_data, item: [
                    item.get("name"),
                    item.get("category"),
                    item.get("purchase_date"),
                    float(item.get("amount_egp") or 0),
                    item.get("notes"),
                ],
                lambda asset_data: asset_data.get("furniture") or [],
            ),
            (
                "Valuations",
                _fixed_asset_report_label(t, lang, "valuation_history", "Valuation History"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "date", "Date"),
                    _fixed_asset_report_label(t, lang, "current_market_value", "Market Value"),
                    _fixed_asset_report_label(t, lang, "valuation_source", "Valuation Source"),
                    _fixed_asset_report_label(t, lang, "notes", "Notes"),
                ],
                lambda asset_data, item: [
                    asset_data.get("name"),
                    item.get("valuation_date"),
                    float(item.get("market_value") or 0),
                    item.get("valuation_source"),
                    item.get("notes"),
                ],
                lambda asset_data: asset_data.get("valuation_history") or [],
            ),
            (
                "Photos",
                _fixed_asset_report_label(t, lang, "photos", "Photos"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "description", "Description"),
                    _fixed_asset_report_label(t, lang, "notes", "Filename"),
                    "URL",
                ],
                lambda asset_data, item: [
                    asset_data.get("name"),
                    item.get("title"),
                    item.get("filename"),
                    item.get("url"),
                ],
                lambda asset_data: asset_data.get("photos") or [],
            ),
        ]

        sale_ws = wb.create_sheet(title="Sale")
        sale_headers = [
            _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
            _fixed_asset_report_label(t, lang, "sale_date", "Sale Date"),
            _fixed_asset_report_label(t, lang, "sale_price_egp", "Sale Price (EGP)"),
            _fixed_asset_report_label(t, lang, "selling_expenses_egp", "Selling Expenses (EGP)"),
            _fixed_asset_report_label(t, lang, "net_sale_amount", "Net Sale Amount"),
            _fixed_asset_report_label(t, lang, "deposit_balance", "Deposit Balance"),
            _fixed_asset_report_label(t, lang, "notes", "Notes"),
        ]
        sale_ws.append(sale_headers)
        for cell in sale_ws[1]:
            cell.font = header_font

        for asset in assets:
            asset_data = asset.to_dict()
            sale = asset_data.get("sale")
            if not sale:
                continue
            sale_ws.append(
                [
                    asset_data.get("name"),
                    sale.get("sale_date"),
                    float(sale.get("sale_price") or 0),
                    float(sale.get("selling_expenses") or 0),
                    float(sale.get("net_sale_amount") or 0),
                    sale.get("deposit_balance_id"),
                    sale.get("notes"),
                ]
            )

        for sheet_name, title, headers, row_builder, collection_getter in collections:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
            for asset in assets:
                asset_data = asset.to_dict()
                for item in collection_getter(asset_data):
                    ws.append(row_builder(asset_data, item))

        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = (
            f"fixed_asset_{assets[0].id}_report.xlsx"
            if scope == "single"
            else "fixed_assets_portfolio_report.xlsx"
        )
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
