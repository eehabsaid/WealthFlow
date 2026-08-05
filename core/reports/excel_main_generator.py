# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
import io
from openpyxl import Workbook
from django.db.models import Max

from core.reports.excel_formatting_helpers import auto_adjust_columns
from core.reports.excel_salary_builder import build_salary_sheet
from core.reports.excel_sheets_builder import (
    build_exchange_rates_sheet,
    build_gold_price_sheet,
    build_bank_certificates_sheet,
    build_balance_sheet,
    build_expenses_sheet,
    build_currency_exchanges_sheet,
)

def generate_excel(output_path=None, lang="ar"):
    from core.models import (
        Company,
        BalanceEntry,
        BankCertificate,
        ExchangeRate,
        GoldPrice,
        Expense,
        CurrencyExchange,
    )

    wb = Workbook()
    wb.remove(wb.active)

    companies = list(Company.objects.all().order_by("order"))
    balance_entries = list(
        BalanceEntry.objects.select_related("currency", "bank").all()
    )

    ws_ex = wb.create_sheet("Exchange Rates")
    latest_ids = ExchangeRate.objects.values("currency_code").annotate(
        latest=Max("fetched_at")
    )
    rates = []
    for item in latest_ids:
        r = ExchangeRate.objects.filter(
            currency_code=item["currency_code"], fetched_at=item["latest"]
        ).first()
        if r:
            rates.append(r)
    build_exchange_rates_sheet(ws_ex, rates, balance_entries)
    auto_adjust_columns(ws_ex)

    ws_gold = wb.create_sheet("Gold Price")
    build_gold_price_sheet(ws_gold, GoldPrice.objects, balance_entries)
    auto_adjust_columns(ws_gold)

    company_sheet_rows = {}
    for company in companies:
        entries = list(company.salary_entries.all())
        ws_sal = wb.create_sheet(company.name)
        sr = build_salary_sheet(ws_sal, company, entries)
        if company.name in ("ElSeweedy Technology", "Dedalus", "Globemed"):
            company_sheet_rows[company.name] = (company.name, sr)
        else:
            company_sheet_rows[company.name] = (company.name, sr)

    ws_cert = wb.create_sheet("Bank-Certificates")
    build_bank_certificates_sheet(ws_cert, BankCertificate.objects.all())
    auto_adjust_columns(ws_cert)

    ws_bal = wb.create_sheet("BALANCE")
    build_balance_sheet(ws_bal, balance_entries, company_sheet_rows)

    ws_ce = wb.create_sheet("Currency Exchanges")
    build_currency_exchanges_sheet(ws_ce, CurrencyExchange.objects.all(), lang=lang)
    auto_adjust_columns(ws_ce)

    ws_exp = wb.create_sheet("Expenses")
    build_expenses_sheet(ws_exp, Expense.objects.all())

    if output_path:
        wb.save(output_path)
        return output_path
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
