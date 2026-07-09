# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false

import json
from decimal import Decimal, InvalidOperation
from django.contrib.auth import authenticate, get_user_model, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import UserPassesTestMixin
from django.http import JsonResponse
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum, Count
from django.db.utils import OperationalError, ProgrammingError
from django.shortcuts import render, get_object_or_404, redirect
from django.core.exceptions import ValidationError
from core.models import (
    Company,
    SalaryEntry,
    Bank,
    BalanceEntry,
    AppSettings,
    ExchangeRate,
    GoldPrice,
    GoldPriceHistory,
    Currency,
    ExpenseCategory,
    ExpenseSubcategory,
    Expense,
    BankCertificate,
    BankCertificateInterestHistory,
    _is_certificate_active,
    PagePermission,
    PAGE_PERMISSION_CHOICES,
    UserProfile,
    ReminderRule,
    CertificateStatus,
    ReminderLog,
    REMINDER_TYPE_CHOICES,
    SALARY_TRIGGER_CHOICES,
    FixedAsset,
    RealEstateDetails,
    VehicleDetails,
    GoldDetails,
    OtherAssetDetails,
    AssetRenovation,
    AssetMaintenance,
    AssetInsurance,
    AssetFurniture,
    AssetValuationHistory,
    AssetPurchasePayment,
    AssetSale,
    AssetPhoto,
    AssetMortgage,
    AssetRental,
    GoldTypeSetting,
    GoldPuritySetting,
    EmailTemplate,
    Goal,
    PerDiem,

)
from django.core.paginator import Paginator, EmptyPage
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import json as _json
import datetime
import os
import io
from django.conf import settings
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    HRFlowable,
    Image as RLImage,
    PageBreak,
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from arabic_reshaper import reshape
from bidi.algorithm import get_display
from django.views.decorators.http import require_http_methods
from core.services.balance.net_worth_service import NetWorthService
from core.services.balance.financial_sync_service import FinancialSyncService
from core.services.shared.document_service import DocumentService
from core.services.shared.exchange_rate_service import ExchangeRateService
from core.services.fixed_assets.gold_valuation_service import GoldValuationService
from core.services.fixed_assets.property_valuation_service import PropertyValuationService
from core.services.shared.reminder_automation_service import ReminderAutomationService
from core.services.shared.auth_workflow_service import AuthWorkflowService, EmailTemplateService
from core.services.financial_advisor.cash_flow_forecast_service import CashFlowForecastService
from core.services.financial_advisor.goal_planning_service import GoalPlanningService
from core.services.financial_advisor.portfolio_optimizer_service import PortfolioOptimizerService
from core.services.financial_advisor.wealth_growth_forecast_service import WealthGrowthForecastService



User = get_user_model()

PAGE_PERMISSION_KEYS = [key for key, _ in PAGE_PERMISSION_CHOICES]

MONTH_ORDER = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

REAL_ESTATE_ASSET_TYPES = {"Real Estate"}
VEHICLE_ASSET_TYPES = {"Vehicles"}
GOLD_ASSET_TYPES = {"Gold"}
OTHER_ASSET_TYPES = {"Other Assets"}

ASSET_PAYMENT_METHOD_CASH = "Cash"
ASSET_PAYMENT_METHOD_CARD = "Card"
ASSET_PAYMENT_METHOD_BANK = "Bank"
ASSET_PAYMENT_METHOD_BANK_TRANSFER = "Bank Transfer"

ASSET_PAYMENT_METHOD_NORMALIZED = {
    "cash": ASSET_PAYMENT_METHOD_CASH,
    "card": ASSET_PAYMENT_METHOD_CARD,
    "bank": ASSET_PAYMENT_METHOD_BANK,
    "bank transfer": ASSET_PAYMENT_METHOD_BANK_TRANSFER,
    "bank_transfer": ASSET_PAYMENT_METHOD_BANK_TRANSFER,
}

GOLD_UNIT_TO_GRAMS = {
    "g": Decimal("1"),
    "gm": Decimal("1"),
    "gram": Decimal("1"),
    "grams": Decimal("1"),
    "kg": Decimal("1000"),
    "kilogram": Decimal("1000"),
    "kilograms": Decimal("1000"),
    "oz": Decimal("31.1034768"),
    "ounce": Decimal("31.1034768"),
    "ounces": Decimal("31.1034768"),
    "tola": Decimal("11.6638038"),
}

from django.db.models.signals import post_save
from django.dispatch import receiver

def _parse_iso_date(value):
    if not value or str(value).strip() in ("", "None"):
        return None
    if isinstance(value, datetime.date):
        return value
    try:
        return datetime.date.fromisoformat(str(value).strip())
    except (ValueError, TypeError):
        return None

def _api_auth_required(request):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Not authenticated"}, status=401)
    return None


try:
    from core.views.auth_views import AdminRequiredMixin
except (ImportError, ValueError):
    pass

try:
    from core.views.certificate_views import _run_certificate_interest_sync
except (ImportError, ValueError):
    pass



class ExportExcelWorkbookGenerator(object):
    """
    Generates a multi-tab Excel Workbook from live DB data,
    matching the original Balance.xlsx format, styles, and formulas,
    with an added Expenses tab.
    """

    def get(self, request):
        return self.post(request)

    def post(self, request):
        from core.reports.excel_generator import generate_excel
        from datetime import date

        buf = generate_excel()
        filename = f"Balance_Tracker_{date.today().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response




# Month sort order — ensures API returns months in calendar order, not alphabetically




def month_sort_key(entry_dict):
    try:
        return MONTH_ORDER.index(entry_dict.get("month", ""))
    except ValueError:
        return len(MONTH_ORDER)




# ══════════════════════════════════════════════════════════════
# PDF REPORT VIEW
# ══════════════════════════════════════════════════════════════
# Helper to load translations
def get_translations(lang):
    path = os.path.join(settings.BASE_DIR, "static", "i18n", f"{lang}.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}




def format_arabic(text):
    return get_display(reshape(str(text)))




def get_text(key, lang, t, default=""):
    text = t.get(key, default)
    return format_arabic(text) if lang == "ar" else text




class GenerateReportGenerator(object):
    """
    POST /api/reports/generate/
    body: { type: "monthly"|"yearly"|"custom",
            year: 2026, month: 5,       # for monthly
            start_date: "2026-01-01",   # for custom
            end_date:   "2026-05-31" }
    Returns: PDF file
    """

    def post(self, request):
        import json as _json, datetime
        from django.http import HttpResponse, JsonResponse
        from django.db.models import Sum

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate,
                Paragraph,
                Spacer,
                Table,
                TableStyle,
                HRFlowable,
            )
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            import io
        except ImportError:
            return JsonResponse(
                {"error": "reportlab not installed. Run: pip install reportlab"},
                status=500,
            )

        data = _json.loads(request.body)
        lang = data.get("lang", "en")
        t = get_translations(lang)

        # Register Arabic-compatible font if the file exists
        font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "arial.ttf")
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont("ArabicFont", font_path))

        # Decide which font style template to use
        pdf_font = "ArabicFont" if lang == "ar" else "Helvetica-Bold"

        rtype = data.get("type", "monthly")
        year = int(data.get("year", datetime.date.today().year))
        month = int(data.get("month", datetime.date.today().month))

        # Accept both parameter styles (with or without suffix) to be fully secure
        start_date = data.get("start_date") or data.get("start")
        end_date = data.get("end_date") or data.get("end")

        # ── Filter expenses safely across all field schemas ──
        qs = Expense.objects.select_related("category", "subcategory").all()
        if rtype == "monthly":
            qs = qs.filter(year=year, month=month)
            month_name = datetime.date(year, month, 1).strftime("%B")
            json_month_key = f"month_{month_name.lower()}"
            translated_month = t.get(json_month_key)

            if not translated_month:
                if lang == "ar":
                    ARABIC_MONTHS = {
                        "January": "يناير",
                        "February": "فبراير",
                        "March": "مارس",
                        "April": "أبريل",
                        "May": "مايو",
                        "June": "يونيو",
                        "July": "يوليو",
                        "August": "أغسطس",
                        "September": "سبتمبر",
                        "October": "أكتوبر",
                        "November": "نوفمبر",
                        "December": "ديسمبر",
                    }
                    translated_month = ARABIC_MONTHS.get(month_name, month_name)
                else:
                    translated_month = month_name

            # FIXED: Changed long em-dash (—) to standard universal hyphen (-)
            title_str = f"{t.get('monthly_report', 'Monthly Report')} - {translated_month} {year}"
            filename = f"report_{year}_{month:02d}.pdf"
        elif rtype == "yearly":
            qs = qs.filter(year=year)
            # FIXED: Changed to standard hyphen
            title_str = f"{t.get('yearly_report', 'Yearly Report')} - {year}"
            filename = f"report_{year}.pdf"
        else:
            from datetime import date as _date

            sd = _date.fromisoformat(start_date)
            ed = _date.fromisoformat(end_date)
            qs = qs.filter(date__gte=sd, date__lte=ed)

            title_str = f"{t.get('report', 'Report')} {start_date} {t.get('to', 'to')} {end_date}"
            filename = f"report_{start_date}_{end_date}.pdf"

        if lang == "ar":
            title_str = format_arabic(title_str)

        expenses = list(qs)
        total_exp = sum(float(e.amount) for e in expenses)

        # Income for period (salary paid amounts)
        from core.services.reports.report_service import ReportService
        total_inc = ReportService.get_period_income(rtype, year, month, start_date, end_date)

        # 2. Add Bank Interest (Summing all certificates)
        total_interest = sum(
            float(c.interest_value or 0) for c in BankCertificate.objects.all()
        )
        # total_interest = 0
        total_inc += total_interest

        # 3. Final Calculations
        net_sav = total_inc - total_exp
        sav_rate = (net_sav / total_inc * 100) if total_inc > 0 else 0

        # Category breakdown
        cat_totals = {}
        for e in expenses:
            cname = e.category.name if e.category else "Uncategorised"
            cat_totals[cname] = cat_totals.get(cname, 0) + float(e.amount)

        # ── Build PDF ──
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        navy = colors.HexColor("#000080")
        blue = colors.HexColor("#1a6ef5")
        green = colors.HexColor("#00d68f")
        red = colors.HexColor("#ff4d6d")
        yellow = colors.HexColor("#ffd166")
        grey = colors.HexColor("#7b97cc")

        H1 = ParagraphStyle(
            "H1",
            fontSize=22,
            textColor=blue,
            spaceAfter=15,
            alignment=TA_CENTER,
            fontName=pdf_font,
        )
        H11 = ParagraphStyle(
            "H11",
            fontSize=18,
            textColor=navy,
            spaceAfter=15,
            alignment=TA_CENTER,
            fontName=pdf_font,
        )
        H2 = ParagraphStyle(
            "H2",
            fontSize=14,
            textColor=navy,
            spaceAfter=4,
            spaceBefore=12,
            fontName=pdf_font,
        )
        BODY = ParagraphStyle("BODY", fontSize=10, textColor=navy, spaceAfter=4)
        SUB = ParagraphStyle("SUB", fontSize=9, textColor=grey, spaceAfter=2)

        story = []

        # Cover
        story.append(Spacer(1, 1 * cm))

        # Fetch the clean, localized text without unstable emojis
        report_text = get_text("financial_report", lang, t, "Financial Report")

        # Append the titles cleanly to the story
        story.append(Paragraph(report_text, H1))
        story.append(Paragraph(title_str, H11))
        story.append(HRFlowable(width="100%", thickness=1, color=blue))
        story.append(Spacer(1, 0.5 * cm))
        # Dynamically set table title alignments based on the document language
        table_title_style = ParagraphStyle(
            "TableTitle", parent=H2, alignment=TA_RIGHT if lang == "ar" else TA_LEFT
        )
        # Summary KPIs
        story.append(
            Paragraph(get_text("summary", lang, t, "Summary"), table_title_style)
        )

        # Define explicit Paragraph styles for table cells to handle Arabic layout flawlessly
        cell_L = ParagraphStyle(
            "CellL", fontName=pdf_font, fontSize=10, textColor=navy, alignment=TA_LEFT
        )
        cell_R = ParagraphStyle(
            "CellR", fontName=pdf_font, fontSize=10, textColor=navy, alignment=TA_RIGHT
        )
        cell_HL = ParagraphStyle(
            "CellHL",
            fontName=pdf_font,
            fontSize=10,
            textColor=colors.white,
            alignment=TA_LEFT,
        )
        cell_HR = ParagraphStyle(
            "CellHR",
            fontName=pdf_font,
            fontSize=10,
            textColor=colors.white,
            alignment=TA_RIGHT,
        )

        kpi_data = [
            [
                Paragraph(get_text("metric", lang, t, "Metric"), cell_HL),
                Paragraph(get_text("amount", lang, t, "Amount (EGP)"), cell_HR),
            ],
            [
                Paragraph(get_text("total_income", lang, t, "Total Income"), cell_L),
                Paragraph(f"{total_inc:,.2f}", cell_R),
            ],
            [
                Paragraph(
                    get_text("total_expenses", lang, t, "Total Expenses"), cell_L
                ),
                Paragraph(f"{total_exp:,.2f}", cell_R),
            ],
            [
                Paragraph(get_text("net_savings", lang, t, "Net Savings"), cell_L),
                Paragraph(
                    f"{net_sav:,.2f}",
                    ParagraphStyle(
                        "NetSavR",
                        parent=cell_R,
                        textColor=green if net_sav >= 0 else red,
                    ),
                ),
            ],
            [
                Paragraph(get_text("savings_rate", lang, t, "Savings Rate"), cell_L),
                Paragraph(f"{sav_rate:.1f}%", cell_R),
            ],
        ]

        kpi_table = Table(kpi_data, colWidths=[9 * cm, 7 * cm])
        kpi_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), blue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, -1), pdf_font),  # Applied globally
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.HexColor("#f0f4ff"), colors.white],
                    ),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1e3a6e")),
                    ("PADDING", (0, 0), (-1, -1), 7),
                ]
            )
        )
        story.append(kpi_table)
        story.append(Spacer(1, 0.5 * cm))

        # Category breakdown
        if cat_totals:
            story.append(
                Paragraph(
                    get_text("cat_breakdown", lang, t, "Expense Breakdown by Category"),
                    table_title_style,
                )
            )

            cell_L9 = ParagraphStyle(
                "CellL9",
                fontName=pdf_font,
                fontSize=9,
                textColor=navy,
                alignment=TA_LEFT,
            )
            cell_R9 = ParagraphStyle(
                "CellR9",
                fontName=pdf_font,
                fontSize=9,
                textColor=navy,
                alignment=TA_RIGHT,
            )
            cell_HL9 = ParagraphStyle(
                "CellHL9",
                fontName=pdf_font,
                fontSize=9,
                textColor=colors.white,
                alignment=TA_LEFT,
            )
            cell_HR9 = ParagraphStyle(
                "CellHR9",
                fontName=pdf_font,
                fontSize=9,
                textColor=colors.white,
                alignment=TA_RIGHT,
            )

            cat_data = [
                [
                    Paragraph(get_text("category", lang, t, "Category"), cell_HL9),
                    Paragraph(get_text("amount", lang, t, "Amount (EGP)"), cell_HR9),
                    Paragraph(get_text("pct", lang, t, "% of Total"), cell_HR9),
                ]
            ]

            for cname, ctotal in sorted(cat_totals.items(), key=lambda x: -x[1]):
                pct = (ctotal / total_exp * 100) if total_exp > 0 else 0
                # Reshape dynamic database category names if language is Arabic
                display_cname = format_arabic(cname) if lang == "ar" else cname

                cat_data.append(
                    [
                        Paragraph(display_cname, cell_L9),
                        Paragraph(f"{ctotal:,.2f}", cell_R9),
                        Paragraph(f"{pct:.1f}%", cell_R9),
                    ]
                )

            cat_data.append(
                [
                    Paragraph(get_text("total", lang, t, "TOTAL"), cell_L9),
                    Paragraph(f"{total_exp:,.2f}", cell_R9),
                    Paragraph("100%", cell_R9),
                ]
            )

            cat_table = Table(cat_data, colWidths=[9 * cm, 5 * cm, 3 * cm])
            cat_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), blue),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        (
                            "FONTNAME",
                            (0, 0),
                            (-1, -1),
                            pdf_font,
                        ),  # Fixed range to cover all cells
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8f0fe")),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -2),
                            [colors.HexColor("#f0f4ff"), colors.white],
                        ),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1e3a6e")),
                        ("PADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            story.append(cat_table)
            story.append(Spacer(1, 0.5 * cm))

        # Detailed expense entries
        if expenses:
            story.append(
                Paragraph(
                    get_text("expense_entries", lang, t, "Expense Entries"),
                    table_title_style,
                )
            )

            cell_L8 = ParagraphStyle(
                "CellL8",
                fontName=pdf_font,
                fontSize=8,
                textColor=navy,
                alignment=TA_LEFT,
            )
            cell_R8 = ParagraphStyle(
                "CellR8",
                fontName=pdf_font,
                fontSize=8,
                textColor=navy,
                alignment=TA_RIGHT,
            )
            cell_HL8 = ParagraphStyle(
                "CellHL8",
                fontName=pdf_font,
                fontSize=8,
                textColor=colors.white,
                alignment=TA_LEFT,
            )
            cell_HR8 = ParagraphStyle(
                "CellHR8",
                fontName=pdf_font,
                fontSize=8,
                textColor=colors.white,
                alignment=TA_RIGHT,
            )

            exp_data = [
                [
                    Paragraph(get_text("date", lang, t, "Date"), cell_HL8),
                    Paragraph(get_text("category", lang, t, "Category"), cell_HL8),
                    Paragraph(
                        get_text("description", lang, t, "Description"), cell_HL8
                    ),
                    Paragraph(get_text("method", lang, t, "Method"), cell_HL8),
                    Paragraph(get_text("amount", lang, t, "Amount"), cell_HR8),
                ]
            ]

            for e in sorted(expenses, key=lambda x: x.date):
                cname = e.category.name if e.category else "—"
                desc = e.description or "—"
                method = e.payment_method or "—"

                # Reshape dynamic Arabic inputs from the database if active language is Arabic
                if lang == "ar":
                    cname = format_arabic(cname)
                    desc = format_arabic(desc[:40])
                    method = format_arabic(method)
                else:
                    desc = desc[:40]

                exp_data.append(
                    [
                        Paragraph(e.date.strftime("%d/%m/%Y"), cell_L8),
                        Paragraph(cname, cell_L8),
                        Paragraph(desc, cell_L8),
                        Paragraph(method, cell_L8),
                        Paragraph(f"{float(e.amount):,.2f}", cell_R8),
                    ]
                )

            exp_table = Table(
                exp_data, colWidths=[2.5 * cm, 3.5 * cm, 6 * cm, 3 * cm, 3 * cm]
            )
            exp_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), blue),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        (
                            "FONTNAME",
                            (0, 0),
                            (-1, -1),
                            pdf_font,
                        ),  # Fixed range to cover all cells
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.HexColor("#f0f4ff"), colors.white],
                        ),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#1e3a6e")),
                        ("PADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            story.append(exp_table)

        # Footer
        story.append(Spacer(1, 1 * cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=grey))

        # 1. Extract current date components
        today = datetime.date.today()
        f_day = today.day
        f_year = today.year
        f_month_name = today.strftime("%B")  # e.g., "June"

        # 2. Look up the translation key across ALL languages matching the header logic
        f_json_key = f"month_{f_month_name.lower()}"
        f_translated_month = t.get(f_json_key)

        # Fallback handling if a language JSON file is missing the specific month key
        if not f_translated_month:
            if lang == "ar":
                ARABIC_MONTHS = {
                    "January": "يناير",
                    "February": "فبراير",
                    "March": "مارس",
                    "April": "أبريل",
                    "May": "مايو",
                    "June": "يونيو",
                    "July": "يوليو",
                    "August": "أغسطس",
                    "September": "سبتمبر",
                    "October": "أكتوبر",
                    "November": "نوفمبر",
                    "December": "ديسمبر",
                }
                f_translated_month = ARABIC_MONTHS.get(f_month_name, f_month_name)
            else:
                f_translated_month = f_month_name

        # 3. Pull the "generated_by" label from the translation context
        raw_label = t.get("generated_by", "Generated by WealthFlow")

        # 4. Construct layout string based on text direction rules
        if lang == "ar":
            # Combined with a standard hyphen, then reshaped once safely
            raw_footer = f"{raw_label} - {f_day} {f_translated_month} {f_year}"
            footer_text = format_arabic(raw_footer)
        else:
            # Handles English, French, and all other LTR languages uniformly
            footer_text = f"{raw_label} - {f_day} {f_translated_month} {f_year}"

        # 5. Append the translated paragraph to the layout story
        story.append(
            Paragraph(
                footer_text,
                ParagraphStyle(
                    "F",
                    fontSize=8,
                    textColor=grey,
                    alignment=TA_CENTER,
                    fontName=pdf_font,
                ),
            )
        )

        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # ── Profile update + avatar upload ───────────────────────────


from django.db.models.signals import post_save
from django.dispatch import receiver




@login_required
def export_excel(request):
    """Generate and download the full Balance tracker Excel workbook."""
    from core.reports.excel_generator import generate_excel
    from datetime import date

    buf = generate_excel()
    filename = f"Balance_Tracker_{date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ════════════════════════════════════════════════════════════════════════════
# REMINDER ENGINE VIEWS
# ════════════════════════════════════════════════════════════════════════════

from core.models import (
    ReminderRule,
    CertificateStatus,
    ReminderLog,
    REMINDER_TYPE_CHOICES,
    SALARY_TRIGGER_CHOICES,
)




@method_decorator(csrf_exempt, name="dispatch")
class SalaryReportView(View):
    """Salary + bonus analytics by year and company."""

    def get(self, request):
        from django.db.models import Sum, Count, Q

        year = request.GET.get("year")
        company_id = request.GET.get("company_id")

        qs = SalaryEntry.objects.all()
        if year:
            qs = qs.filter(year=int(year))
        if company_id:
            qs = qs.filter(company_id=int(company_id))

        # By year
        by_year = list(
            qs.values("year")
            .annotate(
                total_paid=Sum("paid"),
                total_bonus=Sum("bonus"),
                total_expected=Sum("expected"),
                paid_months=Count("id", filter=Q(paid__gt=0)),
            )
            .order_by("year")
        )

        # By company
        by_company = []
        for c in Company.objects.all().order_by("order"):
            cqs = qs.filter(company=c)
            agg = cqs.aggregate(
                total_paid=Sum("paid"),
                total_bonus=Sum("bonus"),
                total_expected=Sum("expected"),
                paid_months=Count("id", filter=Q(paid__gt=0)),
            )
            if agg["paid_months"]:
                by_company.append(
                    {
                        "company_id": c.id,
                        "company_name": c.display_name or c.name,
                        "color_hex": c.color_hex,
                        "total_paid": float(agg["total_paid"] or 0),
                        "total_bonus": float(agg["total_bonus"] or 0),
                        "total_expected": float(agg["total_expected"] or 0),
                        "paid_months": agg["paid_months"] or 0,
                    }
                )

        # Grand totals
        grand = qs.aggregate(
            total_paid=Sum("paid"),
            total_bonus=Sum("bonus"),
            total_expected=Sum("expected"),
            paid_months=Count("id", filter=Q(paid__gt=0)),
        )

        # Available years
        years = list(
            SalaryEntry.objects.values_list("year", flat=True)
            .distinct()
            .order_by("year")
        )
        companies = [
            {"id": c.id, "name": c.display_name or c.name}
            for c in Company.objects.all().order_by("order")
        ]

        return JsonResponse(
            {
                "by_year": by_year,
                "by_company": by_company,
                "grand": {
                    "total_paid": float(grand["total_paid"] or 0),
                    "total_bonus": float(grand["total_bonus"] or 0),
                    "total_expected": float(grand["total_expected"] or 0),
                    "paid_months": grand["paid_months"] or 0,
                },
                "years": years,
                "companies": companies,
            }
        )




@method_decorator(csrf_exempt, name="dispatch")
class BalanceReportView(View):
    """Balance summary across banks and currencies."""

    def get(self, request):
        _run_certificate_interest_sync()
        from django.db.models import Sum

        entries = BalanceEntry.objects.select_related("bank", "currency").all()
        banks = Bank.objects.all()

        # Group by bank
        by_bank = []
        for bank in banks:
            bank_entries = entries.filter(bank=bank)
            total_egp = float(
                bank_entries.filter(currency__code="EGP").aggregate(s=Sum("amount"))[
                    "s"
                ]
                or 0
            )
            by_bank.append(
                {
                    "bank_id": bank.id,
                    "bank_name": bank.name,
                    "total_egp": total_egp,
                    "entries": [e.to_dict() for e in bank_entries],
                }
            )

        # Unbanked entries (cash / home)
        home = entries.filter(bank__isnull=True)
        by_currency = []
        for e in home:
            by_currency.append(e.to_dict())

        net_worth_data = NetWorthService().portfolio_components()
        cert_total = float(net_worth_data["certificate_total_egp"])
        cert_interest_total = float(net_worth_data["certificate_interest_total_egp"])

        cert_monthly_interest = cert_interest_total if cert_interest_total else 0.0

        return JsonResponse(
            {
                "by_bank": by_bank,
                "home_entries": by_currency,
                "cert_total": cert_total,
                "cert_interest": cert_monthly_interest,
                "cert_interest_total": cert_interest_total,
                "fixed_assets_total": float(net_worth_data["fixed_assets_total_egp"]),
                "net_worth": float(net_worth_data["net_worth_egp"]),
            }
        )




@method_decorator(csrf_exempt, name="dispatch")
class CertificateReportView(View):
    """Certificate maturity and analytics report."""

    def get(self, request):
        from datetime import date, timedelta
        from django.db.models import Sum, Count

        today = date.today()
        active_certs = BankCertificate.objects.select_related("bank", "currency").filter(
            status__iexact="active"
        )

        agg = active_certs.aggregate(
            total_count=Count("id"),
            total_amount=Sum("amount"),
            total_interest=Sum("interest_value"),
        )

        # Maturity buckets (configurable label from settings, days from AppSettings)
        bucket_days = [
            ("overdue", 0, -1),
            ("30_days", 0, 30),
            ("90_days", 31, 90),
            ("180_days", 91, 180),
            ("later", 181, 9999),
        ]

        buckets = {}
        for label, low, high in bucket_days:
            if label == "overdue":
                buckets[label] = [
                    c.to_dict()
                    for c in active_certs.filter(expiry_date__lt=today)
                ]
            else:
                buckets[label] = [
                    c.to_dict()
                    for c in active_certs.filter(
                        expiry_date__gte=today + timedelta(days=low),
                        expiry_date__lte=today + timedelta(days=high),
                    )
                ]

        # By status (active certificates only)
        by_status = {}
        for c in active_certs:
            by_status[c.status] = by_status.get(c.status, {"count": 0, "total": 0})
            by_status[c.status]["count"] += 1
            by_status[c.status]["total"] += float(c.amount)

        # Monthly interest cashflow (next 12 months) for active certificates only
        monthly_cf = []
        for i in range(12):
            m_start = today.replace(day=1) + timedelta(days=32 * i)
            m_start = m_start.replace(day=1)
            m_certs = active_certs.filter(
                expiry_date__year=m_start.year,
                expiry_date__month=m_start.month,
            )
            monthly_cf.append(
                {
                    "month": m_start.strftime("%b %Y"),
                    "count": m_certs.count(),
                    "amount": float(m_certs.aggregate(s=Sum("amount"))["s"] or 0),
                }
            )

        total_interest = float(agg["total_interest"] or 0)
        # Keep monthly interest aligned with historical report expectations.
        # VS Code, don't this again to (total_interest / 12.0)
        monthly_interest = (total_interest) if total_interest else 0.0

        return JsonResponse(
            {
                "summary": {
                    "total_count": agg["total_count"] or 0,
                    "total_amount": float(agg["total_amount"] or 0),
                    "total_interest": total_interest,
                    "monthly_interest": monthly_interest,
                },
                "buckets": buckets,
                "by_status": by_status,
                "monthly_cf": monthly_cf,
            }
        )




def _fixed_asset_report_queryset():
    return (
        FixedAsset.objects.select_related(
            "real_estate",
            "vehicle_details",
            "gold_details",
            "other_asset_details",
            "sale",
        )
        .prefetch_related(
            "photos",
            "renovations",
            "maintenance",
            "insurance",
            "furniture",
            "valuation_history",
        )
        .order_by("name")
    )




def _fixed_asset_report_context(request):
    scope = request.GET.get("scope", "single")
    asset_id = request.GET.get("asset_id")
    lang = request.GET.get("lang", "en")
    t = get_translations(lang)

    queryset = _fixed_asset_report_queryset()

    if scope == "single":
        if not asset_id:
            raise ValueError("asset_id is required")
        queryset = queryset.filter(pk=asset_id)

    assets = list(queryset)
    if not assets:
        raise FixedAsset.DoesNotExist()

    return {
        "scope": scope,
        "asset_id": asset_id,
        "lang": lang,
        "t": t,
        "assets": assets,
        "portfolio_snapshot": NetWorthService().fixed_assets_snapshot(),
    }




def _fixed_asset_display_value(value):
    if value in (None, "", []):
        return "-"
    return str(value)




def _fixed_asset_report_label(t, lang, key, default):
    return get_text(key, lang, t, default)




def _fixed_asset_user_text(value, lang):
    if value in (None, ""):
        return "-"
    text = str(value)
    has_arabic = any("\u0600" <= ch <= "\u06FF" for ch in text)
    return format_arabic(text) if lang == "ar" or has_arabic else text




def _fixed_asset_pdf_table(rows, col_widths, font_name):
    table = Table(rows, colWidths=col_widths, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9e1f2")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1f2937")),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#94a3b8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table




def _build_fixed_asset_pdf_story(asset, lang, t, styles, title_style, heading_style, body_style, font_name):
    data = asset.to_dict()
    story = []

    asset_name = _fixed_asset_user_text(asset.name, lang)
    story.append(Paragraph(asset_name, title_style))
    story.append(Spacer(1, 0.25 * cm))

    gain_amount = float(data.get("current_market_value") or 0) - float(data.get("purchase_price") or 0)
    general_rows = [
        [
            _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
            asset_name,
        ],
        [
            _fixed_asset_report_label(t, lang, "asset_type", "Asset Type"),
            _fixed_asset_report_label(
                t,
                lang,
                f"type_{str(data.get('asset_type') or 'other').lower()}",
                data.get("asset_type") or "-",
            ),
        ],
        [
            _fixed_asset_report_label(t, lang, "status", "Status"),
            _fixed_asset_report_label(
                t,
                lang,
                str(data.get("status") or "owned").lower(),
                data.get("status") or "-",
            ),
        ],
        [
            _fixed_asset_report_label(t, lang, "purchase_date", "Purchase Date"),
            _fixed_asset_display_value(data.get("purchase_date")),
        ],
        [
            _fixed_asset_report_label(t, lang, "purchase_price_egp", "Purchase Price (EGP)"),
            f"{float(data.get('purchase_price') or 0):,.2f}",
        ],
        [
            _fixed_asset_report_label(t, lang, "current_market_value", "Current Market Value"),
            f"{float(data.get('current_market_value') or 0):,.2f}",
        ],
        [
            _fixed_asset_report_label(t, lang, "gain_amount", "Gain Amount"),
            f"{gain_amount:,.2f}",
        ],
        [
            _fixed_asset_report_label(t, lang, "notes", "Notes"),
            _fixed_asset_user_text(data.get("notes"), lang),
        ],
    ]

    story.append(Paragraph(_fixed_asset_report_label(t, lang, "general_information", "General Information"), heading_style))
    story.append(_fixed_asset_pdf_table(general_rows, [5 * cm, 10.5 * cm], font_name))
    story.append(Spacer(1, 0.3 * cm))

    real_estate = data.get("real_estate") or {}
    if real_estate:
        property_rows = [
            [_fixed_asset_report_label(t, lang, "country", "Country"), _fixed_asset_user_text(real_estate.get("country"), lang)],
            [_fixed_asset_report_label(t, lang, "governorate", "Governorate"), _fixed_asset_user_text(real_estate.get("governorate"), lang)],
            [_fixed_asset_report_label(t, lang, "city", "City"), _fixed_asset_user_text(real_estate.get("city"), lang)],
            [_fixed_asset_report_label(t, lang, "district", "District"), _fixed_asset_user_text(real_estate.get("district"), lang)],
            [_fixed_asset_report_label(t, lang, "address", "Address"), _fixed_asset_user_text(real_estate.get("address"), lang)],
            [_fixed_asset_report_label(t, lang, "apt_area", "Property Area (Sqm)"), _fixed_asset_display_value(real_estate.get("apartment_area"))],
            [_fixed_asset_report_label(t, lang, "land_area", "Land Area"), _fixed_asset_display_value(real_estate.get("land_area"))],
            [_fixed_asset_report_label(t, lang, "rooms", "Bedrooms"), _fixed_asset_display_value(real_estate.get("rooms"))],
            [_fixed_asset_report_label(t, lang, "bathrooms", "Bathrooms"), _fixed_asset_display_value(real_estate.get("bathrooms"))],
            [_fixed_asset_report_label(t, lang, "description", "Description"), _fixed_asset_user_text(real_estate.get("description"), lang)],
        ]
        story.append(Paragraph(_fixed_asset_report_label(t, lang, "property_details", "Property Details"), heading_style))
        story.append(_fixed_asset_pdf_table(property_rows, [5 * cm, 10.5 * cm], font_name))
        story.append(Spacer(1, 0.3 * cm))

    photos = list(asset.photos.all())
    if photos:
        story.append(Paragraph(_fixed_asset_report_label(t, lang, "photos", "Photos"), heading_style))
        image_rows = []
        current_row = []
        for photo in photos[:4]:
            try:
                img = RLImage(io.BytesIO(photo.image_data), width=6 * cm, height=4.5 * cm)
                current_row.append(img)
            except Exception:
                current_row.append(Paragraph(_fixed_asset_user_text(photo.filename or photo.title or photo.id, lang), body_style))
            if len(current_row) == 2:
                image_rows.append(current_row)
                current_row = []
        if current_row:
            while len(current_row) < 2:
                current_row.append(Paragraph("", body_style))
            image_rows.append(current_row)
        story.append(Table(image_rows, colWidths=[7.7 * cm, 7.7 * cm], hAlign="LEFT"))
        story.append(Spacer(1, 0.3 * cm))

    def build_collection_section(title_key, title_default, items, headers, value_rows):
        if not items:
            return
        story.append(Paragraph(_fixed_asset_report_label(t, lang, title_key, title_default), heading_style))
        rows = [[_fixed_asset_report_label(t, lang, key, default) for key, default in headers]]
        rows.extend(value_rows(item) for item in items)
        story.append(_fixed_asset_pdf_table(rows, [4 * cm, 4 * cm, 3.5 * cm, 4 * cm], font_name))
        story.append(Spacer(1, 0.3 * cm))

    build_collection_section(
        "renovations",
        "Renovations",
        data.get("renovations") or [],
        [("date", "Date"), ("category", "Category"), ("amount_egp", "Amount EGP"), ("notes", "Notes")],
        lambda item: [
            _fixed_asset_display_value(item.get("date")),
            _fixed_asset_user_text(item.get("category"), lang),
            f"{float(item.get('amount_egp') or 0):,.2f}",
            _fixed_asset_user_text(item.get("notes"), lang),
        ],
    )

    build_collection_section(
        "furniture",
        "Furniture",
        data.get("furniture") or [],
        [("asset_name", "Name"), ("category", "Category"), ("amount_egp", "Amount EGP"), ("notes", "Notes")],
        lambda item: [
            _fixed_asset_user_text(item.get("name"), lang),
            _fixed_asset_user_text(item.get("category"), lang),
            f"{float(item.get('amount_egp') or 0):,.2f}",
            _fixed_asset_user_text(item.get("notes"), lang),
        ],
    )

    build_collection_section(
        "valuation_history",
        "Valuation History",
        data.get("valuation_history") or [],
        [("date", "Date"), ("current_market_value", "Market Value"), ("valuation_source", "Valuation Source"), ("notes", "Notes")],
        lambda item: [
            _fixed_asset_display_value(item.get("valuation_date")),
            f"{float(item.get('market_value') or 0):,.2f}",
            _fixed_asset_user_text(item.get("valuation_source"), lang),
            _fixed_asset_user_text(item.get("notes"), lang),
        ],
    )

    sale = data.get("sale") or None
    if sale:
        sale_rows = [
            [_fixed_asset_report_label(t, lang, "sale_date", "Sale Date"), _fixed_asset_display_value(sale.get("sale_date"))],
            [_fixed_asset_report_label(t, lang, "sale_price_egp", "Sale Price (EGP)"), f"{float(sale.get('sale_price') or 0):,.2f}"],
            [_fixed_asset_report_label(t, lang, "selling_expenses_egp", "Selling Expenses (EGP)"), f"{float(sale.get('selling_expenses') or 0):,.2f}"],
            [_fixed_asset_report_label(t, lang, "net_sale_amount", "Net Sale Amount"), f"{float(sale.get('net_sale_amount') or 0):,.2f}"],
            [_fixed_asset_report_label(t, lang, "deposit_balance", "Deposit Balance"), _fixed_asset_display_value(sale.get("deposit_balance_id"))],
            [_fixed_asset_report_label(t, lang, "notes", "Notes"), _fixed_asset_user_text(sale.get("notes"), lang)],
        ]
        story.append(Paragraph(_fixed_asset_report_label(t, lang, "sale_information", "Sale Information"), heading_style))
        story.append(_fixed_asset_pdf_table(sale_rows, [5 * cm, 10.5 * cm], font_name))
        story.append(Spacer(1, 0.3 * cm))

    return story




class FixedAssetPdfReportGenerator(object):

    def get(self, request):
        try:
            context = _fixed_asset_report_context(request)
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        except FixedAsset.DoesNotExist:
            return JsonResponse({"error": "No fixed assets found"}, status=404)

        lang = context["lang"]
        t = context["t"]
        assets = context["assets"]
        scope = context["scope"]
        portfolio_snapshot = context.get("portfolio_snapshot") or {}

        font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "arial.ttf")
        font_exists = os.path.exists(font_path)
        if font_exists:
            pdfmetrics.registerFont(TTFont("ArabicFont", font_path))

        # Always prefer Arabic-capable font when available so mixed-language
        # content (e.g., Arabic names in EN report) renders correctly.
        font_name = "ArabicFont" if font_exists else "Helvetica"
        font_name_bold = "ArabicFont" if font_name == "ArabicFont" else "Helvetica-Bold"

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "FixedAssetTitle",
            parent=styles["Heading1"],
            fontName=font_name_bold,
            fontSize=16,
            textColor=colors.HexColor("#1f2937"),
            spaceAfter=8,
        )
        heading_style = ParagraphStyle(
            "FixedAssetHeading",
            parent=styles["Heading2"],
            fontName=font_name_bold,
            fontSize=12,
            textColor=colors.HexColor("#1a6ef5"),
            spaceBefore=4,
            spaceAfter=6,
        )
        body_style = ParagraphStyle(
            "FixedAssetBody",
            parent=styles["BodyText"],
            fontName=font_name,
            fontSize=9,
            textColor=colors.HexColor("#1f2937"),
            leading=11,
        )

        report_title = _fixed_asset_report_label(
            t,
            lang,
            "fixed_assets_report_title",
            "Fixed Assets Report",
        )
        if scope == "single":
            report_title = f"{report_title} - {_fixed_asset_user_text(assets[0].name, lang)}"

        story = [Paragraph(report_title, title_style), Spacer(1, 0.35 * cm)]

        if scope == "portfolio":
            summary_rows = [
                [
                    _fixed_asset_report_label(t, lang, "total_fixed_assets_value", "Total Fixed Assets"),
                    f"{float(portfolio_snapshot.get('total_fixed_assets_value') or 0):,.2f}",
                ],
                [
                    _fixed_asset_report_label(t, lang, "net_worth", "Net Worth"),
                    f"{float(portfolio_snapshot.get('total_net_worth') or 0):,.2f}",
                ],
                [
                    _fixed_asset_report_label(t, lang, "net_worth_contribution", "Net Worth Contribution"),
                    f"{float(portfolio_snapshot.get('net_worth_contribution') or 0):,.2f}%",
                ],
            ]
            story.append(Paragraph(_fixed_asset_report_label(t, lang, "portfolio_distribution", "Portfolio Distribution"), heading_style))
            story.append(_fixed_asset_pdf_table(summary_rows, [7 * cm, 8.5 * cm], font_name))
            story.append(Spacer(1, 0.35 * cm))

        for index, asset in enumerate(assets):
            story.extend(
                _build_fixed_asset_pdf_story(
                    asset,
                    lang,
                    t,
                    styles,
                    title_style,
                    heading_style,
                    body_style,
                    font_name,
                )
            )
            if index < len(assets) - 1:
                story.append(PageBreak())

        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        filename = (
            f"fixed_asset_{assets[0].id}_report.pdf"
            if scope == "single"
            else "fixed_assets_portfolio_report.pdf"
        )
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response




class FixedAssetExcelReportGenerator(object):

    def get(self, request):
        try:
            context = _fixed_asset_report_context(request)
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        except FixedAsset.DoesNotExist:
            return JsonResponse({"error": "No fixed assets found"}, status=404)

        lang = context["lang"]
        t = context["t"]
        assets = context["assets"]
        scope = context["scope"]
        portfolio_snapshot = context.get("portfolio_snapshot") or {}

        wb = openpyxl.Workbook()
        summary_ws = wb.active
        if summary_ws is None:
            summary_ws = wb.create_sheet(title="Summary")
        else:
            summary_ws.title = "Summary"

        header_font = Font(bold=True)

        summary_headers = [
            _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
            _fixed_asset_report_label(t, lang, "asset_type", "Asset Type"),
            _fixed_asset_report_label(t, lang, "status", "Status"),
            _fixed_asset_report_label(t, lang, "purchase_date", "Purchase Date"),
            _fixed_asset_report_label(t, lang, "purchase_price_egp", "Purchase Price (EGP)"),
            _fixed_asset_report_label(t, lang, "current_market_value", "Current Market Value"),
            _fixed_asset_report_label(t, lang, "country", "Country"),
            _fixed_asset_report_label(t, lang, "city", "City"),
            _fixed_asset_report_label(t, lang, "address", "Address"),
            _fixed_asset_report_label(t, lang, "sale_date", "Sale Date"),
            _fixed_asset_report_label(t, lang, "net_sale_amount", "Net Sale Amount"),
            _fixed_asset_report_label(t, lang, "notes", "Notes"),
        ]
        summary_ws.append(summary_headers)
        for cell in summary_ws[1]:
            cell.font = header_font

        for asset in assets:
            data = asset.to_dict()
            real_estate = data.get("real_estate") or {}
            sale = data.get("sale") or {}
            summary_ws.append(
                [
                    data.get("name"),
                    data.get("asset_type"),
                    data.get("status"),
                    data.get("purchase_date"),
                    float(data.get("purchase_price") or 0),
                    float(data.get("current_market_value") or 0),
                    real_estate.get("country"),
                    real_estate.get("city"),
                    real_estate.get("address"),
                    sale.get("sale_date"),
                    float(sale.get("net_sale_amount") or 0),
                    data.get("notes"),
                ]
            )

        if scope == "portfolio":
            summary_ws.append([])
            summary_ws.append([
                _fixed_asset_report_label(t, lang, "total_fixed_assets_value", "Total Fixed Assets"),
                float(portfolio_snapshot.get("total_fixed_assets_value") or 0),
            ])
            summary_ws.append([
                _fixed_asset_report_label(t, lang, "net_worth", "Net Worth"),
                float(portfolio_snapshot.get("total_net_worth") or 0),
            ])
            summary_ws.append([
                _fixed_asset_report_label(t, lang, "net_worth_contribution", "Net Worth Contribution"),
                float(portfolio_snapshot.get("net_worth_contribution") or 0),
            ])

        collections = [
            (
                "Renovations",
                _fixed_asset_report_label(t, lang, "renovations", "Renovations"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "date", "Date"),
                    _fixed_asset_report_label(t, lang, "category", "Category"),
                    _fixed_asset_report_label(t, lang, "amount_egp", "Amount EGP"),
                    _fixed_asset_report_label(t, lang, "notes", "Notes"),
                ],
                lambda asset_data, item: [
                    asset_data.get("name"),
                    item.get("date"),
                    item.get("category"),
                    float(item.get("amount_egp") or 0),
                    item.get("notes"),
                ],
                lambda asset_data: asset_data.get("renovations") or [],
            ),
            (
                "Furniture",
                _fixed_asset_report_label(t, lang, "furniture", "Furniture"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "category", "Category"),
                    _fixed_asset_report_label(t, lang, "purchase_date", "Purchase Date"),
                    _fixed_asset_report_label(t, lang, "amount_egp", "Amount EGP"),
                    _fixed_asset_report_label(t, lang, "notes", "Notes"),
                ],
                lambda asset_data, item: [
                    item.get("name"),
                    item.get("category"),
                    item.get("purchase_date"),
                    float(item.get("amount_egp") or 0),
                    item.get("notes"),
                ],
                lambda asset_data: asset_data.get("furniture") or [],
            ),
            (
                "Valuations",
                _fixed_asset_report_label(t, lang, "valuation_history", "Valuation History"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "date", "Date"),
                    _fixed_asset_report_label(t, lang, "current_market_value", "Market Value"),
                    _fixed_asset_report_label(t, lang, "valuation_source", "Valuation Source"),
                    _fixed_asset_report_label(t, lang, "notes", "Notes"),
                ],
                lambda asset_data, item: [
                    asset_data.get("name"),
                    item.get("valuation_date"),
                    float(item.get("market_value") or 0),
                    item.get("valuation_source"),
                    item.get("notes"),
                ],
                lambda asset_data: asset_data.get("valuation_history") or [],
            ),
            (
                "Photos",
                _fixed_asset_report_label(t, lang, "photos", "Photos"),
                [
                    _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
                    _fixed_asset_report_label(t, lang, "description", "Description"),
                    _fixed_asset_report_label(t, lang, "notes", "Filename"),
                    "URL",
                ],
                lambda asset_data, item: [
                    asset_data.get("name"),
                    item.get("title"),
                    item.get("filename"),
                    item.get("url"),
                ],
                lambda asset_data: asset_data.get("photos") or [],
            ),
        ]

        sale_ws = wb.create_sheet(title="Sale")
        sale_headers = [
            _fixed_asset_report_label(t, lang, "asset_name", "Asset Name"),
            _fixed_asset_report_label(t, lang, "sale_date", "Sale Date"),
            _fixed_asset_report_label(t, lang, "sale_price_egp", "Sale Price (EGP)"),
            _fixed_asset_report_label(t, lang, "selling_expenses_egp", "Selling Expenses (EGP)"),
            _fixed_asset_report_label(t, lang, "net_sale_amount", "Net Sale Amount"),
            _fixed_asset_report_label(t, lang, "deposit_balance", "Deposit Balance"),
            _fixed_asset_report_label(t, lang, "notes", "Notes"),
        ]
        sale_ws.append(sale_headers)
        for cell in sale_ws[1]:
            cell.font = header_font

        for asset in assets:
            asset_data = asset.to_dict()
            sale = asset_data.get("sale")
            if not sale:
                continue
            sale_ws.append(
                [
                    asset_data.get("name"),
                    sale.get("sale_date"),
                    float(sale.get("sale_price") or 0),
                    float(sale.get("selling_expenses") or 0),
                    float(sale.get("net_sale_amount") or 0),
                    sale.get("deposit_balance_id"),
                    sale.get("notes"),
                ]
            )

        for sheet_name, title, headers, row_builder, collection_getter in collections:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
            for asset in assets:
                asset_data = asset.to_dict()
                for item in collection_getter(asset_data):
                    ws.append(row_builder(asset_data, item))

        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = (
            f"fixed_asset_{assets[0].id}_report.xlsx"
            if scope == "single"
            else "fixed_assets_portfolio_report.xlsx"
        )
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response