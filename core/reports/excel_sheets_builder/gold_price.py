# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
"""NOTE: Part of the excel_sheets_builder package split (files >200 lines are
split into packages per WealthFlow refactoring convention). This file holds
the "Gold Price" sheet builder only.
"""
from openpyxl.styles import Font

from core.reports.excel_formatting_helpers import (
    FMT_EGP_RED,
    WHITE,
    _f,
    _thin,
    _fill,
    _center,
    _apply_zebra_striping,
)


def build_gold_price_sheet(ws, gold_qs, balance_entries):
    ws.column_dimensions["A"].width = 18.6
    ws.column_dimensions["B"].width = 10.7
    ws.column_dimensions["C"].width = 10.7
    ws.column_dimensions["D"].width = 10.7
    ws.column_dimensions["E"].width = 10.7
    ws.column_dimensions["F"].width = 32.7
    ws.column_dimensions["G"].width = 15.8
    ws.column_dimensions["H"].width = 14.3

    FILL_BLACK = _fill("FF000000")
    latest = gold_qs.order_by("-fetched_at").first()

    for c, h in enumerate(["السعر", "شراء", "بيع", "المزيد", "الملاحظات"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial")
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    carats = [
        ("جرام عيار 24", "carat_24k_buy", "carat_24k", "40 ج"),
        ("جرام عيار 22", "carat_22k_buy", "carat_22k", "37 ج"),
        ("جرام عيار 21", "carat_21k_buy", "carat_21k", "35 ج"),
        ("جرام عيار 18", "carat_18k_buy", "carat_18k", "30 ج"),
        ("جرام عيار 14 27 ج", "carat_18k_buy", "carat_18k", "27 ج"),
    ]
    for i, (label, bf, sf, karat) in enumerate(carats, 2):
        _apply_zebra_striping(ws, i, 5)
        for col, val in enumerate(
            [
                label,
                round(float(getattr(latest, bf, 0)), 0) if latest else 0,
                round(float(getattr(latest, sf, 0)), 0) if latest else 0,
                ">",
                karat,
            ],
            1,
        ):
            c = ws.cell(row=i, column=col, value=val)
            c.font = _f(name="Arial")
            c.alignment = _center()
            c.border = _thin()

    rows_data = [
        [
            "الدولار 0 ج",
            float(latest.usd_to_egp) if latest else 0,
            float(latest.usd_to_egp) if latest else 0,
            None,
            "0 ج",
        ],
        [
            "الأونصة 0 $",
            float(latest.usd_per_oz) if latest else 0,
            float(latest.usd_per_oz) if latest else 0,
            None,
            "0 $",
        ],
        [
            "الجنيه الذهب 320 ج",
            round(float(latest.carat_21k_buy) * 8, 0) if latest else 0,
            round(float(latest.carat_21k) * 8, 0) if latest else 0,
            None,
            "320 ج",
        ],
    ]
    for i, row_data in enumerate(rows_data, 7):
        _apply_zebra_striping(ws, i, 5)
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = _f(name="Arial")
            c.alignment = _center()
            c.border = _thin()

    from core.models import Currency

    grams = 0
    try:
        gold_cur = Currency.objects.get(code="Gold")
        grams = sum(
            float(be.amount)
            for be in balance_entries
            if be.bank_id is None and be.currency_id == gold_cur.id
        )
    except Exception:
        pass

    ws.merge_cells("G9:I9")
    c_title = ws.cell(row=9, column=7, value=f"{int(grams)} Grams")
    c_title.font = _f(name="Arial", bold=True)
    c_title.alignment = _center()
    c_title.border = _thin()
    ws.cell(row=9, column=8).border = _thin()
    ws.cell(row=9, column=9).border = _thin()

    for c, title in enumerate(["Now", "Paid", "Diff"], 7):
        c_head = ws.cell(row=10, column=c, value=title)
        c_head.font = _f(name="Arial", bold=True)
        c_head.alignment = _center()
        c_head.border = _thin()

    vals = ["=(C2+28.5)*(BALANCE!F2)", 897375, "=G11-H11"]
    for c, val in enumerate(vals, 7):
        c_val = ws.cell(row=11, column=c, value=val)
        c_val.font = _f(name="Arial")
        c_val.alignment = _center()
        c_val.border = _thin()
        c_val.number_format = FMT_EGP_RED
