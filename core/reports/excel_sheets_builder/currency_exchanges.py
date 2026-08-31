# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
"""NOTE: Part of the excel_sheets_builder package split (files >200 lines are
split into packages per WealthFlow refactoring convention). This file holds
the "Currency Exchanges" sheet builder only.
"""
from openpyxl.styles import Font

from core.reports.excel_formatting_helpers import (
    WHITE,
    _thin,
    _fill,
    _center,
    _apply_zebra_striping,
)


def build_currency_exchanges_sheet(ws, exchanges_qs, lang="ar"):
    from core.reports.report_utils import get_translations, get_text

    t_map = get_translations(lang)
    FILL_BLACK = _fill("FF000000")

    header_keys = [
        ("date", "Date"),
        ("from_balance", "Source Balance"),
        ("from_currency", "From Currency"),
        ("from_amount", "Amount Exchanged"),
        ("to_balance", "Destination Balance"),
        ("to_currency", "To Currency"),
        ("to_amount", "Amount Received"),
        ("exchange_rate", "Exchange Rate"),
        ("status", "Status"),
        ("notes", "Notes"),
    ]
    headers = [get_text(key, lang, t_map, default) for key, default in header_keys]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, size=11, name="Arial", color=WHITE)
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    exchanges = list(
        exchanges_qs.select_related(
            "from_balance", "from_currency", "to_balance", "to_currency"
        ).order_by("-exchange_date", "-id")
    )
    status_keys = {
        "ACTIVE": ("status_active", "Active"),
        "REVERSED": ("status_reversed", "Reversed"),
        "EDITED": ("status_edited", "Edited"),
    }

    for i, ex in enumerate(exchanges, 2):
        _apply_zebra_striping(ws, i, len(headers))
        st_key, st_def = status_keys.get(ex.status, ("status", ex.status))
        st_label = get_text(st_key, lang, t_map, st_def)

        ws.cell(row=i, column=1, value=str(ex.exchange_date))
        ws.cell(row=i, column=2, value=ex.from_balance.title if ex.from_balance else "")
        ws.cell(row=i, column=3, value=ex.from_currency.code if ex.from_currency else "")
        ws.cell(row=i, column=4, value=float(ex.from_amount))
        ws.cell(row=i, column=5, value=ex.to_balance.title if ex.to_balance else "")
        ws.cell(row=i, column=6, value=ex.to_currency.code if ex.to_currency else "")
        ws.cell(row=i, column=7, value=float(ex.to_amount))
        ws.cell(row=i, column=8, value=float(ex.exchange_rate))
        ws.cell(row=i, column=9, value=st_label)
        ws.cell(row=i, column=10, value=ex.notes or "")
