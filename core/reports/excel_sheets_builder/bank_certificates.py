# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
"""NOTE: Part of the excel_sheets_builder package split (files >200 lines are
split into packages per WealthFlow refactoring convention). This file holds
the "Bank-Certificates" sheet builder only.
"""
from openpyxl.styles import Font

from core.reports.excel_formatting_helpers import (
    FMT_EGP_CERT,
    FMT_EGP_CERT_R,
    FMT_PCT,
    FMT_DATE,
    WHITE,
    _thin,
    _fill,
    _center,
    _apply_zebra_striping,
)


def build_bank_certificates_sheet(ws, certs_qs):
    FILL_BLACK = _fill("FF000000")
    hdrs = [
        "Amount",
        "Interest Rate",
        "Interest Value",
        "Frequency",
        "Start Date",
        "End Date",
    ]

    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial")
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    for i, cert in enumerate(certs_qs.filter(status__iexact="active").order_by("issue_date"), 2):
        _apply_zebra_striping(ws, i, 6)
        ws.cell(row=i, column=1, value=float(cert.amount)).number_format = FMT_EGP_CERT
        ws.cell(row=i, column=2, value=float(cert.interest_rate) / 100).number_format = (
            FMT_PCT
        )
        ws.cell(row=i, column=3, value=f"=(A{i}*B{i})/12").number_format = (
            FMT_EGP_CERT_R
        )
        ws.cell(row=i, column=4, value=cert.frequency)
        ws.cell(row=i, column=5, value=cert.issue_date).number_format = FMT_DATE
        ws.cell(row=i, column=6, value=cert.expiry_date).number_format = FMT_DATE
