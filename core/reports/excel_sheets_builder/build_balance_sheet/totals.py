# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
"""NOTE: Part of the excel_sheets_builder/build_balance_sheet subfolder
(promoted because the original build_balance_sheet function was >200 lines
on its own, per WealthFlow refactoring convention). This file holds the
totals phase: certificate total row, EGP total row, all-balances total row,
and the company pay/work-months totals, continuing from ctx.excel_row.
"""
from openpyxl.styles import Alignment, Border, Side

from core.reports.excel_formatting_helpers import FMT_EGP_RED, _f, _thin


def apply_totals(ctx):
    ws = ctx.ws
    excel_row = ctx.excel_row

    from core.models import BankCertificate

    cert_count = BankCertificate.objects.filter(status__iexact="active").count()
    cr = excel_row
    cert_entries = [
        be
        for be in ctx.balance_entries
        if str(be.balance_type).strip().lower() == "certificate"
    ]
    cert_title = cert_entries[0].title if cert_entries else "Certificates Balance"

    ws.cell(row=cr, column=1, value=cert_title).font = _f(bold=True, name="Arial")
    ws.cell(row=cr, column=1).border = _thin()
    bc = ws.cell(
        row=cr, column=2, value=f"=SUM('Bank-Certificates'!A2:A{cert_count+1})"
    )
    bc.font = _f(bold=True, name="Arial")
    bc.border = _thin()
    bc.number_format = FMT_EGP_RED
    excel_row += 1

    ter = excel_row
    ws.cell(row=ter, column=1, value="Total EGP Balance").font = _f(
        bold=True, name="Arial"
    )
    ws.cell(row=ter, column=1).border = _thin()
    te = ws.cell(row=ter, column=2, value=f"=SUM(B2:B{cr})")
    te.font = _f(bold=True, name="Arial")
    te.border = _thin()
    te.number_format = FMT_EGP_RED
    excel_row += 1

    tar = excel_row
    ws.row_dimensions[tar].height = 18.0
    ws.merge_cells(f"B{tar}:F{tar}")

    ws.cell(row=tar, column=1, value="Total all Balances").font = _f(
        bold=True, name="Arial"
    )
    ws.cell(row=tar, column=1).border = _thin()

    formula = (
        f"=B{ter}"
        f"+(C2*('Exchange Rates'!B2))"
        f"+(D2*('Exchange Rates'!B3))"
        f"+(E2*('Exchange Rates'!B11))"
        f"+(F2*(('Gold Price'!C2)+28.5))"
    )

    ta = ws.cell(row=tar, column=2, value=formula)
    ta.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(2, 7):
        ws.cell(row=tar, column=col).border = thin_border

    ta.font = _f(bold=True, name="Arial")
    ta.number_format = FMT_EGP_RED
    excel_row += 1

    tpr = excel_row + 3
    tmr = tpr + 1

    BONUS_COMPANIES = set(ctx.company_sheet_rows.keys())

    pay_parts, month_parts = [], []
    for cname, (sname, srow) in ctx.company_sheet_rows.items():
        ref = f"'{sname}'!{{c}}{srow}" if " " in sname else f"{sname}!{{c}}{srow}"
        if cname in BONUS_COMPANIES:
            pay_parts.append(f"({ref.format(c='D')}+{ref.format(c='F')})")
        else:
            pay_parts.append(ref.format(c="D"))
        month_parts.append(ref.format(c="B"))

    label_tp = ws.cell(row=tpr, column=1, value="Total Pays")
    label_tp.font = _f(bold=True, name="Arial")
    label_tp.border = _thin()

    tp = ws.cell(row=tpr, column=2, value="=" + "+".join(pay_parts) if pay_parts else 0)
    tp.font = _f(bold=True, name="Arial")
    tp.border = _thin()
    tp.number_format = FMT_EGP_RED

    label_tm = ws.cell(row=tmr, column=1, value="Total Work Months")
    label_tm.font = _f(bold=True, name="Arial")
    label_tm.border = _thin()

    tm = ws.cell(
        row=tmr, column=2, value="=" + "+".join(month_parts) if month_parts else 0
    )
    tm.font = _f(bold=True, name="Arial")
    tm.border = _thin()

    ctx.excel_row = excel_row
