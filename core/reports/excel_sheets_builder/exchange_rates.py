# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
"""NOTE: Part of the excel_sheets_builder package split (files >200 lines are
split into packages per WealthFlow refactoring convention). This file holds
the "Exchange Rates" sheet builder only.
"""
from openpyxl.styles import Font

from core.reports.excel_formatting_helpers import (
    FMT_EGP_RED,
    FMT_USD,
    FMT_EUR,
    WHITE,
    _f,
    _thin,
    _fill,
    _center,
    _apply_zebra_striping,
)

CURRENCIES = [
    ("USD", "دولار أمريكى"),
    ("EUR", "يورو"),
    ("GBP", "جنيَــه إسترليـنى"),
    ("CAD", "دولار كنـدى"),
    ("DKK", "كرون دانمركى"),
    ("NOK", "كرون نرويجى"),
    ("SEK", "كرون سَــويدى"),
    ("CHF", "فرنك سويسرى"),
    ("JPY", "100 ين يابانى"),
    ("SAR", "ريـــال سعــودى"),
    ("KWD", "دينــار كويتى"),
    ("AED", "درهم اماراتى"),
    ("AUD", "دولار اســـترالى"),
    ("BHD", "دينــار البحــرين"),
    ("OMR", "ريـــال عمـــانى"),
    ("QAR", "ريـــال قطــــرى"),
    ("JOD", "دينار اردنى"),
    ("CNY", "يوان صينى"),
]


def build_exchange_rates_sheet(ws, rates_list, balance_entries):
    FILL_BLACK = _fill("FF000000")
    for c, h in enumerate(["العملة", "شراء", "بيع"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, size=12, name="Arial", color=WHITE)
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    rate_map = {r.currency_code: r for r in rates_list}
    eur_row = 0

    for i, (code, arabic) in enumerate(CURRENCIES, 2):
        _apply_zebra_striping(ws, i, 3)
        r = rate_map.get(code)
        ws.cell(row=i, column=1, value=arabic)
        if r:
            val_buy = (
                round(float(r.buy_rate) * 100, 4)
                if code == "JPY"
                else float(r.buy_rate)
            )
            val_sell = (
                round(float(r.sell_rate) * 100, 4)
                if code == "JPY"
                else float(r.sell_rate)
            )
            ws.cell(row=i, column=2, value=val_buy)
            ws.cell(row=i, column=3, value=val_sell)

        if code == "EUR":
            eur_row = i

    from core.models import Currency

    try:
        usd_cur = Currency.objects.get(code="USD")
        eur_cur = Currency.objects.get(code="EUR")
        home_usd = sum(
            float(be.amount)
            for be in balance_entries
            if be.bank_id is None and be.currency_id == usd_cur.id
        )
        home_eur = sum(
            float(be.amount)
            for be in balance_entries
            if be.bank_id is None and be.currency_id == eur_cur.id
        )
    except Exception:
        home_usd = home_eur = 0

    def _style_side(r, c, val, fmt=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = _f(name="Arial")
        cell.alignment = _center()
        cell.border = _thin()
        if fmt:
            cell.number_format = fmt
        return cell

    _style_side(5, 6, home_eur, FMT_EUR)
    _style_side(5, 7, home_usd, FMT_USD)
    _style_side(5, 8, "Total")

    _style_side(6, 6, f"=F5*B{eur_row}", FMT_EGP_RED)
    _style_side(6, 7, "=G5*B2", FMT_EGP_RED)
    _style_side(6, 8, "=F6+G6", FMT_EGP_RED)
