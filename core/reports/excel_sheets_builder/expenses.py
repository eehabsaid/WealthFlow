# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
"""NOTE: Part of the excel_sheets_builder package split (files >200 lines are
split into packages per WealthFlow refactoring convention). This file holds
the "Expenses" sheet builder only.
"""
from itertools import groupby
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from core.reports.excel_formatting_helpers import (
    FMT_EGP_CERT,
    FMT_DATE,
    WHITE,
    EXP_BG,
    EXP_MONTH_BG,
    EXP_YEAR_BG,
    MONTH_ORDER,
    _f,
    _fill,
    _center,
)


def build_expenses_sheet(ws, expenses_qs):
    hdrs = [
        "Date",
        "Year",
        "Month",
        "Category",
        "Sub-Category",
        "Description",
        "Amount",
        "Currency",
        "Payment Method",
        "Notes",
    ]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial")
        cell.fill = _fill(EXP_BG)
        cell.alignment = _center()
    widths = [14, 8, 12, 18, 20, 35, 14, 10, 16, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    expenses = list(
        expenses_qs.select_related("category", "subcategory", "currency").order_by(
            "year", "month", "date"
        )
    )
    row = 2
    year_total_rows = {}

    for year, yg in groupby(expenses, key=lambda e: e.year):
        year_entries = list(yg)
        year_start = row
        for month, mg in groupby(year_entries, key=lambda e: e.month):
            month_entries = list(mg)
            mname = MONTH_ORDER[month - 1] if 1 <= month <= 12 else str(month)
            mstart = row
            for exp in month_entries:
                ws.cell(row=row, column=1, value=exp.date).number_format = FMT_DATE
                ws.cell(row=row, column=2, value=exp.year)
                ws.cell(row=row, column=3, value=mname)
                ws.cell(
                    row=row, column=4, value=exp.category.name if exp.category else ""
                )
                ws.cell(
                    row=row,
                    column=5,
                    value=exp.subcategory.name if exp.subcategory else "",
                )
                ws.cell(row=row, column=6, value=exp.description or "")
                ws.cell(row=row, column=7, value=float(exp.amount)).number_format = (
                    FMT_EGP_CERT
                )
                ws.cell(
                    row=row,
                    column=8,
                    value=exp.currency.code if exp.currency else "EGP",
                )
                ws.cell(row=row, column=9, value=exp.payment_method or "")
                ws.cell(row=row, column=10, value=exp.notes or "")
                row += 1
            mend = row - 1
            ws.cell(row=row, column=3, value=f"{mname} Total")
            ws.cell(row=row, column=7, value=f"=SUM(G{mstart}:G{mend})")
            for c in range(1, 10 + 1):
                ws.cell(row=row, column=c).font = _f(bold=True, name="Arial")
                ws.cell(row=row, column=c).fill = _fill(EXP_MONTH_BG)
            row += 1
        year_end = row - 1
        ws.cell(row=row, column=2, value=f"{year} Total")
        ws.cell(
            row=row,
            column=7,
            value=f"=SUMIF(B{year_start}:B{year_end},{year},G{year_start}:G{year_end})",
        )
        for c in range(1, 10 + 1):
            ws.cell(row=row, column=c).font = _f(bold=True, name="Arial")
            ws.cell(row=row, column=c).fill = _fill(EXP_YEAR_BG)
        year_total_rows[year] = row
        row += 2

    if year_total_rows:
        grand = "+".join(f"G{r}" for r in year_total_rows.values())
        ws.cell(row=row, column=1, value="Grand Total")
        ws.cell(row=row, column=7, value=f"={grand}")
        for c in range(1, 10 + 1):
            ws.cell(row=row, column=c).font = Font(bold=True, color=WHITE, name="Arial")
            ws.cell(row=row, column=c).fill = _fill(EXP_BG)
