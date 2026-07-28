# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
from itertools import groupby
from openpyxl.styles import Font, Border, Side
from core.reports.excel_formatting_helpers import (
    FMT_EGP,
    _thin,
    _fill,
    _align,
    _center,
    _msort,
)

SALARY_COL_WIDTHS = {
    "NTG": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
    "Giza Systems": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
    "Giza Systems (2)": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
    "ElSeweedy Technology": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
    "Dedalus": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
    "Globemed": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
    "Giza Systems (3)": {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3},
}

YEAR_ROW_HT = 22.8

SALARY_FREEZE = {
    "NTG": "A4",
    "Giza Systems": "A4",
    "Giza Systems (2)": "A4",
    "ElSeweedy Technology": "A4",
    "Dedalus": "A4",
    "Globemed": "A4",
    "Giza Systems (3)": "A4",
}


def _apply_data_row(ws, row, has_bonus=False):
    cols = 6 if has_bonus else 5
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Arial")
        cell.border = _thin()
        if c in (3, 4, 5):
            cell.number_format = FMT_EGP
        if has_bonus and c == 6:
            cell.number_format = FMT_EGP


def _apply_total_row(ws, row, has_bonus=False):
    cols = 6 if has_bonus else 5
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, name="Arial")
        cell.border = _thin()
        cell.alignment = _center()
        if c in (3, 4):
            cell.number_format = FMT_EGP
        if c == 5:
            cell.number_format = FMT_EGP
        if has_bonus and c == 6:
            cell.number_format = FMT_EGP


def build_salary_sheet(ws, company, entries):
    """Exact replica of original salary sheet styling using theme RGB values."""
    name = company.name
    has_bonus = True
    cols = 6
    last_col = "F"

    _fill("FFFFFFFF")
    FILL_DARK_BLUE = _fill("FF1F497D")
    FILL_RED_DATA = _fill("FFC0504D")
    FILL_BLACK = _fill("FF000000")

    GREY = "FF7F7F7F"
    RED_TTL = "FFFF0000"

    ws.row_dimensions[1].height = 14.25
    ws.row_dimensions[2].height = 14.25
    ws.row_dimensions[3].height = 20.25

    widths = SALARY_COL_WIDTHS.get(
        name, {"A": 13.7, "B": 16.1, "C": 16.0, "D": 19.3, "E": 15.6, "F": 14.3}
    )
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    fp = SALARY_FREEZE.get(name)
    if fp:
        ws.freeze_panes = fp

    hdrs = [
        "Year",
        "Month",
        "Expected",
        "Paid (Salary + Bonus)" if has_bonus else "Paid",
        "Remaining",
    ]
    if has_bonus:
        hdrs.append("Bonus")

    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        if has_bonus:
            cell.font = Font(bold=True, italic=True, size=11, name="Arial")
        else:
            cell.font = Font(bold=False, italic=True, size=11, name="Arial", color=GREY)
        cell.fill = FILL_DARK_BLUE
        cell.alignment = _align("center")
        cell.border = _thin()

    ws.merge_cells(f"A2:{last_col}3")
    c2 = ws.cell(row=2, column=1, value=" Salary Details")
    c2.font = Font(bold=True, size=18, name="Times New Roman", color=RED_TTL)
    c2.fill = FILL_BLACK
    c2.alignment = _align("center")
    if has_bonus:
        c2.border = Border(top=Side(style="thin"))
    else:
        c2.border = Border(bottom=Side(style="thin"))

    row = 4
    total_rows = []
    sorted_entries = sorted(entries, key=lambda e: (e.year, _msort(str(e.month))))

    for year, ygrp in groupby(sorted_entries, key=lambda e: e.year):
        year_entries = list(ygrp)

        yr_row = row
        ws.row_dimensions[yr_row].height = YEAR_ROW_HT
        yr_merge = f"A{yr_row}:{last_col}{yr_row}"
        try:
            ws.merge_cells(yr_merge)
        except Exception:
            pass
        yc = ws.cell(row=yr_row, column=1, value=int(year))
        yc.font = Font(bold=True, size=18, name="Times New Roman")
        yc.fill = FILL_DARK_BLUE
        yc.alignment = _align("center")
        if has_bonus:
            yc.border = Border()
        else:
            yc.border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
        row += 1

        data_start = row
        for entry in year_entries:
            for c in range(1, cols + 1):
                ws.cell(row=row, column=c).fill = FILL_RED_DATA
                ws.cell(row=row, column=c).font = Font(size=11, name="Arial")
                ws.cell(row=row, column=c).border = _thin()

            ws.cell(row=row, column=1, value=int(entry.year))
            ws.cell(row=row, column=2, value=str(entry.month))
            ws.cell(row=row, column=3, value=float(entry.expected))
            ws.cell(row=row, column=3).number_format = FMT_EGP
            ws.cell(row=row, column=4, value=float(entry.paid))
            ws.cell(row=row, column=4).number_format = FMT_EGP

            if name in ("NTG", "Giza Systems", "Giza Systems (2)"):
                rem = f"=D{row}-C{row}"
            else:
                rem = f"=IF(C{row}>D{row},C{row}-D{row},0)"
            ws.cell(row=row, column=5, value=rem)
            ws.cell(row=row, column=5).number_format = FMT_EGP

            if has_bonus:
                bonus_val = float(getattr(entry, "bonus", 0) or 0)
                ws.cell(row=row, column=6, value=bonus_val)
                ws.cell(row=row, column=6).number_format = FMT_EGP
            row += 1

        data_end = row - 1

        paid_count = sum(1 for e in year_entries if float(e.paid) > 0)

        for c in range(1, cols + 1):
            tc = ws.cell(row=row, column=c)
            tc.fill = FILL_DARK_BLUE
            tc.font = Font(bold=True, size=11, name="Arial")
            tc.alignment = _align("center")
            tc.border = _thin()

        ws.cell(row=row, column=1, value="Total")
        if name == "NTG" and len(total_rows) == 0:
            ws.cell(row=row, column=2, value=paid_count)
        else:
            ws.cell(
                row=row,
                column=2,
                value=f'=COUNTIF(D{data_start}:D{data_end}, "<> 0.00")',
            )
        ws.cell(row=row, column=3, value=f"=SUM(C{data_start}:C{data_end})")
        ws.cell(row=row, column=3).number_format = FMT_EGP
        ws.cell(row=row, column=4, value=f"=SUM(D{data_start}:D{data_end})")
        ws.cell(row=row, column=4).number_format = FMT_EGP

        if name == "NTG":
            ws.cell(row=row, column=5, value=f"=D{row}-C{row}")
        else:
            ws.cell(row=row, column=5, value=f"=SUM(E{data_start}:E{data_end})")
        ws.cell(row=row, column=5).number_format = FMT_EGP

        if has_bonus:
            ws.cell(row=row, column=6, value=f"=SUM(F{data_start}:F{data_end})")
            ws.cell(row=row, column=6).number_format = FMT_EGP

        total_rows.append(row)
        row += 1

    sr = row

    label_map = {
        "NTG": "Total",
        "Giza Systems": "Total",
        "Giza Systems (2)": "Total",
        "Giza Systems (3)": "Total",
        "ElSeweedy Technology": "Total",
        "Dedalus": "Total",
        "Globemed": "Total",
    }
    label = label_map.get(name, "Total")

    for c in range(1, cols + 1):
        sc = ws.cell(row=sr, column=c)
        sc.fill = FILL_BLACK
        sc.font = Font(bold=True, size=11, name="Arial", color="FFFFFFFF")
        sc.border = _thin()

    ws.cell(row=sr, column=1, value=label)
    b_ref = "+".join(f"B{r}" for r in total_rows)
    ws.cell(row=sr, column=2, value=f"={b_ref}")
    ws.cell(row=sr, column=2).number_format = "0" if name != "NTG" else "General"
    c_ref = "+".join(f"C{r}" for r in total_rows)
    d_ref = "+".join(f"D{r}" for r in total_rows)
    ws.cell(row=sr, column=3, value=f"={c_ref}")
    ws.cell(row=sr, column=3).number_format = FMT_EGP
    ws.cell(row=sr, column=4, value=f"={d_ref}")
    ws.cell(row=sr, column=4).number_format = FMT_EGP
    ws.cell(row=sr, column=5, value=f"=D{sr}-C{sr}")
    ws.cell(row=sr, column=5).number_format = FMT_EGP

    if has_bonus:
        f_ref = "+".join(f"F{r}" for r in total_rows)
        ws.cell(row=sr, column=6, value=f"={f_ref}")
        ws.cell(row=sr, column=6).number_format = FMT_EGP

    ws.row_dimensions[sr].height = 21.0 if name == "NTG" else None

    return sr
