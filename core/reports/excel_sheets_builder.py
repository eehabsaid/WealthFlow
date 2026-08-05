# pyright: reportMissingTypeStubs=false, reportPrivateUsage=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownLambdaType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportMissingParameterType=false, reportIncompatibleMethodOverride=false, reportOptionalMemberAccess=false, reportRedeclaration=false, reportAssignmentType=false
from itertools import groupby
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.reports.excel_formatting_helpers import (
    FMT_EGP_RED,
    FMT_USD,
    FMT_EUR,
    FMT_SAR,
    FMT_GOLD,
    FMT_EGP_CERT,
    FMT_EGP_CERT_R,
    FMT_PCT,
    FMT_DATE,
    FMT_INT,
    WHITE,
    EXP_BG,
    EXP_MONTH_BG,
    EXP_YEAR_BG,
    MONTH_ORDER,
    _f,
    _thin,
    _thin_lr,
    _fill,
    _center,
    _apply_zebra_striping,
)

CURRENCIES = [
    ("USD", "دولار أمريكى"),
    ("EUR", "يورو"),
    ("GBP", "جنيَــه إسترليـنى"),
    ("CAD", "دولار كنـدى"),
    ("DKK", "كرون دانمركى"),
    ("NOK", "كرون نرويجى"),
    ("SEK", "كرون سَــويدى"),
    ("CHF", "فرنك سويسرى"),
    ("JPY", "100 ين يابانى"),
    ("SAR", "ريـــال سعــودى"),
    ("KWD", "دينــار كويتى"),
    ("AED", "درهم اماراتى"),
    ("AUD", "دولار اســـترالى"),
    ("BHD", "دينــار البحــرين"),
    ("OMR", "ريـــال عمـــانى"),
    ("QAR", "ريـــال قطــــرى"),
    ("JOD", "دينار اردنى"),
    ("CNY", "يوان صينى"),
]

def build_exchange_rates_sheet(ws, rates_list, balance_entries):
    FILL_BLACK = _fill("FF000000")
    for c, h in enumerate(["العملة", "شراء", "بيع"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, size=12, name="Arial", color=WHITE)
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    rate_map = {r.currency_code: r for r in rates_list}
    eur_row = 0

    for i, (code, arabic) in enumerate(CURRENCIES, 2):
        _apply_zebra_striping(ws, i, 3)
        r = rate_map.get(code)
        ws.cell(row=i, column=1, value=arabic)
        if r:
            val_buy = (
                round(float(r.buy_rate) * 100, 4)
                if code == "JPY"
                else float(r.buy_rate)
            )
            val_sell = (
                round(float(r.sell_rate) * 100, 4)
                if code == "JPY"
                else float(r.sell_rate)
            )
            ws.cell(row=i, column=2, value=val_buy)
            ws.cell(row=i, column=3, value=val_sell)

        if code == "EUR":
            eur_row = i

    from core.models import Currency

    try:
        usd_cur = Currency.objects.get(code="USD")
        eur_cur = Currency.objects.get(code="EUR")
        home_usd = sum(
            float(be.amount)
            for be in balance_entries
            if be.bank_id is None and be.currency_id == usd_cur.id
        )
        home_eur = sum(
            float(be.amount)
            for be in balance_entries
            if be.bank_id is None and be.currency_id == eur_cur.id
        )
    except Exception:
        home_usd = home_eur = 0

    def _style_side(r, c, val, fmt=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = _f(name="Arial")
        cell.alignment = _center()
        cell.border = _thin()
        if fmt:
            cell.number_format = fmt
        return cell

    _style_side(5, 6, home_eur, FMT_EUR)
    _style_side(5, 7, home_usd, FMT_USD)
    _style_side(5, 8, "Total")

    _style_side(6, 6, f"=F5*B{eur_row}", FMT_EGP_RED)
    _style_side(6, 7, "=G5*B2", FMT_EGP_RED)
    _style_side(6, 8, "=F6+G6", FMT_EGP_RED)


def build_gold_price_sheet(ws, gold_qs, balance_entries):
    ws.column_dimensions["A"].width = 18.6
    ws.column_dimensions["B"].width = 10.7
    ws.column_dimensions["C"].width = 10.7
    ws.column_dimensions["D"].width = 10.7
    ws.column_dimensions["E"].width = 10.7
    ws.column_dimensions["F"].width = 32.7
    ws.column_dimensions["G"].width = 15.8
    ws.column_dimensions["H"].width = 14.3

    FILL_BLACK = _fill("FF000000")
    latest = gold_qs.order_by("-fetched_at").first()

    for c, h in enumerate(["السعر", "شراء", "بيع", "المزيد", "الملاحظات"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial")
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    carats = [
        ("جرام عيار 24", "carat_24k_buy", "carat_24k", "40 ج"),
        ("جرام عيار 22", "carat_22k_buy", "carat_22k", "37 ج"),
        ("جرام عيار 21", "carat_21k_buy", "carat_21k", "35 ج"),
        ("جرام عيار 18", "carat_18k_buy", "carat_18k", "30 ج"),
        ("جرام عيار 14 27 ج", "carat_18k_buy", "carat_18k", "27 ج"),
    ]
    for i, (label, bf, sf, karat) in enumerate(carats, 2):
        _apply_zebra_striping(ws, i, 5)
        for col, val in enumerate(
            [
                label,
                round(float(getattr(latest, bf, 0)), 0) if latest else 0,
                round(float(getattr(latest, sf, 0)), 0) if latest else 0,
                ">",
                karat,
            ],
            1,
        ):
            c = ws.cell(row=i, column=col, value=val)
            c.font = _f(name="Arial")
            c.alignment = _center()
            c.border = _thin()

    rows_data = [
        [
            "الدولار 0 ج",
            float(latest.usd_to_egp) if latest else 0,
            float(latest.usd_to_egp) if latest else 0,
            None,
            "0 ج",
        ],
        [
            "الأونصة 0 $",
            float(latest.usd_per_oz) if latest else 0,
            float(latest.usd_per_oz) if latest else 0,
            None,
            "0 $",
        ],
        [
            "الجنيه الذهب 320 ج",
            round(float(latest.carat_21k_buy) * 8, 0) if latest else 0,
            round(float(latest.carat_21k) * 8, 0) if latest else 0,
            None,
            "320 ج",
        ],
    ]
    for i, row_data in enumerate(rows_data, 7):
        _apply_zebra_striping(ws, i, 5)
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = _f(name="Arial")
            c.alignment = _center()
            c.border = _thin()

    from core.models import Currency

    grams = 0
    try:
        gold_cur = Currency.objects.get(code="Gold")
        grams = sum(
            float(be.amount)
            for be in balance_entries
            if be.bank_id is None and be.currency_id == gold_cur.id
        )
    except:
        pass

    ws.merge_cells("G9:I9")
    c_title = ws.cell(row=9, column=7, value=f"{int(grams)} Grams")
    c_title.font = _f(name="Arial", bold=True)
    c_title.alignment = _center()
    c_title.border = _thin()
    ws.cell(row=9, column=8).border = _thin()
    ws.cell(row=9, column=9).border = _thin()

    for c, title in enumerate(["Now", "Paid", "Diff"], 7):
        c_head = ws.cell(row=10, column=c, value=title)
        c_head.font = _f(name="Arial", bold=True)
        c_head.alignment = _center()
        c_head.border = _thin()

    vals = ["=(C2+28.5)*(BALANCE!F2)", 897375, "=G11-H11"]
    for c, val in enumerate(vals, 7):
        c_val = ws.cell(row=11, column=c, value=val)
        c_val.font = _f(name="Arial")
        c_val.alignment = _center()
        c_val.border = _thin()
        c_val.number_format = FMT_EGP_RED


def build_bank_certificates_sheet(ws, certs_qs):
    FILL_BLACK = _fill("FF000000")
    hdrs = [
        "Amount",
        "Interest Rate",
        "Interest Value",
        "Frequency",
        "Start Date",
        "End Date",
    ]

    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial")
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    for i, cert in enumerate(certs_qs.filter(status__iexact="active").order_by("issue_date"), 2):
        _apply_zebra_striping(ws, i, 6)
        ws.cell(row=i, column=1, value=float(cert.amount)).number_format = FMT_EGP_CERT
        ws.cell(row=i, column=2, value=float(cert.interest_rate) / 100).number_format = (
            FMT_PCT
        )
        ws.cell(row=i, column=3, value=f"=(A{i}*B{i})/12").number_format = (
            FMT_EGP_CERT_R
        )
        ws.cell(row=i, column=4, value=cert.frequency)
        ws.cell(row=i, column=5, value=cert.issue_date).number_format = FMT_DATE
        ws.cell(row=i, column=6, value=cert.expiry_date).number_format = FMT_DATE


def build_balance_sheet(ws, balance_entries, company_sheet_rows):

    ws.column_dimensions["A"].width = 26.5
    ws.column_dimensions["B"].width = 14.6
    ws.column_dimensions["C"].width = 12.5
    ws.column_dimensions["D"].width = 13.1
    ws.column_dimensions["E"].width = 13.1
    ws.column_dimensions["F"].width = 14.4
    ws.column_dimensions["G"].width = 14.4
    ws.column_dimensions["H"].width = 17.1
    ws.column_dimensions["I"].width = 12.7
    ws.column_dimensions["J"].width = 12.0
    ws.column_dimensions["K"].width = 29.7
    ws.row_dimensions[7].height = 18.0

    hdrs = [
        "Title",
        "EGP",
        "USD",
        "EUR",
        "SAR",
        "Gold",
        "Acct-Number",
        "Card-ID",
        "Swift-Code",
        "Customer-id",
        "Customer-Name",
    ]
    border_map = {
        1: _thin(),
        2: _thin(),
        3: _thin(),
        4: _thin_lr(),
        5: _thin_lr(),
        6: _thin_lr(),
        7: _thin_lr(),
        8: _thin_lr(),
        9: _thin(),
        10: _thin(),
        11: _thin(),
    }
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _f(bold=True, name="Arial")
        cell.border = border_map.get(c, _thin())

    from core.models import Currency, Bank as BankModel

    cur_map = {c.id: c.code for c in Currency.objects.all()}
    bank_map = {b.id: b for b in BankModel.objects.all()}

    home_cash_entries = [be for be in balance_entries if be.bank_id is None and str(be.balance_type).strip().lower() == "cash"]
    gold_entries = [be for be in balance_entries if str(be.balance_type).strip().lower() == "gold"]

    home = {cur_map.get(be.currency_id, "?"): float(be.amount) for be in home_cash_entries}
    if gold_entries:
        home["Gold"] = float(gold_entries[0].amount)

    home_title = home_cash_entries[0].title if home_cash_entries else "Home Balance"
    ws.cell(row=2, column=1, value=home_title).font = _f(bold=True, name="Arial")
    ws.cell(row=2, column=1).border = _thin()

    b2 = ws.cell(row=2, column=2, value=home.get("EGP", 0))
    b2.font = _f(bold=True, name="Arial")
    b2.border = _thin()
    b2.number_format = FMT_EGP_RED

    c2 = ws.cell(row=2, column=3, value=home.get("USD", 0))
    c2.font = _f(bold=True, name="Arial")
    c2.border = _thin()
    c2.number_format = FMT_USD

    d2 = ws.cell(row=2, column=4, value=home.get("EUR", 0))
    d2.font = _f(bold=True, name="Arial")
    d2.border = _thin()
    d2.number_format = FMT_EUR

    e2 = ws.cell(row=2, column=5, value=home.get("SAR", 0))
    e2.font = _f(bold=True, name="Arial")
    e2.border = _thin()
    e2.number_format = FMT_SAR

    f2 = ws.cell(row=2, column=6, value=home.get("Gold", 0))
    f2.font = _f(bold=True, name="Arial")
    f2.border = _thin()
    f2.number_format = FMT_GOLD

    excel_row = 3
    for be in sorted(balance_entries, key=lambda b: b.id):
        if be.bank_id is None or str(be.balance_type).strip().lower() in ("certificate", "gold"):
            continue
        if cur_map.get(be.currency_id) != "EGP":
            continue

        a = ws.cell(row=excel_row, column=1, value=be.title)
        a.font = _f(bold=True, name="Arial")
        a.border = _thin()
        b = ws.cell(row=excel_row, column=2, value=float(be.amount))
        b.font = _f(bold=True, name="Arial")
        b.border = _thin()
        b.number_format = FMT_EGP_RED

        bank = bank_map.get(be.bank_id)
        if bank:
            for col, attr in [
                (7, "account_number"),
                (8, "card_number"),
                (9, "swift_code"),
                (10, "customer_id"),
                (11, "customer_name"),
            ]:
                v = getattr(bank, attr, "") or ""
                cell = ws.cell(row=excel_row, column=col, value=v)
                cell.font = _f(bold=True, name="Arial")
                cell.border = _thin()
                if col in (7, 8):
                    cell.number_format = FMT_INT
        excel_row += 1

    from core.models import BankCertificate

    cert_count = BankCertificate.objects.filter(status__iexact="active").count()
    cr = excel_row
    cert_entries = [be for be in balance_entries if str(be.balance_type).strip().lower() == "certificate"]
    cert_title = cert_entries[0].title if cert_entries else "Certificates Balance"

    ws.cell(row=cr, column=1, value=cert_title).font = _f(bold=True, name="Arial")
    ws.cell(row=cr, column=1).border = _thin()
    bc = ws.cell(
        row=cr, column=2, value=f"=SUM('Bank-Certificates'!A2:A{cert_count+1})"
    )
    bc.font = _f(bold=True, name="Arial")
    bc.border = _thin()
    bc.number_format = FMT_EGP_RED
    excel_row += 1

    ter = excel_row
    ws.cell(row=ter, column=1, value="Total EGP Balance").font = _f(
        bold=True, name="Arial"
    )
    ws.cell(row=ter, column=1).border = _thin()
    te = ws.cell(row=ter, column=2, value=f"=SUM(B2:B{cr})")
    te.font = _f(bold=True, name="Arial")
    te.border = _thin()
    te.number_format = FMT_EGP_RED
    excel_row += 1

    tar = excel_row
    ws.row_dimensions[tar].height = 18.0
    ws.merge_cells(f"B{tar}:F{tar}")

    ws.cell(row=tar, column=1, value="Total all Balances").font = _f(
        bold=True, name="Arial"
    )
    ws.cell(row=tar, column=1).border = _thin()

    formula = (
        f"=B{ter}"
        f"+(C2*('Exchange Rates'!B2))"
        f"+(D2*('Exchange Rates'!B3))"
        f"+(E2*('Exchange Rates'!B11))"
        f"+(F2*(('Gold Price'!C2)+28.5))"
    )

    ta = ws.cell(row=tar, column=2, value=formula)
    ta.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(2, 7):
        ws.cell(row=tar, column=col).border = thin_border

    ta.font = _f(bold=True, name="Arial")
    ta.number_format = FMT_EGP_RED
    excel_row += 1

    tpr = excel_row + 3
    tmr = tpr + 1

    BONUS_COMPANIES = set(company_sheet_rows.keys())

    pay_parts, month_parts = [], []
    for cname, (sname, srow) in company_sheet_rows.items():
        ref = f"'{sname}'!{{c}}{srow}" if " " in sname else f"{sname}!{{c}}{srow}"
        if cname in BONUS_COMPANIES:
            pay_parts.append(f"({ref.format(c='D')}+{ref.format(c='F')})")
        else:
            pay_parts.append(ref.format(c="D"))
        month_parts.append(ref.format(c="B"))

    label_tp = ws.cell(row=tpr, column=1, value="Total Pays")
    label_tp.font = _f(bold=True, name="Arial")
    label_tp.border = _thin()

    tp = ws.cell(row=tpr, column=2, value="=" + "+".join(pay_parts) if pay_parts else 0)
    tp.font = _f(bold=True, name="Arial")
    tp.border = _thin()
    tp.number_format = FMT_EGP_RED

    label_tm = ws.cell(row=tmr, column=1, value="Total Work Months")
    label_tm.font = _f(bold=True, name="Arial")
    label_tm.border = _thin()

    tm = ws.cell(
        row=tmr, column=2, value="=" + "+".join(month_parts) if month_parts else 0
    )
    tm.font = _f(bold=True, name="Arial")
    tm.border = _thin()


def build_expenses_sheet(ws, expenses_qs):
    hdrs = [
        "Date",
        "Year",
        "Month",
        "Category",
        "Sub-Category",
        "Description",
        "Amount",
        "Currency",
        "Payment Method",
        "Notes",
    ]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial")
        cell.fill = _fill(EXP_BG)
        cell.alignment = _center()
    widths = [14, 8, 12, 18, 20, 35, 14, 10, 16, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    expenses = list(
        expenses_qs.select_related("category", "subcategory", "currency").order_by(
            "year", "month", "date"
        )
    )
    row = 2
    year_total_rows = {}

    for year, yg in groupby(expenses, key=lambda e: e.year):
        year_entries = list(yg)
        year_start = row
        for month, mg in groupby(year_entries, key=lambda e: e.month):
            month_entries = list(mg)
            mname = MONTH_ORDER[month - 1] if 1 <= month <= 12 else str(month)
            mstart = row
            for exp in month_entries:
                ws.cell(row=row, column=1, value=exp.date).number_format = FMT_DATE
                ws.cell(row=row, column=2, value=exp.year)
                ws.cell(row=row, column=3, value=mname)
                ws.cell(
                    row=row, column=4, value=exp.category.name if exp.category else ""
                )
                ws.cell(
                    row=row,
                    column=5,
                    value=exp.subcategory.name if exp.subcategory else "",
                )
                ws.cell(row=row, column=6, value=exp.description or "")
                ws.cell(row=row, column=7, value=float(exp.amount)).number_format = (
                    FMT_EGP_CERT
                )
                ws.cell(
                    row=row,
                    column=8,
                    value=exp.currency.code if exp.currency else "EGP",
                )
                ws.cell(row=row, column=9, value=exp.payment_method or "")
                ws.cell(row=row, column=10, value=exp.notes or "")
                row += 1
            mend = row - 1
            ws.cell(row=row, column=3, value=f"{mname} Total")
            ws.cell(row=row, column=7, value=f"=SUM(G{mstart}:G{mend})")
            for c in range(1, 10 + 1):
                ws.cell(row=row, column=c).font = _f(bold=True, name="Arial")
                ws.cell(row=row, column=c).fill = _fill(EXP_MONTH_BG)
            row += 1
        year_end = row - 1
        ws.cell(row=row, column=2, value=f"{year} Total")
        ws.cell(
            row=row,
            column=7,
            value=f"=SUMIF(B{year_start}:B{year_end},{year},G{year_start}:G{year_end})",
        )
        for c in range(1, 10 + 1):
            ws.cell(row=row, column=c).font = _f(bold=True, name="Arial")
            ws.cell(row=row, column=c).fill = _fill(EXP_YEAR_BG)
        year_total_rows[year] = row
        row += 2

    if year_total_rows:
        grand = "+".join(f"G{r}" for r in year_total_rows.values())
        ws.cell(row=row, column=1, value="Grand Total")
        ws.cell(row=row, column=7, value=f"={grand}")
        for c in range(1, 10 + 1):
            ws.cell(row=row, column=c).font = Font(bold=True, color=WHITE, name="Arial")
            ws.cell(row=row, column=c).fill = _fill(EXP_BG)

def build_currency_exchanges_sheet(ws, exchanges_qs, lang="ar"):
    from core.reports.report_utils import get_translations, get_text

    t_map = get_translations(lang)
    FILL_BLACK = _fill("FF000000")

    header_keys = [
        ("date", "Date"),
        ("from_balance", "Source Balance"),
        ("from_currency", "From Currency"),
        ("from_amount", "Amount Exchanged"),
        ("to_balance", "Destination Balance"),
        ("to_currency", "To Currency"),
        ("to_amount", "Amount Received"),
        ("exchange_rate", "Exchange Rate"),
        ("status", "Status"),
        ("notes", "Notes"),
    ]
    headers = [get_text(key, lang, t_map, default) for key, default in header_keys]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, size=11, name="Arial", color=WHITE)
        cell.fill = FILL_BLACK
        cell.alignment = _center()
        cell.border = _thin()

    exchanges = list(
        exchanges_qs.select_related("from_balance", "from_currency", "to_balance", "to_currency").order_by("-exchange_date", "-id")
    )
    status_keys = {
        "ACTIVE": ("status_active", "Active"),
        "REVERSED": ("status_reversed", "Reversed"),
        "EDITED": ("status_edited", "Edited"),
    }

    for i, ex in enumerate(exchanges, 2):
        _apply_zebra_striping(ws, i, len(headers))
        st_key, st_def = status_keys.get(ex.status, ("status", ex.status))
        st_label = get_text(st_key, lang, t_map, st_def)

        ws.cell(row=i, column=1, value=str(ex.exchange_date))
        ws.cell(row=i, column=2, value=ex.from_balance.title if ex.from_balance else "")
        ws.cell(row=i, column=3, value=ex.from_currency.code if ex.from_currency else "")
        ws.cell(row=i, column=4, value=float(ex.from_amount))
        ws.cell(row=i, column=5, value=ex.to_balance.title if ex.to_balance else "")
        ws.cell(row=i, column=6, value=ex.to_currency.code if ex.to_currency else "")
        ws.cell(row=i, column=7, value=float(ex.to_amount))
        ws.cell(row=i, column=8, value=float(ex.exchange_rate))
        ws.cell(row=i, column=9, value=st_label)
        ws.cell(row=i, column=10, value=ex.notes or "")
