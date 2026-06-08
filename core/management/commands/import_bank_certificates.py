import csv
import os
from datetime import datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models import Bank, Currency, BankCertificate


COLUMN_MAP = {
    'bank': ['bank', 'bank name', 'bank_name'],
    'currency': ['currency', 'currency code', 'currency_name', 'currency_code'],
    'issue_date': ['issue date', 'issued', 'date issued', 'start date'],
    'expiry_date': ['expiry date', 'expiry', 'maturity date', 'expiration date', 'end date'],
    'amount': ['amount', 'value', 'balance', 'certificate amount'],
    'interest_rate': ['interest rate', 'rate'],
    'interest_value': ['interest value', 'interest'],
    'frequency': ['frequency', 'payment frequency'],
    'status': ['status', 'state'],
    'notes': ['notes', 'comment', 'description'],
}


def normalize_header(name):
    if not name:
        return ''
    return str(name).strip().lower().replace('_', ' ').replace('-', ' ')


def map_headers(headers):
    mapped = {}
    normalized = [normalize_header(h) for h in headers]
    for idx, norm in enumerate(normalized):
        for key, aliases in COLUMN_MAP.items():
            if norm in aliases:
                mapped[key] = idx
                break
    return mapped


def parse_date(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except (ValueError, TypeError):
            continue
    raise ValueError(f"Unrecognized date format: {value}")


def parse_decimal(value):
    if value is None or value == '':
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(',', '')
    if text == '':
        return 0
    return float(text)


def find_currency(value):
    if not value:
        return None
    value = str(value).strip()
    if not value:
        return None
    try:
        return Currency.objects.get(code__iexact=value)
    except Currency.DoesNotExist:
        try:
            return Currency.objects.get(name__iexact=value)
        except Currency.DoesNotExist:
            return None


class Command(BaseCommand):
    help = 'Import bank certificates from an Excel or CSV sheet into BankCertificate records.'

    def add_arguments(self, parser):
        parser.add_argument('path', help='Path to the Excel or CSV file to import')
        parser.add_argument('--sheet', default='Bank-Certificates', help='Sheet name to import when using Excel')

    def handle(self, *args, **options):
        path = options['path']
        sheet_name = options['sheet']

        if not os.path.exists(path):
            raise CommandError(f'File not found: {path}')

        ext = os.path.splitext(path)[1].lower()
        rows = []

        if ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            try:
                import openpyxl
            except ImportError as exc:
                raise CommandError('openpyxl is required for Excel imports. Install it with pip install openpyxl.') from exc

            workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f'Sheet not found: {sheet_name}. Available sheets: {workbook.sheetnames}')
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            rows = list(iterator)
        elif ext == '.csv':
            with open(path, newline='', encoding='utf-8-sig') as csvfile:
                reader = csv.reader(csvfile)
                rows = list(reader)
        else:
            raise CommandError('Unsupported file type. Use .xlsx, .xlsm, or .csv')

        if not rows:
            raise CommandError('No rows found in the import file.')

        headers = rows[0]
        mapping = map_headers(headers)

        created = 0
        updated = 0
        skipped = 0

        with transaction.atomic():
            for row in rows[1:]:
                if not any(cell not in (None, '') for cell in row):
                    continue
                bank = None
                if 'bank' in mapping:
                    bank_name = row[mapping['bank']]
                    if bank_name not in (None, ''):
                        bank_name = str(bank_name).strip()
                        bank, _ = Bank.objects.get_or_create(name=bank_name, defaults={'is_active': True})

                currency = None
                if 'currency' in mapping:
                    currency = find_currency(row[mapping['currency']])

                issue_date = None
                expiry_date = None
                if 'issue_date' in mapping:
                    try:
                        issue_date = parse_date(row[mapping['issue_date']])
                    except ValueError as exc:
                        raise CommandError(str(exc))
                if 'expiry_date' in mapping:
                    try:
                        expiry_date = parse_date(row[mapping['expiry_date']])
                    except ValueError as exc:
                        raise CommandError(str(exc))

                amount = 0
                if 'amount' in mapping:
                    amount = parse_decimal(row[mapping['amount']])

                interest_rate = 0
                if 'interest_rate' in mapping:
                    interest_rate = parse_decimal(row[mapping['interest_rate']])

                interest_value = 0
                if 'interest_value' in mapping:
                    interest_value = parse_decimal(row[mapping['interest_value']])

                frequency = str(row[mapping['frequency']]).strip() if 'frequency' in mapping and row[mapping['frequency']] not in (None, '') else ''
                status = str(row[mapping['status']]).strip() if 'status' in mapping and row[mapping['status']] not in (None, '') else 'Active'
                notes = str(row[mapping['notes']]).strip() if 'notes' in mapping and row[mapping['notes']] not in (None, '') else ''

                BankCertificate.objects.create(
                    bank=bank,
                    currency=currency,
                    issue_date=issue_date,
                    expiry_date=expiry_date,
                    amount=amount,
                    interest_rate=interest_rate,
                    interest_value=interest_value,
                    frequency=frequency,
                    status=status,
                    notes=notes,
                )
                created += 1

        self.stdout.write(self.style.SUCCESS(f'Import complete: {created} created, {updated} updated, {skipped} skipped.'))
